import asyncio
import csv
import base64
import datetime
import httpx
import smtplib
import json
import ipaddress
import math
import hashlib
import html
import socket
from collections import deque
import io
import logging
import os
import re
import secrets
import shlex
import shutil
import subprocess
import tempfile
import threading
import zipfile
import time
import ssl
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Optional
from uuid import uuid4
from xml.etree import ElementTree as ET
from email.message import EmailMessage

import html5lib
from fastapi import APIRouter, FastAPI, Depends, HTTPException, Query, Request, Response, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.openapi.docs import get_swagger_ui_html
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, selectinload, joinedload
from pypdf import PdfReader, PdfWriter
from reportlab.lib.colors import Color, HexColor
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from PIL import Image, ImageDraw, ImageFont, ImageFilter, ImageOps

from backend.database import (
    DEFAULT_DATA_DIR,
    SessionLocal,
    get_db,
    init_db,
    allocate_series_tab_color,
)
from backend.models import (
    Product,
    Series,
    SeriesImage,
    RpmLine,
    RpmPoint,
    EfficiencyPoint,
    ProductImage,
    ProductType,
    ProductTypeParameterGroupPreset,
    ProductTypeParameterPreset,
    ProductTypeRpmLinePreset,
    ProductTypeRpmPointPreset,
    ProductTypeEfficiencyPointPreset,
    ProductParameterGroup,
    ProductParameter,
    AssociatedDocument,
    QuoteRequest,
    User,
)
from backend.models import AppSettings
from backend.timezone import APP_TIMEZONE, backend_now, backend_now_iso

PERMISSIBLE_USE_MODES = {"dedicated", "upper", "lower", "both", "none"}


def normalize_permissible_use_mode(value, default="both"):
    normalized = str(value or "").strip().lower()
    return normalized if normalized in PERMISSIBLE_USE_MODES else default
from backend.schemas import (
    BandGraphStyleSettings,
    ProductCreate,
    ProductUpdate,
    ProductGraphDataReplace,
    ProductResponse,
    GraphImageMaintenanceResponse,
    MaintenanceJobResponse,
    PdfMaintenanceResponse,
    RpmLineCreate,
    RpmLineUpdate,
    RpmLineResponse,
    RpmPointCreate,
    RpmPointResponse,
    EfficiencyPointCreate,
    EfficiencyPointResponse,
    AuthSessionResponse,
    AuthPasswordChangeRequest,
    CmsProductResponse,
    CmsSeriesResponse,
    LoginRequest,
    UserCreate,
    UserPasswordUpdate,
    UserResponse,
    UserUpdate,
    ProductImageResponse,
    ProductImageReorder,
    SeriesImageResponse,
    SeriesImageReorder,
    ProductTypeResponse,
    ProductTypeCreate,
    ProductTypeParameterGroupPresetUpdate,
    ProductTypePresetUpdate,
    ProductTypeUpdate,
    ProductTypeParameterPresetUpdate,
    SeriesCreate,
    SeriesResponse,
    SeriesUpdate,
    TemplateCreateRequest,
    TemplateAssetUploadRequest,
    TemplateAssetUploadResponse,
    TemplateFileResponse,
    TemplateFileUpdateRequest,
    TemplateRegistryResponse,
    FileManagerCreateFolderRequest,
    FileManagerContentResponse,
    FileManagerContentUpdateRequest,
    FileManagerDeleteRequest,
    FileManagerEntryResponse,
    FileManagerListingResponse,
    FileManagerRenameRequest,
    BulkImportResponse,
    BulkImportSheetNormalizationResponse,
    BulkImportTableSummaryResponse,
    BulkImportManifestSheetResponse,
    BulkImageImportResponse,
    ProductTypePdfResponse,
    SetupLogEntryResponse,
    PublicAccessLogEntryResponse,
    QuoteRequestNotificationSettings,
    QuoteRequestCreate,
    QuoteRequestResponse,
    QuoteRequestEmailTestRequest,
    QuoteRequestEmailTestResponse,
    QuoteRequestStatusUpdate,
    AssociatedDocumentResponse,
)

SAFE_CHARS_RE = re.compile(r"[^a-z0-9]+")
JINJA_PATTERN = re.compile(r"(\{\{[\s\S]*?\}\}|\{%-?[\s\S]*?-?%\}|\{#.*?#\})")
GRAPH_FILTER_GROUP_NAME = "__graph__"
PRODUCT_IMAGES_DIR = Path(DEFAULT_DATA_DIR) / "product_images"
SERIES_IMAGES_DIR = Path(DEFAULT_DATA_DIR) / "series_images"
PRODUCT_GRAPHS_DIR = Path(DEFAULT_DATA_DIR) / "product_graphs"
IMPORTS_DIR = Path(DEFAULT_DATA_DIR) / "bulk_imports"
PRODUCT_PDFS_DIR = Path(DEFAULT_DATA_DIR) / "product_pdfs"
PRODUCT_TYPE_PDFS_DIR = Path(DEFAULT_DATA_DIR) / "product_type_pdfs"
ALL_PRODUCT_TYPES_PDF_FILE_NAME = "product_type_printed_all.pdf"
SERIES_GRAPHS_DIR = Path(DEFAULT_DATA_DIR) / "series_graphs"
SERIES_PDFS_DIR = Path(DEFAULT_DATA_DIR) / "series_pdfs"
ASSOCIATED_DOCUMENTS_DIR = Path(DEFAULT_DATA_DIR) / "associated_documents"
BACKUP_OUTPUT_DIR = Path(DEFAULT_DATA_DIR) / "backups"
DATA_BACKUP_DIRS = [
    PRODUCT_IMAGES_DIR,
    SERIES_IMAGES_DIR,
    PRODUCT_GRAPHS_DIR,
    PRODUCT_PDFS_DIR,
    PRODUCT_TYPE_PDFS_DIR,
    SERIES_GRAPHS_DIR,
    SERIES_PDFS_DIR,
    ASSOCIATED_DOCUMENTS_DIR,
]
DATA_BACKUP_DIR_NAMES = [path.name for path in DATA_BACKUP_DIRS]
FRONTEND_DIR = Path(__file__).resolve().parents[1] / "frontend"
TEMPLATES_DIR = Path(__file__).resolve().parents[1] / "templates"
TEMPLATE_REGISTRY_PATH = TEMPLATES_DIR / "registry.json"
ECHARTS_RENDER_SCRIPT = FRONTEND_DIR / "scripts" / "render_product_graph.mjs"
PRODUCT_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
SERIES_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
PRODUCT_GRAPHS_DIR.mkdir(parents=True, exist_ok=True)
IMPORTS_DIR.mkdir(parents=True, exist_ok=True)
PRODUCT_PDFS_DIR.mkdir(parents=True, exist_ok=True)
PRODUCT_TYPE_PDFS_DIR.mkdir(parents=True, exist_ok=True)
SERIES_GRAPHS_DIR.mkdir(parents=True, exist_ok=True)
SERIES_PDFS_DIR.mkdir(parents=True, exist_ok=True)
BACKUP_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
logger = logging.getLogger(__name__)
PDF_MEDIA_HEADERS = {
    "X-Content-Type-Options": "nosniff",
    "Cache-Control": "public, max-age=3600",
}
APP_LOG_LEVEL_NAME = os.getenv("LOG_LEVEL", "INFO").strip().upper() or "INFO"
APP_LOG_LEVEL = getattr(logging, APP_LOG_LEVEL_NAME, logging.INFO)
LOG_BUFFER_SIZE = int(os.getenv("SETUP_LOG_BUFFER_SIZE", "500"))
LOG_BUFFER = deque(maxlen=LOG_BUFFER_SIZE)
LOG_BUFFER_LOCK = threading.Lock()
LOG_BUFFER_CONDITION = threading.Condition(LOG_BUFFER_LOCK)
LOG_SEQUENCE = 0
LOG_HANDLER_ATTACHED = False
FINDER_DEBUG = os.getenv("FINDER_DEBUG", "false").strip().lower() in {"1", "true", "yes", "on"}
SESSION_SECRET = os.getenv("SESSION_SECRET", "")
AUTH_COOKIE_SECURE = os.getenv("AUTH_COOKIE_SECURE", "false").strip().lower() in {"1", "true", "yes", "on"}
BOOTSTRAP_ADMIN_USERNAME = os.getenv("BOOTSTRAP_ADMIN_USERNAME", "admin").strip()
BOOTSTRAP_ADMIN_PASSWORD = os.getenv("BOOTSTRAP_ADMIN_PASSWORD", "").strip()
CMS_API_TOKEN = os.getenv("CMS_API_TOKEN", "").strip()
PUBLIC_CATALOGUE_SITE_URL = os.getenv("PUBLIC_CATALOGUE_SITE_URL", "").strip().rstrip("/")
PUBLIC_CATALOGUE_REFRESH_LOCK = threading.Lock()
PUBLIC_CATALOGUE_REFRESH_IN_FLIGHT = False
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
POSTGRES_DB = os.getenv("POSTGRES_DB", "").strip()
POSTGRES_USER = os.getenv("POSTGRES_USER", "").strip()
PASSWORD_HASH_ITERATIONS = 600_000
QUOTE_REQUEST_RECIPIENT_EMAILS = [
    email.strip()
    for email in os.getenv("QUOTE_REQUEST_RECIPIENT_EMAILS", os.getenv("QUOTE_REQUEST_RECIPIENT_EMAIL", "admin@venttech.co.nz")).split(",")
    if email.strip()
]
QUOTE_REQUEST_THROTTLE_WINDOW_SECONDS = int(os.getenv("QUOTE_REQUEST_THROTTLE_WINDOW_SECONDS", "900"))
QUOTE_REQUEST_THROTTLE_MAX_ATTEMPTS = int(os.getenv("QUOTE_REQUEST_THROTTLE_MAX_ATTEMPTS", "3"))
QUOTE_REQUEST_THROTTLE_STATE: dict[str, deque[float]] = {}
QUOTE_REQUEST_THROTTLE_LOCK = threading.Lock()
SMTP_HOST = os.getenv("SMTP_HOST", "").strip()
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME", "").strip()
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "").strip()
SMTP_FROM_EMAIL = os.getenv("SMTP_FROM_EMAIL", "").strip()
SMTP_FROM_NAME = os.getenv("SMTP_FROM_NAME", "Vent-Tech website").strip()
SMTP_USE_TLS = os.getenv("SMTP_USE_TLS", "true").strip().lower() in {"1", "true", "yes", "on"}
SMTP_USE_SSL = os.getenv("SMTP_USE_SSL", "false").strip().lower() in {"1", "true", "yes", "on"}
POSTGRES_CLIENT_IMAGE = os.getenv("PG_CLIENT_IMAGE", "docker.io/library/postgres:16").strip() or "docker.io/library/postgres:16"
MAINTENANCE_JOBS: dict[str, dict] = {}


def is_localhost_database_url(database_url: str) -> bool:
    if not database_url:
        return False
    parsed = urllib.parse.urlparse(database_url)
    hostname = (parsed.hostname or "").lower()
    return hostname in {"localhost", "127.0.0.1", "::1"}


AUTH_COOKIE_SECURE = AUTH_COOKIE_SECURE and not is_localhost_database_url(DATABASE_URL)
MAINTENANCE_JOBS_LOCK = threading.Lock()
bulk_import_router = APIRouter(tags=["Maintenance"])


class InMemoryLogHandler(logging.Handler):
    def emit(self, record: logging.LogRecord):
        global LOG_SEQUENCE

        try:
            timestamp = datetime.datetime.fromtimestamp(record.created, tz=APP_TIMEZONE).isoformat(timespec="seconds")
            message = self.format(record)
            entry = {
                "id": 0,
                "timestamp": timestamp,
                "level": record.levelname,
                "logger": record.name,
                "message": message,
                "formatted": f"{timestamp} [{record.levelname:<5}] {record.name}: {message}",
            }
        except Exception:
            self.handleError(record)
            return

        with LOG_BUFFER_CONDITION:
            LOG_SEQUENCE += 1
            entry["id"] = LOG_SEQUENCE
            LOG_BUFFER.append(entry)
            LOG_BUFFER_CONDITION.notify_all()


class NonUvicornLogFilter(logging.Filter):
    def filter(self, record):
        return not str(record.name or "").startswith("uvicorn")


class SuppressPublicAccessStreamFilter(logging.Filter):
    def filter(self, record: logging.LogRecord) -> bool:
        return not str(record.getMessage() or "").startswith(PUBLIC_ACCESS_LOG_PREFIX)


def attach_in_memory_log_handler():
    global LOG_HANDLER_ATTACHED
    if LOG_HANDLER_ATTACHED:
        return

    handler = InMemoryLogHandler()
    handler.setLevel(APP_LOG_LEVEL)
    handler.setFormatter(logging.Formatter("%(message)s"))

    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(APP_LOG_LEVEL)
    stream_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s"))
    stream_handler.addFilter(NonUvicornLogFilter())
    stream_handler.addFilter(SuppressPublicAccessStreamFilter())

    root_logger = logging.getLogger()
    root_logger.setLevel(APP_LOG_LEVEL)
    logging.getLogger("httpx").setLevel(logging.WARNING)
    logging.getLogger("httpcore").setLevel(logging.WARNING)
    logging.getLogger("PIL").setLevel(logging.WARNING)
    logging.getLogger("PIL.PngImagePlugin").setLevel(logging.WARNING)
    root_logger.addHandler(handler)
    root_logger.addHandler(stream_handler)
    for name in ("uvicorn", "uvicorn.error", "uvicorn.access"):
        logging.getLogger(name).addHandler(handler)

    logging.captureWarnings(True)
    LOG_HANDLER_ATTACHED = True


attach_in_memory_log_handler()


def _strip_ip_port(value: str) -> str:
    candidate = (value or "").strip().strip('"').strip("'")
    if not candidate:
        return ""
    if candidate.startswith("[") and "]" in candidate:
        return candidate[1 : candidate.index("]")]
    if candidate.count(":") == 1:
        host, port = candidate.rsplit(":", 1)
        if port.isdigit():
            return host
    return candidate


def _parse_ip_address(value: str) -> ipaddress._BaseAddress | None:
    candidate = _strip_ip_port(value)
    if not candidate or candidate.lower() == "unknown":
        return None
    try:
        return ipaddress.ip_address(candidate)
    except ValueError:
        return None


def _iter_forwarded_ips(request: Request):
    forwarded = request.headers.get("forwarded", "")
    if forwarded:
        for entry in forwarded.split(","):
            for part in entry.split(";"):
                key, separator, raw_value = part.partition("=")
                if not separator or key.strip().lower() != "for":
                    continue
                yield raw_value.strip()
                break

    x_forwarded_for = request.headers.get("x-forwarded-for", "")
    if x_forwarded_for:
        for value in x_forwarded_for.split(","):
            yield value.strip()


def _extract_request_ip_details(request: Request) -> dict[str, str | int | None]:
    peer_host = request.client.host if request.client else ""
    peer_port = request.client.port if request.client else None
    forwarded_for = request.headers.get("forwarded", "")
    x_forwarded_for = request.headers.get("x-forwarded-for", "")
    x_real_ip = request.headers.get("x-real-ip", "")
    public_host = ""
    public_port = None
    public_source = ""
    public_ipv4 = ""
    public_ipv6 = ""
    forwarded_chain: list[str] = []

    candidates: list[tuple[str, str]] = []
    if forwarded_for:
        for entry in forwarded_for.split(","):
            for part in entry.split(";"):
                key, separator, raw_value = part.partition("=")
                if not separator or key.strip().lower() != "for":
                    continue
                cleaned = raw_value.strip()
                if cleaned:
                    candidates.append(("forwarded", cleaned))
                    forwarded_chain.append(cleaned)
                break

    if x_forwarded_for:
        for value in x_forwarded_for.split(","):
            cleaned = value.strip()
            if cleaned:
                candidates.append(("x-forwarded-for", cleaned))
                forwarded_chain.append(cleaned)

    if x_real_ip:
        cleaned = x_real_ip.strip()
        if cleaned:
            candidates.append(("x-real-ip", cleaned))
            forwarded_chain.append(cleaned)

    if peer_host:
        candidates.append(("peer", peer_host))

    for source, candidate in candidates:
        host, port = _split_host_port(candidate)
        parsed = _parse_ip_address(host)
        if parsed is None:
            continue
        if parsed.version == 4 and not public_ipv4:
            public_ipv4 = str(parsed)
        elif parsed.version == 6 and not public_ipv6:
            public_ipv6 = str(parsed)
        if not public_host:
            public_host = str(parsed)
            public_port = port if source != "peer" else peer_port
            public_source = source

    if not public_host:
        peer = _parse_ip_address(peer_host)
        if peer is not None:
            public_host = str(peer)
            public_port = peer_port
            public_source = "peer"
            if peer.version == 4 and not public_ipv4:
                public_ipv4 = str(peer)
            elif peer.version == 6 and not public_ipv6:
                public_ipv6 = str(peer)

    return {
        "peer_host": peer_host,
        "peer_port": peer_port,
        "public_host": public_host,
        "public_port": public_port,
        "public_source": public_source,
        "public_ipv4": public_ipv4,
        "public_ipv6": public_ipv6,
        "forwarded_chain": forwarded_chain,
        "forwarded_for": forwarded_for,
        "x_forwarded_for": x_forwarded_for,
        "x_real_ip": x_real_ip,
    }


def _extract_public_client_ip(request: Request) -> str | None:
    details = _extract_request_ip_details(request)
    if details["public_host"]:
        return str(details["public_host"])
    return None


def _detect_local_device_ip(family: int) -> str | None:
    targets = {
        socket.AF_INET: ("8.8.8.8", 80),
        socket.AF_INET6: ("2001:4860:4860::8888", 80, 0, 0),
    }
    target = targets.get(family)
    if not target:
        return None

    sock = socket.socket(family, socket.SOCK_DGRAM)
    try:
        sock.connect(target)
        candidate = sock.getsockname()[0]
        parsed = ipaddress.ip_address(candidate)
        return str(parsed) if parsed.version == (4 if family == socket.AF_INET else 6) else None
    except OSError:
        return None
    except ValueError:
        return None
    finally:
        sock.close()


def _extract_device_ip_details(request: Request) -> dict[str, str | None]:
    details = _extract_request_ip_details(request)
    device_ipv4 = details["public_ipv4"] or None
    device_ipv6 = details["public_ipv6"] or None
    device_ip = details["public_host"] or None
    device_is_loopback = device_ip in {"127.0.0.1", "::1", "localhost", None}

    if device_is_loopback:
        local_ipv4 = _detect_local_device_ip(socket.AF_INET)
        local_ipv6 = _detect_local_device_ip(socket.AF_INET6)
        if local_ipv4:
            device_ipv4 = local_ipv4
            device_ip = local_ipv4
        elif local_ipv6:
            device_ipv6 = local_ipv6
            device_ip = local_ipv6

    return {
        "device_ip_v4": device_ipv4,
        "device_ip_v6": device_ipv6,
        "device_ip": device_ip,
    }


def _request_path_with_query(request: Request) -> str:
    if request.url.query:
        return f"{request.url.path}?{request.url.query}"
    return request.url.path


def _route_group_for_path(path: str) -> str:
    path = path or ""
    if path == "/api/health":
        return "health"
    if path == "/api/client-telemetry":
        return "telemetry"
    if path.startswith("/api/public/"):
        return "public-api"
    if path.startswith("/api/"):
        return "internal-api"
    return "other"


def _split_host_port(value: str) -> tuple[str, int | None]:
    candidate = _strip_ip_port(value)
    if not candidate:
        return "", None
    if candidate.startswith("[") and "]" in candidate:
        candidate = candidate[1 : candidate.index("]")]
    if candidate.count(":") == 1:
        host, port = candidate.rsplit(":", 1)
        if port.isdigit():
            return host, int(port)
    return candidate, None


def _extract_public_client(request: Request) -> dict[str, object]:
    details = _extract_request_ip_details(request)
    username = ""
    try:
        username = str(request.session.get("username") or "")
    except Exception:
        username = ""
    device_type = "desktop"
    user_agent = request.headers.get("user-agent", "")
    sec_ch_ua_mobile = request.headers.get("sec-ch-ua-mobile", "")
    if sec_ch_ua_mobile.strip().lower() in {'?1', '1', 'true'}:
        device_type = "mobile"
    elif "ipad" in user_agent.lower() or "tablet" in user_agent.lower():
        device_type = "tablet"
    elif "mobile" in user_agent.lower():
        device_type = "mobile"

    return {
        **details,
        "host": request.headers.get("host", ""),
        "scheme": request.url.scheme,
        "method": request.method,
        "path": _request_path_with_query(request),
        "user_agent": user_agent,
        "referer": request.headers.get("referer", ""),
        "accept_language": request.headers.get("accept-language", ""),
        "origin": request.headers.get("origin", ""),
        "sec_ch_ua": request.headers.get("sec-ch-ua", ""),
        "sec_ch_ua_mobile": sec_ch_ua_mobile,
        "sec_ch_ua_platform": request.headers.get("sec-ch-ua-platform", ""),
        "sec_fetch_site": request.headers.get("sec-fetch-site", ""),
        "sec_fetch_mode": request.headers.get("sec-fetch-mode", ""),
        "sec_fetch_dest": request.headers.get("sec-fetch-dest", ""),
        "sec_fetch_user": request.headers.get("sec-fetch-user", ""),
        "username": username,
        "device_type": device_type,
    }


def _log_request_event(site: str, request: Request, status_code: int | None = None, duration_ms: float | None = None):
    payload = {
        "site": site,
        "event": "request",
        "route_group": _route_group_for_path(request.url.path),
        "logged_at": datetime.datetime.now(tz=APP_TIMEZONE).isoformat(timespec="seconds"),
        **_extract_public_client(request),
    }
    if status_code is not None:
        payload["status"] = status_code
    if duration_ms is not None:
        payload["duration_ms"] = round(duration_ms, 1)
    logger.info("public-access %s", json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":")))


def trace_product_filter(message: str, *args):
    if FINDER_DEBUG:
        logger.warning(message, *args)


def pdf_file_response(file_path: Path) -> FileResponse:
    return FileResponse(
        file_path,
        media_type="application/pdf",
        filename=file_path.name,
        content_disposition_type="inline",
        headers=PDF_MEDIA_HEADERS,
    )


def sanitize_name(value: str) -> str:
    cleaned = SAFE_CHARS_RE.sub("_", (value or "").strip().lower()).strip("_")
    return cleaned or "unknown"


def template_token_slug(value: str) -> str:
    return sanitize_name(value)


def json_for_html_script(value) -> str:
    return (
        json.dumps(value, separators=(",", ":"))
        .replace("&", "\\u0026")
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
    )


def product_slug(product: Product) -> str:
    return sanitize_name(product.model)


def public_product_slug(product: Product) -> str:
    name_slug = re.sub(r"[^a-z0-9]+", "-", (product.model or "").strip().lower()).strip("-")
    if name_slug:
        return f"{product.id}-{name_slug}"
    return str(product.id)


def product_public_identifier_candidates(product: Product) -> set[str]:
    name_slug = re.sub(r"[^a-z0-9]+", "-", (product.model or "").strip().lower()).strip("-")
    canonical_slug = product_slug(product)
    return {
        str(product.id),
        public_product_slug(product),
        name_slug,
        canonical_slug,
        canonical_slug.replace("_", "-"),
    }


def series_slug(series: Series) -> str:
    return sanitize_name(f"{series.product_type_key or 'series'}_{series.name}")


def public_series_slug(series: Series) -> str:
    name_slug = re.sub(r"[^a-z0-9]+", "-", (series.name or "").strip().lower()).strip("-")
    if name_slug:
        return f"{series.id}-{name_slug}"
    return str(series.id)


def series_public_identifier_candidates(series: Series) -> set[str]:
    name_slug = re.sub(r"[^a-z0-9]+", "-", (series.name or "").strip().lower()).strip("-")
    product_type_name_slug = re.sub(
        r"[^a-z0-9]+",
        "-",
        f"{series.product_type_key or 'series'} {series.name or ''}".strip().lower(),
    ).strip("-")
    canonical_slug = series_slug(series)
    return {
        str(series.id),
        public_series_slug(series),
        name_slug,
        product_type_name_slug,
        canonical_slug,
        canonical_slug.replace("_", "-"),
    }


def apply_product_type_parameter_presets(product_type: ProductType, preset_groups: list[ProductTypeParameterGroupPresetUpdate]):
    product_type.parameter_group_presets.clear()

    for group_index, group in enumerate(preset_groups or []):
        group_name = (group.group_name or "").strip()
        if not group_name:
            raise HTTPException(status_code=400, detail="Each parameter group needs a name.")

        group_model = ProductTypeParameterGroupPreset(
            group_name=group_name,
            sort_order=group_index,
        )

        seen_parameters: set[str] = set()
        for parameter_index, parameter in enumerate(group.parameters or []):
            parameter_name = (parameter.parameter_name or "").strip()
            if not parameter_name:
                raise HTTPException(status_code=400, detail=f"Each parameter in '{group_name}' needs a name.")

            normalized_parameter_name = parameter_name.casefold()
            if normalized_parameter_name in seen_parameters:
                raise HTTPException(
                    status_code=400,
                    detail=f"Parameter names must be unique within '{group_name}'.",
                )
            seen_parameters.add(normalized_parameter_name)

            preferred_unit = (parameter.preferred_unit or "").strip() or None
            value_string = None if parameter.value_string is None else str(parameter.value_string).strip() or None
            value_number = None if parameter.value_number is None else float(parameter.value_number)
            if value_string is not None and value_number is not None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Each parameter preset in '{group_name}' must be either text or number, not both.",
                )
            if value_string is not None and preferred_unit is not None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Text parameter presets in '{group_name}' cannot define a unit.",
                )
            group_model.parameter_presets.append(
                ProductTypeParameterPreset(
                    parameter_name=parameter_name,
                    sort_order=parameter_index,
                    preferred_unit=preferred_unit,
                    value_string=value_string,
                    value_number=value_number,
                )
            )

        product_type.parameter_group_presets.append(group_model)


def apply_product_type_rpm_line_presets(product_type: ProductType, preset_lines: list):
    product_type.rpm_line_presets.clear()

    for line_index, line in enumerate(preset_lines or []):
        line_model = ProductTypeRpmLinePreset(
            rpm=float(line.rpm),
            band_color=(line.band_color or "").strip() or None,
            sort_order=line_index,
        )

        for point_index, point in enumerate(line.points or []):
            line_model.point_presets.append(
                ProductTypeRpmPointPreset(
                    airflow=float(point.airflow),
                    pressure=float(point.pressure),
                    sort_order=point_index,
                )
            )

        product_type.rpm_line_presets.append(line_model)


def apply_product_type_efficiency_point_presets(product_type: ProductType, preset_points: list):
    product_type.efficiency_point_presets.clear()

    for point_index, point in enumerate(preset_points or []):
        product_type.efficiency_point_presets.append(
            ProductTypeEfficiencyPointPreset(
                airflow=float(point.airflow),
                efficiency_centre=point.efficiency_centre,
                efficiency_lower_end=point.efficiency_lower_end,
                efficiency_higher_end=point.efficiency_higher_end,
                permissible_use=point.permissible_use,
                sort_order=point_index,
            )
        )


def apply_product_type_presets(
    product_type: ProductType,
    preset_groups: list[ProductTypeParameterGroupPresetUpdate],
    preset_lines: list,
    preset_efficiency_points: list,
    product_template_id: str | None = None,
    series_template_id: str | None = None,
    printed_product_template_id: str | None = None,
    online_product_template_id: str | None = None,
):
    apply_product_type_parameter_presets(product_type, preset_groups)
    apply_product_type_rpm_line_presets(product_type, preset_lines)
    apply_product_type_efficiency_point_presets(product_type, preset_efficiency_points)
    product_template_id = validate_template_id(product_template_id, "product")
    product_type.series_template_id = validate_template_id(series_template_id, "series")
    if printed_product_template_id is None and online_product_template_id is None:
        printed_product_template_id = product_template_id
        online_product_template_id = product_template_id
    product_type.printed_product_template_id = validate_template_id(printed_product_template_id, "product")
    product_type.online_product_template_id = validate_template_id(online_product_template_id, "product")
    product_type.product_template_id = (
        product_type.online_product_template_id
        or product_type.printed_product_template_id
        or product_template_id
    )


def resolve_product_type_pdf_template_id(product_type: ProductType) -> str | None:
    candidate = product_type.product_type_template_id or "product_type-default"
    try:
        return validate_template_id(candidate, "product_type")
    except HTTPException:
        return None


def resolve_product_type_default_template_id(product_type: ProductType, variant: str) -> str | None:
    candidate = (
        product_type.printed_product_template_id
        if variant == "printed"
        else product_type.online_product_template_id
    ) or product_type.product_template_id
    if not candidate:
        return None
    try:
        return validate_template_id(candidate, "product")
    except HTTPException:
        return None


def resolve_template_pair(
    template_type: str,
    legacy_template_id: str | None = None,
    printed_template_id: str | None = None,
    online_template_id: str | None = None,
) -> tuple[str | None, str | None]:
    if printed_template_id is None and online_template_id is None and legacy_template_id is not None:
        printed_template_id = legacy_template_id
        online_template_id = legacy_template_id
    return (
        validate_template_id(printed_template_id, template_type),
        validate_template_id(online_template_id, template_type),
    )


def resolve_product_type_band_graph_style_defaults(product_type: ProductType) -> dict:
    return {
        "band_graph_background_color": normalize_color_value(product_type.band_graph_background_color) or "#ffffff",
        "band_graph_label_text_color": normalize_color_value(product_type.band_graph_label_text_color) or "#000000",
        "band_graph_faded_opacity": (
            product_type.band_graph_faded_opacity if product_type.band_graph_faded_opacity is not None else 0.18
        ),
        "band_graph_permissible_label_color": (
            normalize_color_value(product_type.band_graph_permissible_label_color)
            or normalize_color_value(product_type.band_graph_label_text_color)
            or "#000000"
        ),
    }


def build_product_type_rpm_line_presets(product_type: ProductType) -> list[dict]:
    preset_lines: list[dict] = []
    for line in product_type.rpm_line_presets or []:
        preset_lines.append(
            {
                "rpm": line.rpm,
                "band_color": line.band_color,
                "points": [
                    {
                        "airflow": point.airflow,
                        "pressure": point.pressure,
                    }
                    for point in line.point_presets or []
                ],
            }
        )
    return preset_lines


def build_product_type_efficiency_point_presets(product_type: ProductType) -> list[dict]:
    return [
        {
            "airflow": point.airflow,
            "efficiency_centre": point.efficiency_centre,
            "efficiency_lower_end": point.efficiency_lower_end,
            "efficiency_higher_end": point.efficiency_higher_end,
            "permissible_use": point.permissible_use,
        }
        for point in product_type.efficiency_point_presets or []
    ]


def load_template_registry() -> dict:
    if not TEMPLATE_REGISTRY_PATH.exists():
        return {"product_templates": [], "series_templates": [], "product_type_templates": []}
    with TEMPLATE_REGISTRY_PATH.open("r", encoding="utf-8") as handle:
        registry = json.load(handle)
    if not isinstance(registry, dict):
        return {"product_templates": [], "series_templates": [], "product_type_templates": []}
    return {
        "product_templates": list(registry.get("product_templates") or []),
        "series_templates": list(registry.get("series_templates") or []),
        "product_type_templates": list(registry.get("product_type_templates") or []),
    }


def save_template_registry(registry: dict):
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    TEMPLATE_REGISTRY_PATH.write_text(json.dumps(registry, indent=2) + "\n", encoding="utf-8")


FILE_MANAGER_ROOTS = {
    "data": Path(DEFAULT_DATA_DIR),
    "templates": TEMPLATES_DIR,
}
FILE_MANAGER_ALLOWED_TOP_LEVEL = {
    "data": {"product_images", "product_graphs", "product_pdfs", "product_type_pdfs", "series_graphs", "series_pdfs"},
    "templates": {"product", "series", "product_type"},
}
FILE_MANAGER_PROTECTED_RELATIVE_PATHS = {
    "data": {"fans.db", "backups", "product_images", "product_graphs", "product_pdfs", "product_type_pdfs", "series_graphs", "series_pdfs"},
    "templates": {"registry.json", "product", "series", "product_type"},
}


def file_manager_root_path(root_name: str) -> Path:
    normalized = (root_name or "").strip().lower()
    if normalized not in FILE_MANAGER_ROOTS:
        raise HTTPException(status_code=400, detail="Root must be 'data' or 'templates'.")
    root_path = FILE_MANAGER_ROOTS[normalized]
    root_path.mkdir(parents=True, exist_ok=True)
    return root_path


def file_manager_normalize_relative_path(relative_path: str | None) -> Path:
    candidate = Path((relative_path or "").strip())
    if str(candidate).strip() in {"", "."}:
        return Path()
    if candidate.is_absolute() or any(part in {"..", ""} for part in candidate.parts):
        raise HTTPException(status_code=400, detail="Invalid path.")
    return candidate


def file_manager_resolve_path(root_name: str, relative_path: str | None = None) -> Path:
    root_path = file_manager_root_path(root_name)
    relative = file_manager_normalize_relative_path(relative_path)
    resolved = (root_path / relative).resolve()
    if not resolved.is_relative_to(root_path.resolve()):
        raise HTTPException(status_code=400, detail="Invalid path.")
    return resolved


def file_manager_relative_path(root_name: str, resolved_path: Path) -> str:
    root_path = file_manager_root_path(root_name).resolve()
    return str(resolved_path.resolve().relative_to(root_path))


def file_manager_is_protected(root_name: str, relative_path: str | None) -> bool:
    normalized = file_manager_normalize_relative_path(relative_path)
    protected = FILE_MANAGER_PROTECTED_RELATIVE_PATHS.get((root_name or "").strip().lower(), set())
    return str(normalized) in protected


def file_manager_visible_entries(root_name: str, current_relative: str | None, entries: list[Path]) -> list[Path]:
    normalized_root = (root_name or "").strip().lower()
    if str(file_manager_normalize_relative_path(current_relative)) != "":
        return [entry for entry in entries if not entry.name.startswith(".")]

    allowed_top_level = FILE_MANAGER_ALLOWED_TOP_LEVEL.get(normalized_root)
    if allowed_top_level is None:
        return [entry for entry in entries if not entry.name.startswith(".")]

    visible: list[Path] = []
    for entry in entries:
        if entry.name.startswith("."):
            continue
        if entry.name in allowed_top_level:
            visible.append(entry)
        elif normalized_root == "templates" and entry.name == "registry.json":
            visible.append(entry)
    return visible


def file_manager_entry_to_response(root_name: str, entry: Path, root_relative: str | None = None) -> FileManagerEntryResponse:
    stat = entry.stat()
    relative = str(entry.resolve().relative_to(file_manager_root_path(root_name).resolve()))
    return FileManagerEntryResponse(
        name=entry.name,
        path=relative,
        type="directory" if entry.is_dir() else "file",
        size_bytes=None if entry.is_dir() else stat.st_size,
        modified_at=datetime.datetime.fromtimestamp(stat.st_mtime, tz=APP_TIMEZONE).isoformat(),
        protected=file_manager_is_protected(root_name, relative),
    )


def file_manager_list_directory(root_name: str, relative_path: str | None = None) -> FileManagerListingResponse:
    current_path = file_manager_resolve_path(root_name, relative_path)
    if not current_path.exists():
        raise HTTPException(status_code=404, detail="Folder not found.")
    if not current_path.is_dir():
        raise HTTPException(status_code=400, detail="Path is not a directory.")

    entries = sorted(current_path.iterdir(), key=lambda item: (not item.is_dir(), item.name.casefold()))
    entries = file_manager_visible_entries(root_name, relative_path, entries)
    parent_path = None
    if current_path.resolve() != file_manager_root_path(root_name).resolve():
        parent_path = str(current_path.resolve().relative_to(file_manager_root_path(root_name).resolve()).parent)
        if parent_path == ".":
            parent_path = ""

    return FileManagerListingResponse(
        root=root_name,
        path=str(file_manager_normalize_relative_path(relative_path)),
        parent_path=parent_path,
        entries=[file_manager_entry_to_response(root_name, entry) for entry in entries],
    )


def file_manager_read_text_file(root_name: str, relative_path: str) -> FileManagerContentResponse:
    target_path = file_manager_resolve_path(root_name, relative_path)
    if not target_path.exists():
        raise HTTPException(status_code=404, detail="File not found.")
    if not target_path.is_file():
        raise HTTPException(status_code=400, detail="Path is not a file.")

    try:
        content = target_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="This file is not UTF-8 text and cannot be edited here.")

    return FileManagerContentResponse(
        name=target_path.name,
        path=file_manager_relative_path(root_name, target_path),
        content=content,
    )


def file_manager_write_text_file(root_name: str, relative_path: str, content: str) -> FileManagerContentResponse:
    target_path = file_manager_resolve_path(root_name, relative_path)
    if not target_path.exists():
        raise HTTPException(status_code=404, detail="File not found.")
    if not target_path.is_file():
        raise HTTPException(status_code=400, detail="Path is not a file.")

    if file_manager_normalize_relative_path(relative_path).name == "registry.json" and root_name.strip().lower() == "templates":
        try:
            parsed_registry = json.loads(content)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail="registry.json must contain valid JSON.") from exc
        if not isinstance(parsed_registry, dict):
            raise HTTPException(status_code=400, detail="registry.json must contain a JSON object.")

    with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=str(target_path.parent), delete=False) as handle:
        handle.write(content)
        temp_path = Path(handle.name)

    try:
        temp_path.replace(target_path)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)

    if root_name.strip().lower() == "templates":
        sync_templates_after_file_change()

    return file_manager_read_text_file(root_name, relative_path)


def normalize_bulk_import_name(value: str | None) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    alias_map = {
        "producttype": "product_type_key",
        "producttypekey": "product_type_key",
        "product_type": "product_type_key",
        "seriesname": "series_name",
        "seriesid": "series_id",
        "productid": "product_id",
        "productmodel": "model",
        "model": "model",
        "name": "name",
        "filename": "file_name",
        "file": "file_name",
        "filepath": "file_name",
        "path": "file_name",
        "descriptionhtml": "description1_html",
        "description1html": "description1_html",
        "description2html": "description2_html",
        "description3html": "description3_html",
        "description4html": "description4_html",
        "printedtemplateid": "printed_template_id",
        "onlinetemplateid": "online_template_id",
        "templateid": "template_id",
        "showrpmbandshading": "show_rpm_band_shading",
        "green_system": "efficiency_centre",
        "upper_red_curve": "efficiency_higher_end",
        "lower_red_curve": "efficiency_lower_end",
        "red_high": "efficiency_higher_end",
        "red_low": "efficiency_lower_end",
        "grey_curve": "permissible_use",
    }
    if normalized in alias_map:
        return alias_map[normalized]

    rpm_match = re.fullmatch(r"(?:pressure_)?([0-9]+(?:\.[0-9]+)?)rpm", normalized)
    if rpm_match:
        return f"pressure_{rpm_match.group(1)}rpm"

    return normalized


def bulk_import_is_missing_value(value) -> bool:
    return isinstance(value, str) and value.strip().upper() == "#N/A"


def bulk_import_parse_numeric_candidate(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(parsed):
        return None
    return parsed


def bulk_import_copy_zero_airflow_values(rows: list[dict]) -> list[dict]:
    if not rows:
        return rows

    next_rows = [dict(row) for row in rows]
    first_row = next_rows[0]
    keys = list(first_row.keys())

    for key in keys[1:]:
        if not bulk_import_is_missing_value(first_row.get(key)):
            continue

        for row in next_rows[1:]:
            candidate = row.get(key)
            if bulk_import_parse_numeric_candidate(candidate) is None:
                continue
            first_row[key] = candidate
            break

    return next_rows


def bulk_import_find_highest_efficiency_overlay_key(rows: list[dict]) -> str | None:
    overlay_keys = ["efficiency_centre", "efficiency_higher_end", "efficiency_lower_end"]
    best_key = None
    best_value = float("-inf")

    for key in overlay_keys:
        for row in rows or []:
            candidate = bulk_import_parse_numeric_candidate(row.get(key))
            if candidate is None or candidate <= best_value:
                continue
            best_value = candidate
            best_key = key

    return best_key


def bulk_import_copy_permissible_use_from_highest_efficiency_line(rows: list[dict]) -> list[dict]:
    if not rows:
        return rows

    next_rows = [dict(row) for row in rows]
    source_key = bulk_import_find_highest_efficiency_overlay_key(next_rows)
    if not source_key:
        return next_rows

    for row in next_rows:
        permissible_use = row.get("permissible_use")
        if permissible_use is not None and not bulk_import_is_missing_value(permissible_use):
            continue
        highest_efficiency_value = row.get(source_key)
        if bulk_import_parse_numeric_candidate(highest_efficiency_value) is None:
            continue
        row["permissible_use"] = highest_efficiency_value
    return next_rows


def normalize_bulk_import_row(row: dict) -> dict:
    normalized: dict = {}
    for raw_key, raw_value in row.items():
        if raw_key is None:
            continue
        key = normalize_bulk_import_name(str(raw_key))
        if not key:
            continue
        value = raw_value
        if isinstance(value, str):
            value = value.strip()
            if not value:
                continue
            if value.upper() == "#N/A":
                continue
            if value.lower() in {"true", "false"}:
                value = value.lower() == "true"
        normalized[key] = value
    return normalized


def normalize_bulk_import_rows(rows: list[dict], copy_permissible_use: bool = True) -> list[dict]:
    copied_rows = bulk_import_copy_zero_airflow_values(rows)
    if copy_permissible_use:
        copied_rows = bulk_import_copy_permissible_use_from_highest_efficiency_line(copied_rows)
    return [normalize_bulk_import_row(row) for row in copied_rows]


def normalize_bulk_import_source_name(filename: str) -> str:
    return Path(str(filename or "").replace("\\", "/")).as_posix().lstrip("./")


def bulk_import_sheet_key(value: str | None) -> str:
    return str(value or "").strip().casefold()


def load_bulk_import_csv(file_name: str, content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = [row for row in reader if any(str(value or "").strip() for value in row.values())]
    return normalize_bulk_import_rows(rows, copy_permissible_use=not bulk_import_is_graph_sheet(rows))


def load_bulk_import_workbook(file_name: str, content: bytes) -> tuple[dict[str, list[dict]], dict[str, dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Excel workbook imports require the openpyxl package.") from exc

    workbook = load_workbook(io.BytesIO(content), data_only=True)
    tables: dict[str, list[dict]] = {}
    sheet_meta: dict[str, dict] = {}
    for sheet in workbook.worksheets:
        sheet_name = str(sheet.title or "").strip() or "Sheet"
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            tables[sheet_name] = []
            sheet_meta[sheet_name] = {
                "sheet_name": sheet_name,
                "row_count": 0,
                "raw_headers": [],
                "normalized_headers": [],
            }
            continue

        raw_headers = [str(header or "").strip() for header in rows[0]]
        normalized_headers = [normalize_bulk_import_name(header) for header in rows[0]]
        raw_table_rows: list[dict] = []
        for raw_row in rows[1:]:
            row_data: dict = {}
            for index, header in enumerate(normalized_headers):
                if not header or index >= len(raw_row):
                    continue
                value = raw_row[index]
                if value is None:
                    continue
                if isinstance(value, str):
                    value = value.strip()
                    if not value:
                        continue
                    if value.lower() in {"true", "false"}:
                        value = value.lower() == "true"
                row_data[header] = value
            if any(value is not None for value in row_data.values()):
                raw_table_rows.append(row_data)
        table_rows = normalize_bulk_import_rows(
            raw_table_rows,
            copy_permissible_use=not bulk_import_is_graph_sheet(raw_table_rows),
        )
        tables[sheet_name] = table_rows
        sheet_meta[sheet_name] = {
            "sheet_name": sheet_name,
            "row_count": len(table_rows),
            "raw_headers": raw_headers,
            "normalized_headers": normalized_headers,
        }
    return tables, sheet_meta


def load_graph_import_rows(file_name: str, content: bytes) -> list[list]:
    suffix = Path(str(file_name or "")).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        return [
            ["" if cell is None else cell for cell in row]
            for row in reader
            if any(str(cell or "").strip() for cell in row)
        ]

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="Excel workbook imports require the openpyxl package.") from exc

        workbook = load_workbook(io.BytesIO(content), data_only=True)
        for sheet in workbook.worksheets:
            rows = []
            for raw_row in sheet.iter_rows(values_only=True):
                row = ["" if cell is None else cell for cell in raw_row]
                if any(str(cell or "").strip() for cell in row):
                    rows.append(row)
            if rows:
                return rows
        return []

    raise HTTPException(status_code=400, detail="Please upload a CSV or Excel workbook.")


def load_bulk_import_manifest(content: bytes | str) -> dict:
    try:
        if isinstance(content, str):
            content = content.encode("utf-8")
        manifest = json.loads(content.decode("utf-8-sig"))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Import manifest must be valid JSON.") from exc

    if not isinstance(manifest, dict):
        raise HTTPException(status_code=400, detail="Import manifest must be a JSON object.")
    return manifest


def parse_int_or_none(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    try:
        text = str(value).strip()
        if not text:
            return None
        return int(float(text))
    except (TypeError, ValueError):
        return None


def bulk_import_find_series(db: Session, row: dict) -> Series | None:
    series_id = parse_int_or_none(row.get("id") or row.get("series_id"))
    if series_id is not None:
        return db.get(Series, series_id)

    name = str(row.get("name") or row.get("series_name") or "").strip()
    if not name:
        return None
    product_type_key = str(row.get("product_type_key") or "fan").strip() or "fan"
    return (
        db.query(Series)
        .join(ProductType)
        .filter(ProductType.key == product_type_key, func.lower(Series.name) == name.lower())
        .first()
    )


def bulk_import_find_product(db: Session, row: dict) -> Product | None:
    product_id = parse_int_or_none(row.get("id") or row.get("product_id"))
    if product_id is not None:
        return db.get(Product, product_id)

    model = str(row.get("model") or "").strip()
    if not model:
        return None

    query = db.query(Product).filter(func.lower(Product.model) == model.lower())
    series_id = parse_int_or_none(row.get("series_id"))
    if series_id is not None:
        query = query.filter(Product.series_id == series_id)
    else:
        series_name = str(row.get("series_name") or "").strip()
        if series_name:
            query = query.filter(func.lower(Product.series_name) == series_name.lower())
    product_type_key = str(row.get("product_type_key") or "").strip()
    if product_type_key:
        query = query.join(ProductType).filter(ProductType.key == product_type_key)
    return query.first()


def bulk_import_resolve_image_bytes(image_sources: dict[str, bytes], image_name: str) -> tuple[bytes | None, str | None]:
    normalized_name = normalize_bulk_import_source_name(image_name)
    if normalized_name in image_sources:
        return image_sources[normalized_name], normalized_name

    basename = Path(normalized_name).name
    matching = [key for key in image_sources if Path(key).name == basename]
    if len(matching) == 1:
        return image_sources[matching[0]], matching[0]
    if len(matching) > 1:
        raise HTTPException(
            status_code=400,
            detail=f'Image reference "{image_name}" is ambiguous. Use a unique relative path or filename.',
        )
    return None, None


def bulk_import_resolve_series_lookup(db: Session, row: dict, known_series: dict[tuple[str, str], int]) -> Series | None:
    series = bulk_import_find_series(db, row)
    if series is not None:
        return series

    series_name = str(row.get("series_name") or "").strip()
    if not series_name:
        return None
    product_type_key = str(row.get("product_type_key") or "fan").strip() or "fan"
    series_id = known_series.get((product_type_key.lower(), series_name.lower()))
    if series_id is None:
        return None
    return db.get(Series, series_id)


def bulk_import_parse_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(parsed):
        return None
    return parsed


def bulk_import_parse_integer(value):
    parsed = bulk_import_parse_number(value)
    if parsed is None:
        return None
    return int(round(parsed))


def bulk_import_parse_graph_line_token(header: str) -> float | None:
    normalized = bulk_import_sheet_key(header)
    candidates = [
        r"^(?:pressure[_ -]?)?([0-9]+(?:\.[0-9]+)?)rpm$",
        r"^([0-9]+(?:\.[0-9]+)?)rpm$",
        r"^pressure[_ -]?([0-9]+(?:\.[0-9]+)?)$",
        r"^rpm[_ -]?([0-9]+(?:\.[0-9]+)?)$",
        r"^([0-9]+(?:\.[0-9]+)?)$",
    ]
    for pattern in candidates:
        match = re.match(pattern, normalized)
        if not match:
            continue
        try:
            return float(match.group(1))
        except ValueError:
            continue
    return None


def bulk_import_interpolate_value(points: list[dict], axis_value: float):
    if not points:
        return None
    if len(points) == 1:
        return points[0]["value"]

    if axis_value <= points[0]["axis"]:
        return points[0]["value"]
    if axis_value >= points[-1]["axis"]:
        return points[-1]["value"]

    for index in range(len(points) - 1):
        left = points[index]
        right = points[index + 1]
        if axis_value < left["axis"] or axis_value > right["axis"]:
            continue
        span = right["axis"] - left["axis"]
        if span == 0:
            return right["value"]
        ratio = (axis_value - left["axis"]) / span
        return left["value"] + (right["value"] - left["value"]) * ratio

    return points[-1]["value"]


def bulk_import_build_interpolated_series(
    points: list[dict],
    axis_key: str = "airflow",
    value_key: str = "pressure",
    step: float = 1,
):
    numeric_points = (
        [
            {
                "point": point,
                "axis": bulk_import_parse_number(point.get(axis_key)),
                "value": bulk_import_parse_number(point.get(value_key)),
            }
            for point in points or []
        ]
    )
    numeric_points = [item for item in numeric_points if item["axis"] is not None and item["value"] is not None]
    numeric_points.sort(key=lambda item: item["axis"])

    if not numeric_points:
        return []

    if len(numeric_points) == 1:
        only = numeric_points[0]
        return [
            {
                **only["point"],
                axis_key: round(only["axis"]),
                value_key: round(only["value"]),
            }
        ]

    safe_step = step if isinstance(step, (int, float)) and step > 0 else 1
    min_axis = numeric_points[0]["axis"]
    max_axis = numeric_points[-1]["axis"]
    sampled = []
    axis = min_axis
    while axis <= max_axis:
        interpolated_value = bulk_import_interpolate_value(numeric_points, axis)
        if interpolated_value is None:
            axis += safe_step
            continue
        sampled.append(
            {
                **numeric_points[0]["point"],
                axis_key: round(axis, 3),
                value_key: round(interpolated_value),
            }
        )
        axis += safe_step

    if sampled and sampled[-1][axis_key] != round(max_axis, 3):
        interpolated_value = bulk_import_interpolate_value(numeric_points, max_axis)
        if interpolated_value is not None:
            sampled.append(
                {
                    **numeric_points[0]["point"],
                    axis_key: round(max_axis, 3),
                    value_key: round(interpolated_value),
                }
            )

    seen = set()
    result = []
    for point in sampled:
        key = point.get(axis_key)
        if key in seen:
            continue
        seen.add(key)
        result.append(point)
    return result


def bulk_import_downsample_series(
    points: list[dict],
    axis_key: str = "airflow",
    value_key: str = "pressure",
    target_count: int = 5,
    precision: int = 0,
):
    numeric_points = (
        [
            {
                "point": point,
                "axis": bulk_import_parse_number(point.get(axis_key)),
                "value": bulk_import_parse_number(point.get(value_key)),
            }
            for point in points or []
        ]
    )
    numeric_points = [item for item in numeric_points if item["axis"] is not None and item["value"] is not None]
    numeric_points.sort(key=lambda item: item["axis"])

    if len(numeric_points) <= target_count:
        return [item["point"] for item in numeric_points]

    sample_axes = []
    for index in range(target_count):
        t = 0 if target_count == 1 else index / (target_count - 1)
        sample_axes.append(
            round(
                numeric_points[0]["axis"]
                + (numeric_points[-1]["axis"] - numeric_points[0]["axis"]) * t
            )
        )

    template = numeric_points[0]["point"]
    sampled = []
    for axis in sample_axes:
        interpolated_value = bulk_import_interpolate_value(numeric_points, axis)
        if interpolated_value is None:
            continue
        sampled.append(
            {
                **template,
                axis_key: round(axis, precision),
                value_key: round(interpolated_value, precision),
            }
        )

    seen = set()
    result = []
    for point in sampled:
        axis_value = point.get(axis_key)
        if axis_value in seen:
            continue
        seen.add(axis_value)
        result.append(point)
    return result


def bulk_import_downsample_overlay_points(points: list[dict], value_keys: list[str], target_count: int = 5):
    merged_points = {}

    for value_key in value_keys:
        series_points = [point for point in (points or []) if point.get(value_key) is not None]
        sampled_points = bulk_import_downsample_series(
            series_points,
            "airflow",
            value_key,
            target_count,
            precision=0,
        )
        peak_point = max(
            (
                point
                for point in series_points
                if bulk_import_parse_number(point.get("airflow")) is not None
                and bulk_import_parse_number(point.get(value_key)) is not None
            ),
            key=lambda point: bulk_import_parse_number(point.get(value_key)) or float("-inf"),
            default=None,
        )
        if peak_point is not None and not any(
            bulk_import_parse_number(sampled_point.get("airflow"))
            == bulk_import_parse_number(peak_point.get("airflow"))
            for sampled_point in sampled_points
        ):
            sampled_points.append(dict(peak_point))
        for sampled_point in sampled_points:
            airflow = bulk_import_parse_number(sampled_point.get("airflow"))
            value = bulk_import_parse_number(sampled_point.get(value_key))
            if airflow is None or value is None:
                continue
            merge_key = str(airflow)
            if merge_key not in merged_points:
                merged_points[merge_key] = {
                    "airflow": airflow,
                    "efficiency_centre": None,
                    "efficiency_lower_end": None,
                    "efficiency_higher_end": None,
                    "permissible_use": None,
                }
            merged_points[merge_key][value_key] = round(value)

    return sorted(merged_points.values(), key=lambda point: point["airflow"])


def bulk_import_scale_overlay_points_to_highest_rpm_line(
    points: list[dict],
    rpm_lines: list[dict],
    rpm_points: list[dict],
):
    if not points:
        return points

    highest_line = sorted(
        [
            line
            for line in (rpm_lines or [])
            if bulk_import_parse_number(line.get("id")) is not None
            and bulk_import_parse_number(line.get("rpm")) is not None
        ],
        key=lambda line: bulk_import_parse_number(line.get("rpm")) or 0,
        reverse=True,
    )
    if not highest_line:
        return points
    highest_line = highest_line[0]
    highest_rpm_line_points = sorted(
        [
            {
                "airflow": bulk_import_parse_number(point.get("airflow")),
                "pressure": bulk_import_parse_number(point.get("pressure")),
            }
            for point in (rpm_points or [])
            if bulk_import_parse_number(point.get("rpm_line_id"))
            == bulk_import_parse_number(highest_line.get("id"))
        ],
        key=lambda point: point["airflow"] or 0,
    )
    highest_rpm_line_points = [
        point
        for point in highest_rpm_line_points
        if point["airflow"] is not None and point["pressure"] is not None
    ]
    if not highest_rpm_line_points:
        return points
    rpm_profile = bulk_import_build_highest_rpm_profile(highest_rpm_line_points)

    overlay_keys = [
        "efficiency_centre",
        "efficiency_lower_end",
        "efficiency_higher_end",
        "permissible_use",
    ]
    scaled_points: list[dict] = [dict(point) for point in points]

    for key in overlay_keys:
        key_points = sorted(
            [
                {
                    "airflow": bulk_import_parse_number(point.get("airflow")),
                    "value": bulk_import_parse_number(point.get(key)),
                }
                for point in (points or [])
                if bulk_import_parse_number(point.get("airflow")) is not None
                and bulk_import_parse_number(point.get(key)) is not None
            ],
            key=lambda point: point["airflow"] or 0,
        )
        if not key_points:
            continue

        peak_point = max(
            key_points,
            key=lambda point: (
                point["value"] if point["value"] is not None else float("-inf"),
                point["airflow"] if point["airflow"] is not None else float("-inf"),
            ),
        )
        scale_factor = bulk_import_find_best_overlay_scale_factor(peak_point, rpm_profile)
        if scale_factor is None or not math.isfinite(scale_factor):
            continue

        for scaled_point in scaled_points:
            value = bulk_import_parse_number(scaled_point.get(key))
            if value is None:
                continue
            scaled_point[key] = round(value * scale_factor)

    return scaled_points


def bulk_import_build_high_resolution_highest_rpm_line(rpm_lines: list[dict], rpm_points: list[dict], step: float = 1):
    highest_line = sorted(
        [
            line
            for line in (rpm_lines or [])
            if bulk_import_parse_number(line.get("id")) is not None and bulk_import_parse_number(line.get("rpm")) is not None
        ],
        key=lambda line: bulk_import_parse_number(line.get("rpm")) or 0,
        reverse=True,
    )
    if not highest_line:
        return []

    highest = highest_line[0]
    highest_line_points = sorted(
        [
            {
                "airflow": bulk_import_parse_number(point.get("airflow")),
                "pressure": bulk_import_parse_number(point.get("pressure")),
            }
            for point in (rpm_points or [])
            if bulk_import_parse_number(point.get("rpm_line_id")) == bulk_import_parse_number(highest.get("id"))
        ],
        key=lambda point: point["airflow"] or 0,
    )
    highest_line_points = [point for point in highest_line_points if point["airflow"] is not None and point["pressure"] is not None]
    return bulk_import_build_interpolated_series(highest_line_points, "airflow", "pressure", step)


def bulk_import_solve_overlay_scale_factor(terminal_point: dict, rpm_profile: dict):
    if not terminal_point or not rpm_profile:
        return None

    terminal_airflow = bulk_import_parse_number(terminal_point.get("airflow"))
    terminal_value = bulk_import_parse_number(terminal_point.get("value"))
    if terminal_airflow is None or terminal_value is None or terminal_airflow <= 0 or terminal_value <= 0:
        return None

    points = bulk_import_normalised_graph_series(rpm_profile.get("points") or [], "axis", "value")
    if len(points) < 2:
        return None

    candidates = []
    for index in range(len(points) - 1):
        left = points[index]
        right = points[index + 1]
        left_axis = bulk_import_parse_number(left.get("axis"))
        left_value = bulk_import_parse_number(left.get("value"))
        right_axis = bulk_import_parse_number(right.get("axis"))
        right_value = bulk_import_parse_number(right.get("value"))
        if None in (left_axis, left_value, right_axis, right_value):
            continue
        if right_axis <= left_axis:
            continue

        slope = (right_value - left_value) / (right_axis - left_axis)
        denominator = terminal_value - slope * terminal_airflow
        if abs(denominator) < 1e-12:
            continue

        scale_factor = (left_value - slope * left_axis) / denominator
        if not math.isfinite(scale_factor) or scale_factor <= 0:
            continue

        scaled_airflow = terminal_airflow * scale_factor
        if scaled_airflow < left_axis - 1e-9 or scaled_airflow > right_axis + 1e-9:
            continue

        target_value = bulk_import_interpolate_value(points, scaled_airflow)
        if target_value is None or not math.isfinite(target_value):
            continue

        residual = abs(target_value - terminal_value * scale_factor)
        candidates.append((residual, abs(scale_factor - 1), scale_factor))

    if candidates:
        candidates.sort(key=lambda item: (item[0], item[1], item[2]))
        return candidates[0][2]

    return None


def bulk_import_normalised_graph_series(points: list[dict], axis_key: str = "airflow", value_key: str = "pressure"):
    return sorted(
        [
            {
                "axis": bulk_import_parse_number(point.get(axis_key)),
                "value": bulk_import_parse_number(point.get(value_key)),
            }
            for point in (points or [])
        ],
        key=lambda point: point["axis"] or 0,
    )


def bulk_import_chart_space_axis_extents(line_points: list[dict]):
    flow_raw_max = max(
        [bulk_import_parse_number(point.get("airflow")) for point in (line_points or []) if bulk_import_parse_number(point.get("airflow")) is not None]
        or [1]
    )
    pressure_raw_max = max(
        [bulk_import_parse_number(point.get("pressure")) for point in (line_points or []) if bulk_import_parse_number(point.get("pressure")) is not None]
        or [1]
    )
    return {
        "flowMax": (flow_raw_max * 1.05) if flow_raw_max > 0 else 1,
        "pressureMax": (pressure_raw_max * 1.05) if pressure_raw_max > 0 else 1,
        # Overlay values are pressure-based now, so keep them on the same scale
        # as the RPM curve rather than the old percentage-style 0-100 range.
        "overlayMax": (pressure_raw_max * 1.05) if pressure_raw_max > 0 else 1,
    }


def bulk_import_build_highest_rpm_profile(rpm_line_points: list[dict]):
    numeric_points = bulk_import_normalised_graph_series(rpm_line_points, "airflow", "pressure")
    return {
        "points": numeric_points,
        "axis_extents": bulk_import_chart_space_axis_extents(rpm_line_points),
    }


def bulk_import_chart_space_difference_at_scaled_terminal_point(
    terminal_point: dict,
    scale_factor: float,
    rpm_profile: dict,
):
    scaled_airflow = bulk_import_parse_number(terminal_point.get("airflow"))
    scaled_value = bulk_import_parse_number(terminal_point.get("value"))
    if scaled_airflow is None or scaled_value is None:
        return float("inf")

    scaled_point = {
        "airflow": scaled_airflow * scale_factor,
        "pressure": scaled_value * scale_factor,
    }
    target_pressure = bulk_import_interpolate_value(rpm_profile.get("points") or [], scaled_point["airflow"])
    if target_pressure is None or not math.isfinite(target_pressure):
        return float("inf")

    axis_extents = rpm_profile.get("axis_extents") or {"overlayMax": 1, "pressureMax": 1}
    overlay_position = scaled_point["pressure"] / axis_extents["overlayMax"]
    rpm_position = target_pressure / axis_extents["pressureMax"]
    return abs(overlay_position - rpm_position)


def bulk_import_find_best_overlay_scale_factor(terminal_point: dict, rpm_profile: dict):
    if not terminal_point or not rpm_profile:
        return None

    terminal_airflow = bulk_import_parse_number(terminal_point.get("airflow"))
    terminal_value = bulk_import_parse_number(terminal_point.get("value"))
    if terminal_airflow is None or terminal_value is None or terminal_airflow <= 0 or terminal_value <= 0:
        return None

    target_pressure = bulk_import_interpolate_value(
        rpm_profile.get("points") or [],
        terminal_airflow,
    )
    if target_pressure is None or not math.isfinite(target_pressure):
        return None

    scale_factor = target_pressure / terminal_value
    if not math.isfinite(scale_factor) or scale_factor <= 0:
        return None

    return max(0.01, scale_factor)


def bulk_import_is_graph_sheet(rows: list[dict]) -> bool:
    if not rows:
        return False
    first_row = rows[0]
    headers = [bulk_import_sheet_key(header) for header in first_row.keys()]
    if not headers:
        return False
    first_header = headers[0]
    if first_header not in {"airflow_l_s", "airflow"}:
        return False
    return any(
        header.startswith("pressure_") or header in {"efficiency_centre", "efficiency_lower_end", "efficiency_higher_end", "permissible_use"}
        for header in headers[1:]
    )


def bulk_import_manifest_sheet_config(manifest: dict | None, sheet_name: str) -> dict:
    if not isinstance(manifest, dict):
        return {}
    raw_sheet_name = str(sheet_name or "").strip()
    desired = bulk_import_sheet_key(sheet_name)
    sheets = manifest.get("sheets")
    if isinstance(sheets, dict):
        if raw_sheet_name and raw_sheet_name in sheets and isinstance(sheets.get(raw_sheet_name), dict):
            return dict(sheets[raw_sheet_name])
        for key, value in sheets.items():
            if bulk_import_sheet_key(key) == desired and isinstance(value, dict):
                return dict(value)
    elif isinstance(sheets, list):
        for item in sheets:
            if not isinstance(item, dict):
                continue
            candidates = [
                item.get("sheet_name"),
                item.get("sheet"),
                item.get("name"),
                item.get("product_name"),
                item.get("model"),
            ]
            if any(str(candidate or "").strip() == raw_sheet_name for candidate in candidates if candidate is not None):
                return dict(item)
            if any(bulk_import_sheet_key(candidate) == desired for candidate in candidates if candidate is not None):
                return dict(item)
    products = manifest.get("products")
    if isinstance(products, dict):
        if raw_sheet_name and raw_sheet_name in products and isinstance(products.get(raw_sheet_name), dict):
            return dict(products[raw_sheet_name])
        for key, value in products.items():
            if bulk_import_sheet_key(key) == desired and isinstance(value, dict):
                return dict(value)
    elif isinstance(products, list):
        for item in products:
            if not isinstance(item, dict):
                continue
            candidates = [
                item.get("sheet_name"),
                item.get("sheet"),
                item.get("name"),
                item.get("product_name"),
                item.get("model"),
            ]
            if any(str(candidate or "").strip() == raw_sheet_name for candidate in candidates if candidate is not None):
                return dict(item)
            if any(bulk_import_sheet_key(candidate) == desired for candidate in candidates if candidate is not None):
                return dict(item)
    return {}


def bulk_import_build_graph_state(
    rows: list[dict],
    downsample_imported_curves: bool = True,
    downsample_point_count: int = 5,
    permissible_use_mode: str = "both",
    permissible_use_source_key: str = "efficiency_higher_end",
):
    if not rows:
        return {"rpmLines": [], "rpmPoints": [], "efficiencyPoints": []}

    headers = [bulk_import_sheet_key(header) for header in rows[0].keys()]
    ordered_headers = list(rows[0].keys())
    airflow_header = headers[0]
    if airflow_header not in {"airflow_l_s", "airflow"}:
        raise HTTPException(status_code=400, detail='The first column must be "airflow_l_s".')

    pressure_columns = []
    overlay_columns = {"efficiency_centre", "efficiency_lower_end", "efficiency_higher_end", "permissible_use"}
    for index, header in enumerate(headers[1:], start=1):
        original_header = ordered_headers[index]
        if not header:
            continue
        if header in overlay_columns:
            continue
        rpm = bulk_import_parse_graph_line_token(header)
        if rpm is None:
            raise HTTPException(status_code=400, detail=f'Column "{original_header}" is not recognised.')
        pressure_columns.append({"index": index, "header": original_header, "rpm": rpm})

    rpm_lines = [
        {"id": idx + 1, "rpm": column["rpm"], "band_color": None}
        for idx, column in enumerate(pressure_columns)
    ]
    rpm_line_by_rpm = {str(line["rpm"]): line for line in rpm_lines}

    next_rpm_points = []
    next_efficiency_points = []
    seen_airflows = set()
    previous_airflow = None

    for row_index, row in enumerate(rows):
        rounded_airflow = bulk_import_parse_integer(row.get(ordered_headers[0]))
        if rounded_airflow is None:
            raise HTTPException(status_code=400, detail=f"Row {row_index + 2} is missing an airflow value.")
        if rounded_airflow in seen_airflows:
            raise HTTPException(status_code=400, detail=f"Duplicate airflow value found: {rounded_airflow}.")
        if previous_airflow is not None and rounded_airflow <= previous_airflow:
            raise HTTPException(status_code=400, detail=f"Airflow must increase row by row. Row {row_index + 2} is out of order.")
        seen_airflows.add(rounded_airflow)
        previous_airflow = rounded_airflow

        for column in pressure_columns:
            pressure = bulk_import_parse_number(row.get(column["header"]))
            if pressure is None:
                continue
            rounded_pressure = bulk_import_parse_integer(row.get(column["header"]))
            line = rpm_line_by_rpm.get(str(column["rpm"]))
            if not line:
                continue
            next_rpm_points.append(
                {
                    "id": len(next_rpm_points) + 1,
                    "product_id": None,
                    "rpm_line_id": line["id"],
                    "rpm": line["rpm"],
                    "airflow": rounded_airflow,
                    "pressure": rounded_pressure,
                }
            )

        efficiency_point = {
            "id": len(next_efficiency_points) + 1,
            "product_id": None,
            "airflow": rounded_airflow,
            "efficiency_centre": None,
            "efficiency_lower_end": None,
            "efficiency_higher_end": None,
            "permissible_use": None,
        }
        has_overlay = False
        for overlay_key in overlay_columns:
            if overlay_key not in headers:
                continue
            value = bulk_import_parse_integer(row.get(overlay_key))
            if value is not None:
                efficiency_point[overlay_key] = value
                has_overlay = True
        if has_overlay:
            next_efficiency_points.append(efficiency_point)

    source_overlay_key = permissible_use_source_key if permissible_use_source_key in {
        "efficiency_higher_end",
        "efficiency_lower_end",
    } else "efficiency_higher_end"
    if normalize_permissible_use_mode(permissible_use_mode) == "dedicated":
        for point in next_efficiency_points:
            permissible_use = point.get("permissible_use")
            if permissible_use is not None and permissible_use != "":
                continue
            source_value = point.get(source_overlay_key)
            if source_value is None or source_value == "":
                continue
            point["permissible_use"] = source_value

    next_efficiency_points = bulk_import_scale_overlay_points_to_highest_rpm_line(
        next_efficiency_points,
        rpm_lines,
        next_rpm_points,
    )

    if downsample_imported_curves:
        next_rpm_points_by_line = {}
        for point in next_rpm_points:
            line_id = point.get("rpm_line_id")
            next_rpm_points_by_line.setdefault(line_id, []).append(point)
        adjusted_rpm_points = []
        for line_points in next_rpm_points_by_line.values():
            adjusted_rpm_points.extend(
                bulk_import_downsample_series(line_points, "airflow", "pressure", downsample_point_count)
            )
    else:
        adjusted_rpm_points = next_rpm_points

    adjusted_efficiency_points = (
        bulk_import_downsample_overlay_points(next_efficiency_points, list(overlay_columns), downsample_point_count)
        if downsample_imported_curves
        else next_efficiency_points
    )

    return {
        "rpmLines": rpm_lines,
        "rpmPoints": adjusted_rpm_points,
        "efficiencyPoints": adjusted_efficiency_points,
    }


def bulk_import_attach_rpm_points_to_lines(rpm_lines: list[dict], rpm_points: list[dict]) -> list[dict]:
    points_by_line_id: dict[int, list[dict]] = {}
    for point in rpm_points or []:
        line_id = parse_int_or_none(point.get("rpm_line_id"))
        if line_id is None:
            continue
        points_by_line_id.setdefault(line_id, []).append(
            {
                "airflow": point.get("airflow"),
                "pressure": point.get("pressure"),
            }
        )

    enriched_lines: list[dict] = []
    for line in rpm_lines or []:
        line_id = parse_int_or_none(line.get("id"))
        if line_id is None:
            continue
        enriched_lines.append(
            {
                "rpm": line.get("rpm"),
                "band_color": line.get("band_color"),
                "points": sorted(points_by_line_id.get(line_id, []), key=lambda point: bulk_import_parse_number(point.get("airflow")) or 0),
            }
        )
    return enriched_lines


def bulk_import_describe_sheet_normalization(
    sheet_name: str,
    rows: list[dict],
    sheet_meta: dict | None = None,
    include_in_import: bool = True,
    error: str | None = None,
) -> dict:
    meta = sheet_meta or {}
    raw_headers = list(meta.get("raw_headers") or [])
    normalized_headers = list(meta.get("normalized_headers") or [])
    if not normalized_headers and rows:
        normalized_headers = list(rows[0].keys())
    if not raw_headers and normalized_headers:
        raw_headers = list(normalized_headers)
    headers = normalized_headers or (list(rows[0].keys()) if rows else [])
    raw_headers = raw_headers[: len(headers)]
    if len(raw_headers) < len(headers):
        raw_headers.extend([""] * (len(headers) - len(raw_headers)))

    columns = []
    if headers:
        overlay_columns = {"efficiency_centre", "efficiency_lower_end", "efficiency_higher_end", "permissible_use"}
        for index, normalized_header in enumerate(headers):
            raw_header = str(raw_headers[index] or "").strip() if index < len(raw_headers) else ""
            role = "ignored"
            rpm = None
            reason = ""
            if not raw_header and not normalized_header:
                reason = "Blank header, so it is ignored."
            elif index == 0:
                role = "airflow"
                if raw_header and normalized_header and normalized_header != raw_header:
                    reason = f"Normalized from '{raw_header}' and used as the airflow column."
                else:
                    reason = "First column is used as the airflow axis."
            elif normalized_header in overlay_columns:
                role = "overlay"
                if raw_header and normalized_header and normalized_header != raw_header:
                    reason = f"Normalized from '{raw_header}' and matched an efficiency overlay column."
                else:
                    reason = "Matched an efficiency overlay column."
            else:
                rpm = bulk_import_parse_graph_line_token(normalized_header)
                if rpm is not None:
                    role = "rpm_line"
                    if raw_header and normalized_header and normalized_header != raw_header:
                        reason = f"Normalized from '{raw_header}' and parsed as the {rpm:g} RPM curve."
                    else:
                        reason = f"Parsed as the {rpm:g} RPM curve."
                elif not reason:
                    if raw_header and normalized_header and normalized_header != raw_header:
                        reason = f"Normalized from '{raw_header}' but it did not match airflow, overlay, or RPM curve patterns."
                    else:
                        reason = "Did not match airflow, overlay, or RPM curve patterns."
            columns.append(
                {
                    "raw_header": raw_header,
                    "normalized_header": normalized_header,
                    "role": role,
                    "rpm": rpm,
                    "reason": reason,
                }
            )

    graph_state = None
    if not error and bulk_import_is_graph_sheet(rows):
        try:
            graph_state = bulk_import_build_graph_state(rows)
        except Exception as exc:
            error = str(exc)

    efficiency_points = graph_state["efficiencyPoints"] if graph_state else []
    return {
        "sheet_name": sheet_name,
        "row_count": int(meta.get("row_count") or len(rows) or 0),
        "include_in_import": include_in_import,
        "raw_headers": raw_headers,
        "normalized_headers": headers,
        "columns": columns,
        "rpm_line_count": len(graph_state["rpmLines"]) if graph_state else 0,
        "rpm_point_count": len(graph_state["rpmPoints"]) if graph_state else 0,
        "efficiency_point_count": len(graph_state["efficiencyPoints"]) if graph_state else 0,
        "has_efficiency_upper": any(point.get("efficiency_higher_end") is not None for point in efficiency_points),
        "has_efficiency_lower": any(point.get("efficiency_lower_end") is not None for point in efficiency_points),
        "error": error,
    }


def bulk_import_process_payloads(
    db: Session,
    tables: dict[str, list[dict]],
    image_sources: dict[str, bytes],
    sheet_meta: dict[str, dict],
    dry_run: bool,
    default_downsample_imported_curves: bool = True,
    default_downsample_point_count: int = 5,
) -> BulkImportResponse:
    report = BulkImportResponse(dry_run=dry_run)
    manifest = tables.get("__manifest__") if isinstance(tables.get("__manifest__"), dict) else {}
    defaults = dict(manifest.get("defaults") or {}) if isinstance(manifest, dict) else {}
    defaults.setdefault("downsample_imported_curves", default_downsample_imported_curves)
    defaults.setdefault("downsample_point_count", default_downsample_point_count)
    default_series_id = parse_int_or_none(defaults.get("series_id")) if "series_id" in defaults else None
    default_series_name = str(defaults.get("series_name") or "").strip() or None
    report.tables = [
        BulkImportTableSummaryResponse(name=table_name, kind="sheet" if bulk_import_is_graph_sheet(rows) else "table", row_count=len(rows))
        for table_name, rows in sorted(((name, rows) for name, rows in tables.items() if name != "__manifest__"), key=lambda item: item[0].casefold())
        if isinstance(rows, list)
    ]
    report.sheet_normalizations = [
        BulkImportSheetNormalizationResponse(**bulk_import_describe_sheet_normalization(table_name, rows, sheet_meta.get(table_name), True))
        for table_name, rows in sorted(((name, rows) for name, rows in tables.items() if name != "__manifest__"), key=lambda item: item[0].casefold())
        if isinstance(rows, list)
    ]

    known_series: dict[tuple[str, str], int] = {}
    touched_products: dict[int, Product] = {}
    touched_series: dict[int, Series] = {}
    product_sheet_lookup: dict[str, dict] = {}
    manifest_images = manifest.get("images") if isinstance(manifest, dict) else []
    if isinstance(manifest_images, dict):
        manifest_images = list(manifest_images.values())
    if not isinstance(manifest_images, list):
        manifest_images = []
    manifest_image_counts: dict[str, int] = {}
    for image_spec in manifest_images:
        if not isinstance(image_spec, dict):
            continue
        sheet_key = bulk_import_sheet_key(
            image_spec.get("sheet_name")
            or image_spec.get("sheet")
            or image_spec.get("product_name")
            or image_spec.get("series_name")
        )
        if sheet_key:
            manifest_image_counts[sheet_key] = manifest_image_counts.get(sheet_key, 0) + 1

    for table_name, rows in tables.items():
        if table_name == "__manifest__" or not isinstance(rows, list) or not bulk_import_is_graph_sheet(rows):
            continue

        sheet_config = bulk_import_manifest_sheet_config(manifest, table_name)
        if bool(sheet_config.get("skip_import")):
            report.skipped_sheets.append(table_name)
            for entry in report.sheet_normalizations:
                if entry.sheet_name == table_name:
                    entry.include_in_import = False
                    break
            continue
        product_model = str(
            sheet_config.get("product_model")
            or sheet_config.get("product_name")
            or sheet_config.get("model")
            or table_name
        ).strip() or table_name
        product_type_key = str(
            sheet_config.get("product_type_key")
            or defaults.get("product_type_key")
            or "fan"
        ).strip() or "fan"
        if "series_name" in sheet_config:
            series_name = str(sheet_config.get("series_name") or "").strip() or None
        else:
            series_name = default_series_name
        if "series_id" in sheet_config:
            series_id = parse_int_or_none(sheet_config.get("series_id"))
        else:
            series_id = default_series_id
        product_id = parse_int_or_none(sheet_config.get("product_id") or sheet_config.get("id"))
        downsample_imported_curves = bool(
            sheet_config.get(
                "downsample_imported_curves",
                defaults.get("downsample_imported_curves", True),
            )
        )
        downsample_point_count = parse_int_or_none(
            sheet_config.get("downsample_point_count", defaults.get("downsample_point_count", 5))
        ) or 5
        permissible_use_mode = normalize_permissible_use_mode(
            sheet_config.get("permissible_use_mode", defaults.get("permissible_use_mode", "both"))
        )
        permissible_use_source_key = (
            "efficiency_lower_end"
            if bool(sheet_config.get("generate_missing_permissible_use_from_lower", defaults.get("generate_missing_permissible_use_from_lower", False)))
            else "efficiency_higher_end"
        )
        try:
            graph_state = bulk_import_build_graph_state(
                rows,
                downsample_imported_curves=downsample_imported_curves,
                downsample_point_count=downsample_point_count,
                permissible_use_mode=permissible_use_mode,
                permissible_use_source_key=permissible_use_source_key,
            )
            graph_rpm_lines = bulk_import_attach_rpm_points_to_lines(graph_state["rpmLines"], graph_state["rpmPoints"])
            for entry in report.sheet_normalizations:
                if entry.sheet_name == table_name:
                    entry.rpm_line_count = len(graph_state["rpmLines"])
                    entry.rpm_point_count = len(graph_state["rpmPoints"])
                    entry.efficiency_point_count = len(graph_state["efficiencyPoints"])
                    entry.error = None
                    entry.include_in_import = True
                    break
            payload = {
                "model": product_model,
                "product_type_key": product_type_key,
                "series_id": series_id,
                "series_name": series_name,
                "template_id": sheet_config.get("template_id"),
                "printed_template_id": sheet_config.get("printed_template_id"),
                "online_template_id": sheet_config.get("online_template_id"),
                "description1_html": sheet_config.get("description1_html"),
                "description2_html": sheet_config.get("description2_html"),
                "description3_html": sheet_config.get("description3_html"),
                "comments_html": sheet_config.get("comments_html"),
                "show_rpm_band_shading": sheet_config.get("show_rpm_band_shading", True),
                "permissible_use_mode": permissible_use_mode,
                "band_graph_background_color": sheet_config.get("band_graph_background_color"),
                "band_graph_label_text_color": sheet_config.get("band_graph_label_text_color"),
                "band_graph_faded_opacity": sheet_config.get("band_graph_faded_opacity"),
                "band_graph_permissible_label_color": sheet_config.get("band_graph_permissible_label_color"),
                "rpm_lines": graph_rpm_lines,
                "efficiency_points": graph_state["efficiencyPoints"],
            }

            if series_id is None and series_name:
                series = bulk_import_find_series(db, {"name": series_name, "product_type_key": product_type_key})
                if series is None:
                    if dry_run:
                        report.created_series += 1
                    else:
                        series = create_series(SeriesCreate(name=series_name, product_type_key=product_type_key), db)
                        report.created_series += 1
                if series is not None:
                    payload["series_id"] = series.id
                    payload["series_name"] = series.name
            product = None
            if product_id is not None:
                product = db.get(Product, product_id)
            if product is None:
                product = bulk_import_find_product(db, {"model": product_model, "series_id": payload.get("series_id"), "series_name": series_name, "product_type_key": product_type_key})

            if product is None:
                if dry_run:
                    report.created_products += 1
                    product_sheet_lookup[bulk_import_sheet_key(table_name)] = {
                        "model": product_model,
                        "product_type_key": product_type_key,
                        "series_name": series_name,
                        "series_id": payload.get("series_id"),
                    }
                    continue
                body = ProductCreate(**payload)
                product = create_product(body, db)
                report.created_products += 1
            else:
                if dry_run:
                    report.updated_products += 1
                else:
                    update_body = ProductUpdate(**{key: value for key, value in payload.items() if value is not None})
                    product = update_product(product.id, update_body, db)
                    replace_product_graph_data(db, product, graph_rpm_lines, graph_state["efficiencyPoints"])
                    db.commit()
                    db.refresh(product)
                    report.updated_products += 1

            if product is not None and product.id is not None:
                touched_products[product.id] = product
                product_sheet_lookup[bulk_import_sheet_key(table_name)] = {
                    "product_id": product.id,
                    "model": product.model,
                    "series_name": product.series_name,
                    "product_type_key": product.product_type_key,
                }
                product_sheet_lookup[bulk_import_sheet_key(product_model)] = {
                    "product_id": product.id,
                    "model": product.model,
                    "series_name": product.series_name,
                    "product_type_key": product.product_type_key,
                }
                report.manifest_sheets.append(
                    BulkImportManifestSheetResponse(
                        sheet_name=table_name,
                        product_model=product_model,
                        product_type_key=product_type_key,
                        series_id=parse_int_or_none(payload.get("series_id")),
                        series_name=str(payload.get("series_name") or series_name or "").strip() or None,
                        image_count=manifest_image_counts.get(bulk_import_sheet_key(table_name), 0),
                    )
                )
        except Exception as exc:
            report.errors.append(f"Sheet import failed for '{table_name}': {exc}")
            for entry in report.sheet_normalizations:
                if entry.sheet_name == table_name:
                    entry.error = str(exc)
                    break

    def attach_product_image(product: Product, file_name: str, image_bytes: bytes, sort_order: int | None = None):
        if dry_run:
            report.created_product_images += 1
            return
        existing_image = next((image for image in product.product_images if image.file_name == file_name), None)
        if existing_image is None:
            existing_image = ProductImage(
                product_id=product.id,
                file_name=file_name,
                sort_order=sort_order if sort_order is not None else len(product.product_images),
            )
            db.add(existing_image)
            db.flush()
        else:
            existing_image.sort_order = sort_order if sort_order is not None else existing_image.sort_order
        target_path = product_image_target_path(product.id, existing_image.file_name)
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_bytes(image_bytes)
        touched_products[product.id] = product
        report.created_product_images += 1

    def attach_series_image(series: Series, file_name: str, image_bytes: bytes, sort_order: int | None = None):
        if dry_run:
            report.created_series_images += 1
            return
        existing_image = next((image for image in series.series_images if image.file_name == file_name), None)
        if existing_image is None:
            existing_image = SeriesImage(
                series_id=series.id,
                file_name=file_name,
                sort_order=sort_order if sort_order is not None else len(series.series_images),
            )
            db.add(existing_image)
            db.flush()
        else:
            existing_image.sort_order = sort_order if sort_order is not None else existing_image.sort_order
        target_path = series_image_target_path(series.id, existing_image.file_name)
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_bytes(image_bytes)
        touched_series[series.id] = series
        report.created_series_images += 1

    for image_spec in manifest_images:
        if not isinstance(image_spec, dict):
            continue
        sheet_name = image_spec.get("sheet_name") or image_spec.get("sheet") or image_spec.get("product_name") or image_spec.get("series_name")
        file_name = str(image_spec.get("file_name") or Path(str(image_spec.get("source") or "")).name or "").strip()
        source_name = str(image_spec.get("source") or file_name).strip()
        kind = bulk_import_sheet_key(image_spec.get("kind") or "product")
        if not sheet_name or not file_name or not source_name:
            continue
        image_bytes, _ = bulk_import_resolve_image_bytes(image_sources, source_name)
        if image_bytes is None:
            report.errors.append(f'Image "{source_name}" listed in the manifest was not uploaded.')
            continue

        if kind == "series":
            series = bulk_import_find_series(db, {"name": sheet_name, "product_type_key": image_spec.get("product_type_key") or defaults.get("product_type_key") or "fan"})
            if series is None and not dry_run:
                continue
            if series is None and dry_run:
                report.created_series_images += 1
                continue
            attach_series_image(series, file_name, image_bytes, parse_int_or_none(image_spec.get("sort_order")))
        else:
            lookup = product_sheet_lookup.get(bulk_import_sheet_key(sheet_name))
            product = db.get(Product, lookup["product_id"]) if lookup and lookup.get("product_id") else bulk_import_find_product(db, {"model": sheet_name, "product_type_key": image_spec.get("product_type_key") or defaults.get("product_type_key") or "fan"})
            if product is None and not dry_run:
                continue
            if product is None and dry_run:
                report.created_product_images += 1
                continue
            attach_product_image(product, file_name, image_bytes, parse_int_or_none(image_spec.get("sort_order")))

    for table_name in ("series", "products", "product_images", "series_images"):
        rows = tables.get(table_name, [])
        if not isinstance(rows, list):
            continue
        if table_name == "series":
            for row in rows:
                try:
                    series = bulk_import_find_series(db, row)
                    payload = dict(row)
                    if series is None:
                        if "name" not in payload or not str(payload.get("name") or "").strip():
                            report.errors.append("Series row is missing a name.")
                            continue
                        if dry_run:
                            report.created_series += 1
                            continue
                        series = create_series(SeriesCreate(**payload), db)
                        report.created_series += 1
                    else:
                        if dry_run:
                            report.updated_series += 1
                        else:
                            update_series(series.id, SeriesUpdate(**payload), db)
                            series = db.get(Series, series.id)
                            report.updated_series += 1
                    if series is not None and series.id is not None:
                        touched_series[series.id] = series
                except Exception as exc:
                    report.errors.append(f"Series import failed for row {row!r}: {exc}")
        elif table_name == "products":
            for row in rows:
                try:
                    payload = dict(row)
                    series = bulk_import_resolve_series_lookup(db, payload, known_series)
                    if series is not None:
                        payload.setdefault("series_id", series.id)
                        payload.setdefault("series_name", series.name)
                        payload.setdefault("product_type_key", series.product_type.key)
                    payload.setdefault("product_type_key", "fan")

                    product = bulk_import_find_product(db, payload)
                    if product is None:
                        if "model" not in payload or not str(payload.get("model") or "").strip():
                            report.errors.append("Product row is missing a model.")
                            continue
                        if dry_run:
                            report.created_products += 1
                            continue
                        product = create_product(ProductCreate(**payload), db)
                        report.created_products += 1
                    else:
                        if dry_run:
                            report.updated_products += 1
                        else:
                            update_product(product.id, ProductUpdate(**payload), db)
                            product = db.get(Product, product.id)
                            report.updated_products += 1
                    if product is not None and product.id is not None:
                        touched_products[product.id] = product
                except Exception as exc:
                    report.errors.append(f"Product import failed for row {row!r}: {exc}")
        else:
            for row in rows:
                try:
                    file_name = str(row.get("file_name") or "").strip()
                    if not file_name:
                        continue
                    image_bytes, _ = bulk_import_resolve_image_bytes(image_sources, file_name)
                    if image_bytes is None:
                        continue
                    if table_name == "product_images":
                        product = bulk_import_find_product(db, row)
                        if product is None:
                            continue
                        attach_product_image(product, file_name, image_bytes, parse_int_or_none(row.get("sort_order")))
                    elif table_name == "series_images":
                        series = bulk_import_find_series(db, row)
                        if series is None:
                            series = bulk_import_resolve_series_lookup(db, row, known_series)
                        if series is None:
                            continue
                        attach_series_image(series, file_name, image_bytes, parse_int_or_none(row.get("sort_order")))
                except Exception as exc:
                    report.errors.append(f"{table_name} import failed for row {row!r}: {exc}")

    if not dry_run:
        for product in touched_products.values():
            sync_graph_image(product, list(product.rpm_lines), list(product.efficiency_points))
            sync_product_image_files(product)
        for series in touched_series.values():
            sync_series_image_files(series)
        db.commit()
        for product in touched_products.values():
            db.refresh(product)
        for series in touched_series.values():
            db.refresh(series)
        notify_public_catalogue_cache_refresh()

    return report


def bulk_import_sources_from_uploads(
    files: list[UploadFile], manifest_json: str | None = None
) -> tuple[dict[str, list[dict]], dict[str, bytes], dict[str, dict]]:
    tables: dict[str, list[dict]] = {}
    image_sources: dict[str, bytes] = {}
    sheet_meta: dict[str, dict] = {}
    duplicates: set[str] = set()

    for upload in files:
        raw_name = normalize_bulk_import_source_name(upload.filename or "")
        suffix = Path(raw_name).suffix.lower()
        contents = upload.file.read() if hasattr(upload, "file") else None
        if contents is None:
            continue
        if suffix in {".xlsx", ".xlsm"}:
            workbook_tables, workbook_sheet_meta = load_bulk_import_workbook(raw_name, contents)
            for sheet_name, rows in workbook_tables.items():
                tables[sheet_name] = rows
            for sheet_name, meta in workbook_sheet_meta.items():
                sheet_meta[sheet_name] = meta
        elif suffix == ".csv":
            table_name = str(Path(raw_name).stem or raw_name).strip() or raw_name
            tables[table_name] = load_bulk_import_csv(raw_name, contents)
            sheet_meta[table_name] = {
                "sheet_name": table_name,
                "row_count": len(tables[table_name]),
                "raw_headers": list(tables[table_name][0].keys()) if tables[table_name] else [],
                "normalized_headers": list(tables[table_name][0].keys()) if tables[table_name] else [],
            }
        elif suffix in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tif", ".tiff"}:
            if raw_name in image_sources:
                duplicates.add(raw_name)
            image_sources[raw_name] = contents

    if manifest_json:
        tables["__manifest__"] = load_bulk_import_manifest(manifest_json)

    if duplicates:
        logger.warning("Bulk import received duplicate image filenames: %s", ", ".join(sorted(duplicates)))

    return tables, image_sources, sheet_meta


async def bulk_import_assets(
    dry_run: bool = False,
    downsample_imported_curves: bool = True,
    downsample_point_count: int = 5,
    manifest_json: str | None = Form(None),
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    try:
        if not files:
            raise HTTPException(status_code=400, detail="Please upload at least one workbook, CSV, or image file.")

        # Read all uploads before any database writes so dry runs and imports share the same parsing path.
        for upload in files:
            await upload.seek(0)

        tables, image_sources, sheet_meta = bulk_import_sources_from_uploads(files, manifest_json=manifest_json)
        if not tables and not image_sources:
            raise HTTPException(status_code=400, detail="No supported workbook, CSV, or image files were provided.")

        return bulk_import_process_payloads(
            db,
            tables,
            image_sources,
            sheet_meta,
            dry_run,
            default_downsample_imported_curves=downsample_imported_curves,
            default_downsample_point_count=downsample_point_count,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Bulk import failed unexpectedly: %s", exc)
        raise HTTPException(status_code=500, detail="Bulk import failed unexpectedly.") from exc


def sync_templates_after_file_change():
    sync_template_registry_with_disk()


def copy_tree_into_directory(source_dir: Path, target_dir: Path):
    target_dir.mkdir(parents=True, exist_ok=True)
    shutil.copytree(source_dir, target_dir, dirs_exist_ok=True)


def template_collection_name(template_type: str) -> str:
    normalized = (template_type or "").strip().lower()
    if normalized not in {"product", "series", "product_type"}:
        raise HTTPException(status_code=400, detail="Template type must be 'product', 'series', or 'product_type'.")
    return f"{normalized}_templates"


def template_type_directory(template_type: str) -> Path:
    normalized = (template_type or "").strip().lower()
    if normalized not in {"product", "series", "product_type"}:
        raise HTTPException(status_code=400, detail="Template type must be 'product', 'series', or 'product_type'.")
    directory = TEMPLATES_DIR / normalized
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def infer_template_label_from_slug(template_type: str, slug: str) -> str:
    prefix = "Product" if template_type == "product" else "Series" if template_type == "series" else "Product Type"
    readable = slug.replace("_", " ").replace("-", " ").strip()
    readable = " ".join(word.capitalize() for word in readable.split()) or "Template"
    return f"{prefix} {readable}"


def sync_template_registry_with_disk() -> dict:
    registry = load_template_registry()
    synchronized: dict[str, list[dict]] = {"product_templates": [], "series_templates": [], "product_type_templates": []}

    for template_type in ("product", "series", "product_type"):
        collection_name = template_collection_name(template_type)
        existing_by_path = {}
        existing_by_id = {}
        for item in registry.get(collection_name, []):
            if not isinstance(item, dict):
                continue
            item_path = str(item.get("path") or "").strip()
            item_id = str(item.get("id") or "").strip()
            if item_path:
                existing_by_path[item_path] = dict(item)
            if item_id:
                existing_by_id[item_id] = dict(item)

        template_dir = template_type_directory(template_type)
        for child in sorted([entry for entry in template_dir.iterdir() if entry.is_dir()], key=lambda entry: entry.name):
            template_path = child / "template.html"
            if not template_path.is_file():
                continue

            relative_html_path = str(template_path.relative_to(TEMPLATES_DIR.parent))
            relative_css_path = child / "template.css"
            entry = (
                existing_by_path.get(relative_html_path)
                or existing_by_id.get(f"{template_type}-{sanitize_name(child.name)}")
                or {}
            )
            template_id = str(entry.get("id") or f"{template_type}-{sanitize_name(child.name)}").strip()
            label = str(entry.get("label") or infer_template_label_from_slug(template_type, child.name)).strip()
            stylesheet = (
                str(relative_css_path.relative_to(TEMPLATES_DIR.parent))
                if relative_css_path.is_file()
                else None
            )
            synchronized[collection_name].append(
                {
                    "id": template_id,
                    "label": label,
                    "type": template_type,
                    "path": relative_html_path,
                    "stylesheet": stylesheet,
                }
            )

    if synchronized != registry:
        save_template_registry(synchronized)
    return synchronized


def scaffold_blank_template(template_type: str, destination_dir: Path):
    if template_type == "product":
        html_content = """<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>{{product.model}} | Product PDF</title>
    <link rel="stylesheet" href="./template.css" />
  </head>
  <body>
    <main class="sheet">
      <h1>{{product.model}}</h1>
      <div>{{product.description1_html}}</div>
      <div>{{product.grouped_specs_table}}</div>
      <img src="{{product.graph_image_url}}" alt="{{product.model}} graph" />
    </main>
  </body>
</html>
"""
    elif template_type == "product_type":
        html_content = """<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>{{product_type.label}} | Product Type PDF</title>
    <link rel="stylesheet" href="./template.css" />
  </head>
  <body>
    <main class="sheet">
      <h1>{{product_type.label}}</h1>
      <div>{{product_type.contents_icon_url}}</div>
      <div>{{product_type.series_names_html}}</div>
      <div>{{product_type.series_legend_html}}</div>
      <div>{{product_type.series_groups_html}}</div>
    </main>
  </body>
</html>
"""
    else:
        html_content = """<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>{{series.name}} | Series PDF</title>
    <link rel="stylesheet" href="./template.css" />
  </head>
  <body>
    <main class="sheet">
      <h1>{{series.name}}</h1>
      <div>{{series.description1_html}}</div>
      <img src="{{series.graph_image_url}}" alt="{{series.name}} graph" />
      <div>{{series.products_summary_table}}</div>
    </main>
  </body>
</html>
"""
    css_content = """.sheet { font-family: Arial, sans-serif; padding: 24px; }\nimg { max-width: 100%; height: auto; }\n"""
    destination_dir.mkdir(parents=True, exist_ok=False)
    (destination_dir / "template.html").write_text(html_content, encoding="utf-8")
    (destination_dir / "template.css").write_text(css_content, encoding="utf-8")


def validate_template_id(template_id: str | None, template_type: str) -> str | None:
    normalized = (template_id or "").strip() or None
    if normalized is None:
        return None
    registry = sync_template_registry_with_disk()
    collection_name = template_collection_name(template_type)
    valid_ids = {
        str(item.get("id")).strip()
        for item in registry.get(collection_name, [])
        if isinstance(item, dict) and item.get("id")
    }
    if normalized not in valid_ids:
        raise HTTPException(status_code=400, detail=f"Unknown {template_type} template id: {normalized}")
    return normalized


def get_template_definition(template_id: str | None, template_type: str) -> dict | None:
    normalized = validate_template_id(template_id, template_type)
    if normalized is None:
        return None
    registry = sync_template_registry_with_disk()
    collection_name = template_collection_name(template_type)
    for item in registry.get(collection_name, []):
        if isinstance(item, dict) and str(item.get("id")).strip() == normalized:
            return item
    return None


def require_template_definition(template_id: str, template_type: str) -> dict:
    template_definition = get_template_definition(template_id, template_type)
    if template_definition is None:
        raise HTTPException(status_code=404, detail="Template not found.")
    return template_definition


def resolve_template_stylesheet_path(template_definition: dict, template_path: Path) -> Path | None:
    project_root = Path(__file__).resolve().parents[1]
    stylesheet_value = str(template_definition.get("stylesheet") or "").strip()
    candidate_paths: list[Path] = []
    if stylesheet_value:
        candidate_paths.append(project_root / stylesheet_value)
    candidate_paths.append(template_path.parent / "template.css")
    for candidate in candidate_paths:
        if candidate.is_file():
            return candidate
    return candidate_paths[0] if candidate_paths else None


def inline_template_stylesheet(html_template: str, stylesheet_text: str) -> str:
    replacement = f"<style>\n{stylesheet_text}\n</style>"
    pattern = re.compile(
        r"<link\b[^>]*rel=(['\"])stylesheet\1[^>]*href=(['\"])\.\/template\.css\2[^>]*\/?>",
        re.IGNORECASE,
    )
    return pattern.sub(replacement, html_template, count=1)


def find_chromium_binary() -> str:
    candidates = [
        os.getenv("CHROMIUM_BIN", "").strip(),
        "chromium",
        "chromium-browser",
        "google-chrome",
        "google-chrome-stable",
    ]
    for candidate in candidates:
        if not candidate:
            continue
        resolved = shutil.which(candidate)
        if resolved:
            return resolved
    raise RuntimeError("No Chromium-compatible browser was found for PDF rendering.")


def product_image_file_name(product: Product, index: int, extension: str) -> str:
    ext = extension.lower()
    if not ext.startswith("."):
        ext = f".{ext}"
    return f"pic_{product_slug(product)}_{index}{ext}"


def series_image_file_name(series: Series, index: int, extension: str) -> str:
    ext = extension.lower()
    if not ext.startswith("."):
        ext = f".{ext}"
    return f"series_pic_{series_slug(series)}_{index}{ext}"


def graph_file_name(product: Product) -> str:
    return f"graph_{product_slug(product)}.png"


def product_pdf_file_name(product: Product) -> str:
    return f"product_printed_{product_slug(product)}.pdf"


def product_pdf_path(product: Product) -> Path:
    return PRODUCT_PDFS_DIR / product_pdf_file_name(product)


def series_graph_file_name(series: Series) -> str:
    return f"series_graph_{series_slug(series)}.png"


def series_pdf_file_name(series: Series) -> str:
    return f"series_printed_{series_slug(series)}.pdf"


def series_graph_path(series: Series) -> Path:
    return SERIES_GRAPHS_DIR / series_graph_file_name(series)


def series_pdf_path(series: Series) -> Path:
    return SERIES_PDFS_DIR / series_pdf_file_name(series)


def product_type_pdf_file_name(product_type: ProductType) -> str:
    return f"product_type_printed_{sanitize_name(product_type.key or product_type.label or 'unknown')}.pdf"


def product_type_pdf_path(product_type: ProductType) -> Path:
    return PRODUCT_TYPE_PDFS_DIR / product_type_pdf_file_name(product_type)


def all_product_types_pdf_path() -> Path:
    return PRODUCT_TYPE_PDFS_DIR / ALL_PRODUCT_TYPES_PDF_FILE_NAME


def all_product_types_pdf_public_url(file_path: Path | None = None) -> str:
    target = file_path or all_product_types_pdf_path()
    try:
        version = str(target.stat().st_mtime_ns)
    except OSError:
        version = ""
    suffix = f"?v={version}" if version else ""
    return f"/api/public/media/all-product-types-pdf{suffix}"


def _normalize_media_relative_path(relative_path: str | None) -> Path:
    candidate = Path((relative_path or "").strip())
    if str(candidate).strip() in {"", "."}:
        raise HTTPException(status_code=400, detail="A valid file name is required.")
    if candidate.is_absolute() or any(part in {"..", ""} for part in candidate.parts):
        raise HTTPException(status_code=400, detail="Invalid file name.")
    return candidate


def product_image_directory(product_id: int) -> Path:
    directory = PRODUCT_IMAGES_DIR / f"product_{product_id}"
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def series_image_directory(series_id: int) -> Path:
    directory = SERIES_IMAGES_DIR / f"series_{series_id}"
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def product_image_target_path(product_id: int, file_name: str) -> Path:
    return product_image_directory(product_id) / _normalize_media_relative_path(file_name)


def series_image_target_path(series_id: int, file_name: str) -> Path:
    return series_image_directory(series_id) / _normalize_media_relative_path(file_name)


def product_image_path(product_id: int, file_name: str) -> Path:
    target_path = product_image_target_path(product_id, file_name)
    if target_path.exists():
        return target_path
    legacy_path = PRODUCT_IMAGES_DIR / _normalize_media_relative_path(file_name)
    return legacy_path if legacy_path.exists() else target_path


def series_image_path(series_id: int, file_name: str) -> Path:
    target_path = series_image_target_path(series_id, file_name)
    if target_path.exists():
        return target_path
    legacy_path = SERIES_IMAGES_DIR / _normalize_media_relative_path(file_name)
    return legacy_path if legacy_path.exists() else target_path


def associated_document_path(owner_type: str, owner_id: int, file_name: str) -> Path:
    return ASSOCIATED_DOCUMENTS_DIR / owner_type / str(owner_id) / _normalize_media_relative_path(file_name)


def associated_document_directory(owner_type: str, owner_id: int) -> Path:
    directory = ASSOCIATED_DOCUMENTS_DIR / owner_type / str(owner_id)
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def remove_file(path: str | os.PathLike | None):
    if not path:
        return
    try:
        Path(path).unlink(missing_ok=True)
    except OSError:
        pass


def normalize_color_value(value: str | None) -> str | None:
    if value is None:
        return None
    stripped = value.strip()
    return stripped or None


def protect_jinja_tokens(source: str, prefix: str = "template-save") -> tuple[str, list[dict[str, str]]]:
    tokens: list[dict[str, str]] = []

    def replacer(match: re.Match[str]) -> str:
        placeholder = f"__{prefix}_{len(tokens)}__"
        tokens.append({"placeholder": placeholder, "token": match.group(0)})
        return placeholder

    encoded = JINJA_PATTERN.sub(replacer, str(source or ""))
    return encoded, tokens


def restore_jinja_tokens(source: str, tokens: list[dict[str, str]] | None = None) -> str:
    result = str(source or "")
    for entry in tokens or []:
        result = result.replace(entry.get("placeholder", ""), entry.get("token", ""))
    return result


def split_template_document(html_content: str) -> tuple[str, str, str]:
    source = str(html_content or "")
    body_match = re.search(r"<body\b[^>]*>", source, re.IGNORECASE)
    closing_match = re.search(r"</body>", source, re.IGNORECASE)
    if not body_match or not closing_match:
        return "", source, ""

    prefix_end = body_match.end()
    return source[:prefix_end], source[prefix_end:closing_match.start()], source[closing_match.start():]


def _html_local_name(tag: str | None) -> str:
    if not tag:
        return ""
    if isinstance(tag, str) and "}" in tag:
        return tag.rsplit("}", 1)[-1]
    return str(tag)


def _format_html_node(node, depth: int, indent: str) -> list[str]:
    pad = indent * max(depth, 0)

    if node.tag is ET.Comment:
        return [f"{pad}<!--{node.text or ''}-->"]

    tag_name = _html_local_name(getattr(node, "tag", ""))
    attrs = "".join(
        f' {name}="{html.escape(str(value), quote=True)}"' for name, value in (node.attrib or {}).items()
    )
    self_closing = tag_name.lower() in {
        "area",
        "base",
        "br",
        "col",
        "embed",
        "hr",
        "img",
        "input",
        "link",
        "meta",
        "source",
        "track",
        "wbr",
    }
    child_nodes = list(node)
    meaningful_children = bool(child_nodes) or (node.text or "").strip() != "" or any(
        (child.tail or "").strip() if isinstance(getattr(child, "tag", None), str) else True
        for child in child_nodes
    )

    if self_closing:
        return [f"{pad}<{tag_name}{attrs} />"]

    if not meaningful_children:
        return [f"{pad}<{tag_name}{attrs}></{tag_name}>"]

    text_content = " ".join((node.text or "").split()).strip()
    if len(child_nodes) == 0 and text_content:
        return [f"{pad}<{tag_name}{attrs}>{html.escape(text_content)}</{tag_name}>"]

    lines = [f"{pad}<{tag_name}{attrs}>"]

    if text_content:
        lines.append(f"{indent * (depth + 1)}{html.escape(text_content)}")

    for child in child_nodes:
        lines.extend(_format_html_node(child, depth + 1, indent))
        tail_text = " ".join((child.tail or "").split()).strip()
        if tail_text:
            lines.append(f"{indent * (depth + 1)}{html.escape(tail_text)}")

    lines.append(f"{pad}</{tag_name}>")
    return lines


def format_html_source(html_content: str, indent: str = "  ") -> str:
    source = str(html_content or "").strip()
    if not source:
        return ""

    protected_source, tokens = protect_jinja_tokens(source, "format-html")
    body_open = re.search(r"<body\b[^>]*>", protected_source, re.IGNORECASE)
    body_close = re.search(r"</body>", protected_source, re.IGNORECASE)

    if body_open and body_close:
        prefix = protected_source[: body_open.end()]
        body_inner = protected_source[body_open.end() : body_close.start()]
        suffix = protected_source[body_close.start() :]
        fragment = html5lib.parseFragment(body_inner, treebuilder="etree")
        body_lines: list[str] = []
        if fragment.text and fragment.text.strip():
            body_lines.append(f"{indent}{' '.join(fragment.text.split()).strip()}")
        for child in list(fragment):
            body_lines.extend(_format_html_node(child, 1, indent))
            tail_text = " ".join((child.tail or "").split()).strip()
            if tail_text:
                body_lines.append(f"{indent}{html.escape(tail_text)}")
        formatted = prefix
        if body_lines:
            formatted += "\n" + "\n".join(body_lines) + "\n"
        formatted += suffix
        return restore_jinja_tokens(formatted, tokens).strip()

    if re.search(r"<!doctype|<html\b|<head\b|<body\b", protected_source, re.IGNORECASE):
        document = html5lib.parse(protected_source, treebuilder="etree")
        lines: list[str] = []
        if "<!doctype" in protected_source.lower():
            lines.append("<!DOCTYPE html>")
        lines.extend(_format_html_node(document, 0, indent))
        return restore_jinja_tokens("\n".join(lines).strip(), tokens)

    fragment = html5lib.parseFragment(protected_source, treebuilder="etree")
    lines: list[str] = []
    if fragment.text and fragment.text.strip():
        lines.append(" ".join(fragment.text.split()).strip())
    for child in list(fragment):
        lines.extend(_format_html_node(child, 0, indent))
        tail_text = " ".join((child.tail or "").split()).strip()
        if tail_text:
            lines.append(html.escape(tail_text))

    return restore_jinja_tokens("\n".join(lines).strip(), tokens)


def format_css_source(css_content: str) -> str:
    source = str(css_content or "").strip()
    if not source:
        return ""

    protected_source, tokens = protect_jinja_tokens(source, "format-css")
    indent = "  "
    result = ""
    depth = 0
    in_string: str | None = None
    in_comment = False

    def append_indent() -> None:
        nonlocal result
        result += indent * max(depth, 0)

    def trim_line_end() -> None:
        nonlocal result
        result = re.sub(r"[ \t]+$", "", result)

    index = 0
    while index < len(protected_source):
        char = protected_source[index]
        next_char = protected_source[index + 1] if index + 1 < len(protected_source) else ""

        if in_comment:
            result += char
            if char == "*" and next_char == "/":
                result += next_char
                in_comment = False
                result += "\n"
                append_indent()
                index += 2
                continue
            index += 1
            continue

        if in_string:
            result += char
            if char == "\\" and next_char:
                result += next_char
                index += 2
                continue
            if char == in_string:
                in_string = None
            index += 1
            continue

        if char == "/" and next_char == "*":
            trim_line_end()
            if result and not result.endswith("\n"):
                result += "\n"
            append_indent()
            result += "/*"
            in_comment = True
            index += 2
            continue

        if char in {'"', "'"}:
            result += char
            in_string = char
            index += 1
            continue

        if char == "{":
            trim_line_end()
            result += " {\n"
            depth += 1
            append_indent()
            index += 1
            continue

        if char == "}":
            trim_line_end()
            result = re.sub(r"\n[ \t]*$", "\n", result)
            if not result.endswith("\n"):
                result += "\n"
            depth = max(0, depth - 1)
            append_indent()
            result += "}\n\n"
            append_indent()
            index += 1
            continue

        if char == ";":
            result += ";\n"
            append_indent()
            index += 1
            continue

        if char in {"\n", "\r", "\t"}:
            if not result.endswith(" ") and not result.endswith("\n"):
                result += " "
            index += 1
            continue

        if char == " " and (not result or result.endswith("\n") or result.endswith(" ")):
            index += 1
            continue

        result += char
        index += 1

    formatted = "\n".join(line.rstrip() for line in result.splitlines())
    formatted = "\n".join(
        (
            f"{line.split(':', 1)[0].rstrip()}: {line.split(':', 1)[1].lstrip()}"
            if ":" in line and not line.lstrip().startswith("@") and not line.lstrip().startswith("/*")
            else line
        )
        for line in formatted.splitlines()
    )
    formatted = re.sub(r"\n{3,}", "\n\n", formatted).strip()
    return restore_jinja_tokens(formatted, tokens)


SERIES_TAB_FALLBACK_COLOR = "#64748b"
SERIES_TAB_TEXT_COLOR = "#ffffff"
SERIES_TAB_STRIP_WIDTH = 25
SERIES_TAB_OUTER_MARGIN_TOP = 2
SERIES_TAB_OUTER_MARGIN_BOTTOM = 2
SERIES_TAB_SLOT_GAP = 3
SERIES_TAB_SLOT_FILL_RATIO = 0.9
SERIES_TAB_MAX_LABEL_CHARS = 26
SERIES_TAB_CORNER_RADIUS = 4


def series_tab_color_for_identity(identity: int | str | None) -> str:
    if identity in (None, ""):
        return SERIES_TAB_FALLBACK_COLOR
    digest = hashlib.sha1(str(identity).encode("utf-8")).digest()
    hue = int.from_bytes(digest[:2], "big") / 65535.0
    saturation = 0.62 + (digest[2] / 255.0) * 0.18
    lightness = 0.44 + (digest[3] / 255.0) * 0.08

    import colorsys

    red, green, blue = colorsys.hls_to_rgb(hue, lightness, saturation)
    return "#{:02x}{:02x}{:02x}".format(int(red * 255), int(green * 255), int(blue * 255))


def _shorten_tab_label(label: str) -> str:
    value = " ".join((label or "").split()).strip()
    if len(value) <= SERIES_TAB_MAX_LABEL_CHARS:
        return value
    return value[: SERIES_TAB_MAX_LABEL_CHARS - 1].rstrip() + "…"


def _hex_to_rgb(value: str) -> tuple[float, float, float]:
    text = (value or "").strip().lstrip("#")
    if len(text) != 6:
        return (0.39, 0.47, 0.53)
    return tuple(int(text[index : index + 2], 16) / 255.0 for index in (0, 2, 4))


def _darken_hex_color(value: str, factor: float = 0.72) -> str:
    red, green, blue = _hex_to_rgb(value)
    scale = max(0.0, min(float(factor), 1.0))
    return "#{:02x}{:02x}{:02x}".format(
        int(red * 255 * scale),
        int(green * 255 * scale),
        int(blue * 255 * scale),
    )


def _series_tab_layout(page_height: float, tab_count: int) -> list[tuple[float, float]]:
    count = max(int(tab_count), 1)
    usable_height = max(page_height - SERIES_TAB_OUTER_MARGIN_TOP - SERIES_TAB_OUTER_MARGIN_BOTTOM, 0)
    total_slot_gap = SERIES_TAB_SLOT_GAP * max(count - 1, 0)
    slot_height = (usable_height - total_slot_gap) / count if count else usable_height
    slot_height = max(slot_height, 1)
    tab_height = max(slot_height * SERIES_TAB_SLOT_FILL_RATIO, 1)
    positions: list[tuple[float, float]] = []

    for index in range(count):
        slot_top = page_height - SERIES_TAB_OUTER_MARGIN_TOP - (index * (slot_height + SERIES_TAB_SLOT_GAP))
        y_position = slot_top - tab_height
        positions.append((y_position, tab_height))

    return positions


def render_pdf_from_html(html_content: str, output_path: Path, stylesheet_text: str | None = None) -> None:
    browser_binary = find_chromium_binary()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    render_started = time.perf_counter()
    logger.info(
        "[pdf-render] starting browser=%s html_bytes=%d css_bytes=%d images=%d tables=%d",
        browser_binary,
        len(html_content.encode("utf-8")),
        len((stylesheet_text or "").encode("utf-8")),
        html_content.lower().count("<img"),
        html_content.lower().count("<table"),
    )
    timeout_seconds = max(float(os.getenv("PDF_RENDER_TIMEOUT_SECONDS", "120")), 1.0)

    with tempfile.TemporaryDirectory(prefix="pdf-render-") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        html_path = temp_dir / "document.html"
        html_path.write_text(html_content, encoding="utf-8")
        if stylesheet_text is not None:
            (temp_dir / "template.css").write_text(stylesheet_text, encoding="utf-8")

        last_result: subprocess.CompletedProcess[str] | None = None
        for headless_flag in ("--headless", "--headless=new"):
            if output_path.exists():
                output_path.unlink()
            attempt_started = time.perf_counter()
            try:
                result = subprocess.run(
                    [
                        browser_binary,
                        headless_flag,
                        "--disable-gpu",
                        "--no-sandbox",
                        "--disable-dev-shm-usage",
                        "--allow-file-access-from-files",
                        "--print-to-pdf-no-header",
                        f"--print-to-pdf={output_path}",
                        html_path.as_uri(),
                    ],
                    capture_output=True,
                    text=True,
                    timeout=timeout_seconds,
                )
            except subprocess.TimeoutExpired as exc:
                logger.error(
                    "[pdf-render] Chromium timed out after %.1fs flag=%s total_elapsed=%.1fs",
                    timeout_seconds,
                    headless_flag,
                    time.perf_counter() - render_started,
                )
                raise RuntimeError(f"Chromium PDF rendering exceeded {timeout_seconds:g} seconds.") from exc
            logger.info(
                "[pdf-render] browser finished flag=%s elapsed=%.1fs exit_code=%s output_bytes=%d",
                headless_flag,
                time.perf_counter() - attempt_started,
                result.returncode,
                output_path.stat().st_size if output_path.is_file() else 0,
            )
            last_result = result
            if output_path.is_file() and output_path.stat().st_size > 0:
                logger.info("[pdf-render] completed total_elapsed=%.1fs", time.perf_counter() - render_started)
                return

        if last_result is None:
            raise RuntimeError("Chromium failed to render the PDF.")
        stderr = (last_result.stderr or "").strip()
        stdout = (last_result.stdout or "").strip()
        details = stderr or stdout or f"Chromium failed to render the PDF (exit code {last_result.returncode})."
        raise RuntimeError(details)


def merge_pdf_files(source_paths: list[Path], output_path: Path) -> None:
    started = time.perf_counter()
    writer = PdfWriter()
    for source_path in source_paths:
        reader = PdfReader(str(source_path))
        for page in reader.pages:
            writer.add_page(page)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    merged_page_count = len(writer.pages)
    with output_path.open("wb") as handle:
        writer.write(handle)
    logger.info(
        "[pdf-merge] sources=%d pages=%d output=%s bytes=%d elapsed=%.1fs",
        len(source_paths),
        merged_page_count,
        output_path,
        output_path.stat().st_size,
        time.perf_counter() - started,
    )


def _pdf_first_page_size(pdf_path: Path) -> tuple[float, float]:
    reader = PdfReader(str(pdf_path))
    if not reader.pages:
        return float(A4[0]), float(A4[1])
    first_page = reader.pages[0]
    return float(first_page.mediabox.width), float(first_page.mediabox.height)


def _append_pdf_pages(writer: PdfWriter, pdf_path: Path) -> int:
    reader = PdfReader(str(pdf_path))
    for page in reader.pages:
        writer.add_page(page)
    return len(reader.pages)


def _build_pdf_overlay(page_width: float, page_height: float, page_number: int, tabs: list[dict]) -> PdfReader:
    from io import BytesIO

    buffer = BytesIO()
    overlay = canvas.Canvas(buffer, pagesize=(page_width, page_height))
    tab_items = tabs or []
    tab_x = page_width - SERIES_TAB_STRIP_WIDTH
    tab_layout = _series_tab_layout(page_height, len(tab_items))

    for index, tab in enumerate(tab_items):
        tab_label = str(tab.get("tab_label") or "")
        tab_color = str(tab.get("tab_color") or SERIES_TAB_FALLBACK_COLOR)
        tab_opacity = float(tab.get("tab_opacity") if tab.get("tab_opacity") is not None else 1.0)
        y_position, tab_height = tab_layout[index] if index < len(tab_layout) else (0.0, page_height)
        red, green, blue = _hex_to_rgb(tab_color)
        fill_alpha = max(0.05, min(1.0, tab_opacity))
        fill_color = Color(red, green, blue, alpha=fill_alpha)
        border_color = Color(max(0.0, red * 0.72), max(0.0, green * 0.72), max(0.0, blue * 0.72), alpha=min(0.9, fill_alpha))
        overlay.saveState()
        overlay.setFillColor(fill_color)
        overlay.setStrokeColor(border_color)
        overlay.setLineWidth(0.5)
        overlay.roundRect(tab_x, y_position, SERIES_TAB_STRIP_WIDTH, tab_height, SERIES_TAB_CORNER_RADIUS, stroke=1, fill=1)
        overlay.restoreState()

        overlay.saveState()
        overlay.translate(page_width - (SERIES_TAB_STRIP_WIDTH / 2), y_position + (tab_height / 2))
        overlay.rotate(90)
        overlay.setFillColor(HexColor(SERIES_TAB_TEXT_COLOR))
        overlay.setFont("Helvetica-Bold", 8.25)
        overlay.drawCentredString(0, -3, _shorten_tab_label(tab_label))
        overlay.restoreState()

    overlay.setFillColor(HexColor("#000000"))
    overlay.setFont("Helvetica", 9)
    overlay.drawRightString(page_width - SERIES_TAB_STRIP_WIDTH - 10, 10, str(page_number))
    overlay.showPage()
    overlay.save()
    buffer.seek(0)
    return PdfReader(buffer)


def stamp_pdf_file(input_path: Path, output_path: Path, page_decorations: list[dict]) -> None:
    started = time.perf_counter()
    reader = PdfReader(str(input_path))
    writer = PdfWriter()
    for index, page in enumerate(reader.pages):
        decoration = page_decorations[index] if index < len(page_decorations) else {}
        page_width = float(page.mediabox.width)
        page_height = float(page.mediabox.height)
        overlay_tabs = decoration.get("tabs")
        if overlay_tabs is None:
            overlay_tabs = []
            tab_label = str(decoration.get("tab_label") or "")
            tab_color = decoration.get("tab_color")
            tab_opacity = decoration.get("tab_opacity")
            if tab_label or tab_color is not None or tab_opacity is not None:
                overlay_tabs = [
                    {
                        "tab_label": tab_label,
                        "tab_color": str(tab_color or SERIES_TAB_FALLBACK_COLOR),
                        "tab_opacity": float(tab_opacity if tab_opacity is not None else 1.0),
                    }
                ]
        overlay_reader = _build_pdf_overlay(page_width, page_height, index + 1, list(overlay_tabs))
        page.merge_page(overlay_reader.pages[0])
        writer.add_page(page)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("wb") as handle:
        writer.write(handle)
    logger.info(
        "[pdf-stamp] input=%s pages=%d output=%s bytes=%d elapsed=%.1fs",
        input_path,
        len(reader.pages),
        output_path,
        output_path.stat().st_size,
        time.perf_counter() - started,
    )


def pdf_page_count(pdf_path: Path) -> int:
    return len(PdfReader(str(pdf_path)).pages)


def get_product_type_by_key(db: Session, product_type_key: str | None) -> ProductType:
    desired_key = (product_type_key or "fan").strip() or "fan"
    product_type = db.query(ProductType).filter(ProductType.key == desired_key).first()
    if product_type is None:
        raise HTTPException(status_code=400, detail=f"Unknown product type: {desired_key}")
    return product_type


def get_series_by_id(db: Session, series_id: int | None) -> Series | None:
    if series_id is None:
        return None
    series = db.get(Series, series_id)
    if series is None:
        raise HTTPException(status_code=400, detail=f"Unknown series id: {series_id}")
    return series


def load_series_graph_ready_series(db: Session, series_id: int) -> Series:
    series = (
        db.query(Series)
        .options(
            joinedload(Series.product_type),
            selectinload(Series.series_images),
            selectinload(Series.products).selectinload(Product.rpm_lines).selectinload(RpmLine.points),
            selectinload(Series.products).selectinload(Product.parameter_groups).selectinload(ProductParameterGroup.parameters),
            selectinload(Series.products).selectinload(Product.efficiency_points),
        )
        .filter(Series.id == series_id)
        .first()
    )
    if not series:
        raise HTTPException(status_code=404, detail="Series not found")
    assign_series_product_counts([series], load_series_product_counts(db, [series.id]))
    return series


def series_response_with_graph_payload(series: Series) -> SeriesResponse:
    response = SeriesResponse.model_validate(series, from_attributes=True)
    return response.model_copy(
        update={
            "performance_table_html": render_series_performance_table_html(series),
            "series_graph_payload": build_series_graph_payload(series),
        }
    )


def sync_product_series(product: Product, series: Series | None) -> None:
    product.series = series
    product.series_id = series.id if series else None
    product.series_name = series.name if series else None


def ensure_series_tab_color(db: Session, series: Series) -> str:
    existing_color = (series.series_tab_color or "").strip()
    if existing_color:
        return existing_color

    used_colors = {
        str(value[0]).strip().lower()
        for value in db.query(Series.series_tab_color)
        .filter(Series.series_tab_color.isnot(None), Series.id != series.id)
        .all()
        if value[0]
    }
    seed = series.id if series.id is not None else f"{series.product_type_id}:{series.name}"
    series.series_tab_color = allocate_series_tab_color(seed, used_colors)
    return series.series_tab_color


def sync_product_parameter_groups(product: Product, groups_payload: list[dict]):
    normalized_groups: list[dict] = []
    for group_index, group in enumerate(groups_payload or []):
        group_name = str(group.get("group_name", "")).strip()
        if not group_name:
            raise HTTPException(status_code=400, detail="Each parameter group must have a name.")

        seen_parameter_names: set[str] = set()
        normalized_parameters: list[dict] = []
        for parameter_index, parameter in enumerate(group.get("parameters") or []):
            parameter_name = str(parameter.get("parameter_name", "")).strip()
            if not parameter_name:
                raise HTTPException(status_code=400, detail=f"Each parameter in '{group_name}' must have a name.")
            if parameter_name.lower() in seen_parameter_names:
                raise HTTPException(
                    status_code=400,
                    detail=f"Parameter names must be unique within '{group_name}'.",
                )
            seen_parameter_names.add(parameter_name.lower())

            raw_value_string = parameter.get("value_string")
            value_string = None if raw_value_string is None else str(raw_value_string).strip() or None
            value_number = parameter.get("value_number")
            unit = None if parameter.get("unit") is None else str(parameter.get("unit")).strip() or None

            if value_string is not None and value_number is not None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Parameter '{parameter_name}' in '{group_name}' cannot have both text and numeric values.",
                )
            if value_string is not None and unit is not None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Parameter '{parameter_name}' in '{group_name}' cannot have a unit without a numeric value.",
                )

            normalized_parameters.append(
                {
                    "parameter_name": parameter_name,
                    "sort_order": parameter_index,
                    "value_string": value_string,
                    "value_number": None if value_number is None else float(value_number),
                    "unit": unit,
                }
            )

        normalized_groups.append(
            {
                "group_name": group_name,
                "sort_order": group_index,
                "parameters": normalized_parameters,
            }
        )

    product.parameter_groups.clear()
    for group_data in normalized_groups:
        group = ProductParameterGroup(
            group_name=group_data["group_name"],
            sort_order=group_data["sort_order"],
        )
        for parameter_data in group_data["parameters"]:
            group.parameters.append(ProductParameter(**parameter_data))
        product.parameter_groups.append(group)


def get_or_create_app_settings(db: Session) -> AppSettings:
    settings = db.get(AppSettings, 1)
    if settings is None:
        settings = AppSettings(id=1)
        db.add(settings)
        db.flush()
    return settings


def sync_product_image_files(product: Product):
    ordered_images = sorted(product.product_images, key=lambda image: (image.sort_order, image.id))
    temp_paths = {}

    for image in ordered_images:
        current_path = product_image_path(product.id, image.file_name)
        if current_path.exists():
            temp_path = product_image_target_path(product.id, f"tmp_{image.id}_{Path(image.file_name).name}")
            temp_path.parent.mkdir(parents=True, exist_ok=True)
            current_path.rename(temp_path)
            temp_paths[image.id] = temp_path

    for index, image in enumerate(ordered_images, start=1):
        suffix = Path(image.file_name).suffix or ".jpg"
        final_name = product_image_file_name(product, index, suffix)
        final_path = product_image_target_path(product.id, final_name)
        temp_path = temp_paths.get(image.id)
        if temp_path and temp_path.exists():
            final_path.parent.mkdir(parents=True, exist_ok=True)
            if final_path.exists():
                final_path.unlink()
            temp_path.rename(final_path)
        image.file_name = final_name
        image.sort_order = index - 1


def sync_series_image_files(series: Series):
    ordered_images = sorted(series.series_images, key=lambda image: (image.sort_order, image.id))
    temp_paths = {}

    for image in ordered_images:
        current_path = series_image_path(series.id, image.file_name)
        if current_path.exists():
            temp_path = series_image_target_path(series.id, f"tmp_{image.id}_{Path(image.file_name).name}")
            temp_path.parent.mkdir(parents=True, exist_ok=True)
            current_path.rename(temp_path)
            temp_paths[image.id] = temp_path

    for index, image in enumerate(ordered_images, start=1):
        suffix = Path(image.file_name).suffix or ".jpg"
        final_name = series_image_file_name(series, index, suffix)
        final_path = series_image_target_path(series.id, final_name)
        temp_path = temp_paths.get(image.id)
        if temp_path and temp_path.exists():
            final_path.parent.mkdir(parents=True, exist_ok=True)
            if final_path.exists():
                final_path.unlink()
            temp_path.rename(final_path)
        image.file_name = final_name
        image.sort_order = index - 1


def render_richtext_html(value: str | None) -> str:
    return value or ""


def remove_empty_richtext_sections(rendered: str) -> str:
    """Remove PDF template blocks whose optional rich-text value is empty."""
    patterns = (
        re.compile(
            r'<div\b[^>]*class="[^"]*\bcopy-section\b[^"]*"[^>]*>'
            r'(?:(?!<div\b[^>]*class="[^"]*\bcopy-section\b).)*?'
            r'<div class="(?:richtext|compact-richtext)">\s*</div>\s*</div>',
            re.DOTALL,
        ),
        re.compile(
            r'<section\b[^>]*class="[^"]*\b(?:panel|spec-card)\b[^"]*"[^>]*>'
            r'(?:(?!<section\b[^>]*class="[^"]*\b(?:panel|spec-card)\b).)*?'
            r'<div class="(?:richtext|compact-richtext)">\s*</div>.*?</section>',
            re.DOTALL,
        ),
    )
    previous = None
    while previous != rendered:
        previous = rendered
        for pattern in patterns:
            rendered = pattern.sub("", rendered)
    return rendered


def render_grouped_specs_table(product: Product) -> str:
    groups = sorted(product.parameter_groups, key=lambda group: (group.sort_order, group.id))
    if not groups:
        return '<p class="placeholder">No grouped specifications available.</p>'

    sections: list[str] = []
    for group in groups:
        rows: list[str] = []
        for parameter in sorted(group.parameters, key=lambda item: (item.sort_order, item.id)):
            if parameter.value_string not in {None, ""}:
                value_html = html.escape(parameter.value_string)
            elif parameter.value_number is not None:
                number_value = f"{parameter.value_number:g}"
                value_html = html.escape(f"{number_value} {parameter.unit}".strip())
            else:
                value_html = "—"
            rows.append(
                "<tr>"
                f"<th>{html.escape(parameter.parameter_name)}</th>"
                f"<td>{value_html}</td>"
                "</tr>"
            )

        sections.append(
            '<section class="spec-group">'
            f"<h3>{html.escape(group.group_name)}</h3>"
            '<table class="spec-table"><tbody>'
            + "".join(rows)
            + "</tbody></table></section>"
        )

    return "".join(sections)


def render_grouped_specs_group_html(product: Product, group_name: str) -> str:
    rows_html = render_grouped_specs_rows_html(product, group_name)
    if not rows_html:
        return f'<p class="placeholder">No {html.escape(group_name.lower())} grouped specifications available.</p>'

    return '<div class="spec-list">' + rows_html + "</div>"


def render_grouped_specs_rows_html(product: Product, group_name: str) -> str:
    target_slug = template_token_slug(group_name)
    matching_groups = [
        group
        for group in sorted(product.parameter_groups, key=lambda item: (item.sort_order, item.id))
        if template_token_slug(group.group_name or "") == target_slug
    ]
    if not matching_groups:
        return f'<p class="placeholder">No {html.escape(group_name.lower())} grouped specifications available.</p>'

    rows: list[str] = []
    for group in matching_groups:
        for parameter in sorted(group.parameters, key=lambda item: (item.sort_order, item.id)):
            if parameter.value_string not in {None, ""}:
                value_html = html.escape(parameter.value_string)
            elif parameter.value_number is not None:
                number_value = f"{parameter.value_number:g}"
                value_html = html.escape(f"{number_value} {parameter.unit}".strip())
            else:
                value_html = "—"
            rows.append(
                "<div class=\"spec-list__row\">"
                f"<dt>{html.escape(parameter.parameter_name)}</dt>"
                f"<dd>{value_html}</dd>"
                "</div>"
            )

    if not rows:
        return f'<p class="placeholder">No {html.escape(group_name.lower())} grouped specifications available.</p>'

    return "".join(rows)


def render_grouped_specs_table_for_group(product: Product, group_name: str) -> str:
    target_slug = template_token_slug(group_name)
    matching_groups = [
        group
        for group in sorted(product.parameter_groups, key=lambda item: (item.sort_order, item.id))
        if template_token_slug(group.group_name or "") == target_slug
    ]
    if not matching_groups:
        return f'<p class="placeholder">No {html.escape(group_name.lower())} grouped specifications available.</p>'

    rows: list[str] = []
    for group in matching_groups:
        for parameter in sorted(group.parameters, key=lambda item: (item.sort_order, item.id)):
            if parameter.value_string not in {None, ""}:
                value_html = html.escape(parameter.value_string)
            elif parameter.value_number is not None:
                number_value = f"{parameter.value_number:g}"
                value_html = html.escape(f"{number_value} {parameter.unit}".strip())
            else:
                value_html = "—"
            rows.append(
                "<tr>"
                f"<th>{html.escape(parameter.parameter_name)}</th>"
                f"<td>{value_html}</td>"
                "</tr>"
            )

    if not rows:
        return f'<p class="placeholder">No {html.escape(group_name.lower())} grouped specifications available.</p>'

    return '<table class="spec-table"><tbody>' + "".join(rows) + "</tbody></table>"


def format_parameter_value(parameter: ProductParameter) -> str:
    if parameter.value_string not in {None, ""}:
        return parameter.value_string
    if parameter.value_number is not None:
        number_value = f"{parameter.value_number:g}"
        return f"{number_value} {parameter.unit}".strip()
    return ""


def build_grouped_spec_token_map(product: Product) -> dict[str, str]:
    replacements: dict[str, str] = {}
    groups = sorted(product.parameter_groups, key=lambda group: (group.sort_order, group.id))
    for group in groups:
        group_key = template_token_slug(group.group_name or "")
        if not group_key:
            continue
        for parameter in sorted(group.parameters, key=lambda item: (item.sort_order, item.id)):
            parameter_key = template_token_slug(parameter.parameter_name or "")
            if not parameter_key:
                continue
            replacements[f"{{{{spec.{group_key}.{parameter_key}}}}}"] = html.escape(format_parameter_value(parameter))
    return replacements


def build_grouped_spec_group_token_map(product: Product) -> dict[str, str]:
    replacements: dict[str, str] = {}
    for group_name in ("impeller", "motor", "fan", "main"):
        rows_html = render_grouped_specs_rows_html(product, group_name)
        replacements[f"{{{{product.grouped_specs_{group_name}_html}}}}"] = (
            '<div class="spec-list">' + rows_html + "</div>"
            if rows_html.startswith("<div class=\"spec-list__row\">")
            else rows_html
        )
        replacements[f"{{{{product.grouped_specs_{group_name}_table}}}}"] = render_grouped_specs_table_for_group(product, group_name)
    return replacements


def render_grouped_specs_cards(product: Product, excluded_group_names: set[str] | None = None) -> str:
    excluded_group_names = excluded_group_names or set()
    groups = sorted(product.parameter_groups, key=lambda group: (group.sort_order, group.id))
    if not groups:
        return '<p class="placeholder">No grouped specifications available.</p>'

    cards: list[str] = []
    for group in groups:
        group_key = template_token_slug(group.group_name or "")
        if group_key in excluded_group_names:
            continue
        rows: list[str] = []
        for parameter in sorted(group.parameters, key=lambda item: (item.sort_order, item.id)):
            if parameter.value_string not in {None, ""}:
                value_html = html.escape(parameter.value_string)
            elif parameter.value_number is not None:
                number_value = f"{parameter.value_number:g}"
                value_html = html.escape(f"{number_value} {parameter.unit}".strip())
            else:
                value_html = "—"
            rows.append(
                "<div class=\"spec-list__row\">"
                f"<dt>{html.escape(parameter.parameter_name)}</dt>"
                f"<dd>{value_html}</dd>"
                "</div>"
            )
        if not rows:
            rows.append('<p class="placeholder">No specifications available.</p>')

        group_class = f"grouped-spec-group--{group_key}" if group_key else "grouped-spec-group--custom"
        cards.append(
            '<section class="spec-card">'
            f'<h2 class="spec-card__title">{html.escape(group.group_name)}</h2>'
            f'<div class="grouped-spec-group {group_class}">'
            + (
                '<dl class="spec-list">'
                + "".join(rows)
                + "</dl>"
                if rows and rows[0].startswith("<div class=\"spec-list__row\">")
                else "".join(rows)
            )
            + "</div></section>"
        )

    if not cards:
        return '<p class="placeholder">No grouped specifications available.</p>'

    return "".join(cards)


def render_fan_acoustic_table(product: Product) -> str:
    table = product.fan_acoustic_table or {}
    if product.product_type_key != "fan":
        return ""
    sound_power_columns = table.get("sound_power_columns") or []
    if not sound_power_columns:
        sound_power_columns = ["63", "125", "250", "500", "1k", "2k", "4k", "8k"]
    rows = table.get("rows") or []

    def format_numeric(value):
        if value is None or value == "":
            return ""
        try:
            return html.escape(f"{float(value):g}")
        except (TypeError, ValueError):
            return html.escape(str(value))

    header_cells = [
        '<th rowspan="2" class="fan-acoustic-table__cell fan-acoustic-table__primary-heading fan-acoustic-table__cell--speed">Speed (rpm)</th>',
        '<th rowspan="2" class="fan-acoustic-table__cell fan-acoustic-table__primary-heading fan-acoustic-table__cell--peak-pressure">Peak Pressure (Pa)</th>',
        '<th rowspan="2" class="fan-acoustic-table__cell fan-acoustic-table__primary-heading fan-acoustic-table__cell--peak-power">Peak Power (kW)</th>',
        '<th rowspan="2" class="fan-acoustic-table__cell fan-acoustic-table__primary-heading fan-acoustic-table__cell--running-frequency">Running Frequency (Hz)</th>',
        '<th rowspan="2" class="fan-acoustic-table__cell fan-acoustic-table__primary-heading fan-acoustic-table__cell--sound-pressure">Sound Pressure Level (dB) @ 3 meters</th>',
    ]

    body_rows: list[str] = []
    if rows:
        for row in rows:
            sound_power_levels = row.get("sound_power_levels") or {}
            body_rows.append(
                "<tr>"
                f'<td class="fan-acoustic-table__cell fan-acoustic-table__primary-cell fan-acoustic-table__cell--speed">{format_numeric(row.get("speed_rpm"))}</td>'
                f'<td class="fan-acoustic-table__cell fan-acoustic-table__primary-cell fan-acoustic-table__cell--peak-pressure">{format_numeric(row.get("peak_pressure_pa"))}</td>'
                f'<td class="fan-acoustic-table__cell fan-acoustic-table__primary-cell fan-acoustic-table__cell--peak-power">{format_numeric(row.get("peak_power_kw"))}</td>'
                f'<td class="fan-acoustic-table__cell fan-acoustic-table__primary-cell fan-acoustic-table__cell--running-frequency">{format_numeric(row.get("running_frequency_hz"))}</td>'
                f'<td class="fan-acoustic-table__cell fan-acoustic-table__primary-cell fan-acoustic-table__cell--sound-pressure">{format_numeric(row.get("sound_pressure_db_3m"))}</td>'
                + "".join(
                    f'<td class="fan-acoustic-table__cell fan-acoustic-table__sound-power-cell fan-acoustic-table__cell--sound-power">{format_numeric(sound_power_levels.get(column))}</td>'
                    for column in sound_power_columns
                )
                + "</tr>"
            )
    else:
        body_rows.append(
            "<tr>"
            f'<td colspan="{5 + len(sound_power_columns)}">No fan acoustic rows available.</td>'
            "</tr>"
        )

    return (
        '<table class="fan-acoustic-table">'
        + '<colgroup>'
        + '<col class="fan-acoustic-table__col fan-acoustic-table__col--speed" />'
        + '<col class="fan-acoustic-table__col fan-acoustic-table__col--peak-pressure" />'
        + '<col class="fan-acoustic-table__col fan-acoustic-table__col--peak-power" />'
        + '<col class="fan-acoustic-table__col fan-acoustic-table__col--running-frequency" />'
        + '<col class="fan-acoustic-table__col fan-acoustic-table__col--sound-pressure" />'
        + "".join(
            '<col class="fan-acoustic-table__col fan-acoustic-table__col--sound-power" />'
            for _ in sound_power_columns
        )
        + '</colgroup>'
        + "<thead>"
        + "<tr>"
        + "".join(header_cells)
        + f'<th colspan="{len(sound_power_columns)}" class="fan-acoustic-table__cell fan-acoustic-table__sound-power-band-heading fan-acoustic-table__band-heading">Sound (Hz) Power Level (dB) SWL re 1pw</th>'
        + "</tr>"
        + "<tr>"
        + "".join(
            f'<th class="fan-acoustic-table__cell fan-acoustic-table__sound-power-subhead fan-acoustic-table__subhead">{html.escape(str(column))}</th>'
            for column in sound_power_columns
        )
        + "</tr></thead><tbody>"
        + "".join(body_rows)
        + "</tbody></table>"
    )


def _renderable_product_images(product: Product) -> list[tuple[int, ProductImage, Path]]:
    ordered_images = sorted(product.product_images, key=lambda image: (image.sort_order, image.id))
    renderable_images: list[tuple[int, ProductImage, Path]] = []
    for index, image in enumerate(ordered_images, start=1):
        image_path = product_image_path(product.id, image.file_name)
        if image_path.is_file():
            renderable_images.append((index, image, image_path))
    return renderable_images


def render_product_image_html(product: Product, image_index: int, css_class: str, alt_text: str) -> str:
    renderable_images = _renderable_product_images(product)
    if image_index < 1 or len(renderable_images) < image_index:
        return '<p class="placeholder">No product image available.</p>'

    _, image, image_path = renderable_images[image_index - 1]

    return (
        f'<img src="{image_path.as_uri()}" alt="{html.escape(alt_text)}" class="{css_class}" />'
    )


def render_image_gallery_html(product: Product, start_index: int = 1) -> str:
    renderable_images = _renderable_product_images(product)
    if not renderable_images:
        return '<p class="placeholder">No product images available.</p>'

    items: list[str] = []
    for index, (_, image, image_path) in enumerate(renderable_images[start_index - 1 :], start=start_index):
        items.append(
            '<figure class="gallery-item">'
            f'<img src="{image_path.as_uri()}" alt="{html.escape(product.model or "")} image {index}" class="gallery-image" />'
            f'<figcaption>{html.escape("Primary image" if index == 1 else f"Image {index}")}</figcaption>'
            "</figure>"
        )

    return "".join(items) if items else '<p class="placeholder">No product images available.</p>'


def render_bottom_image_card_html(product: Product, image_index: int, alt_text: str) -> str:
    renderable_images = _renderable_product_images(product)
    if image_index < 1 or len(renderable_images) < image_index:
        return '<p class="placeholder">No product image available.</p>'

    _, _, image_path = renderable_images[image_index - 1]
    return (
        '<figure class="bottom-image-card">'
        f'<img src="{image_path.as_uri()}" alt="{html.escape(alt_text)}" class="bottom-image-card__img" />'
        "</figure>"
    )


def render_lower_visual_panels_html(product: Product) -> tuple[str, str]:
    renderable_images = _renderable_product_images(product)
    image_count = len(renderable_images)
    product_label = product.model or ""

    if image_count == 2:
        return (
            "lower-grid--single",
            (
                '<section class="bottom-image-stage bottom-image-stage--single">'
                f'{render_bottom_image_card_html(product, 2, f"{product_label} image 2")}'
                "</section>"
            ),
        )

    if image_count == 3:
        return (
            "lower-grid--pair",
            (
                '<section class="bottom-image-stage bottom-image-stage--pair">'
                f'{render_bottom_image_card_html(product, 2, f"{product_label} image 2")}'
                f'{render_bottom_image_card_html(product, 3, f"{product_label} image 3")}'
                "</section>"
            ),
        )

    return (
        "lower-grid--split",
        (
            '<section class="drawing-panel">'
            '<div class="drawing-panel__inner">'
            f'{render_product_image_html(product, 2, "drawing-image", f"{product_label} detailed view")}'
            "</div>"
            "</section>"
            '<section class="gallery-panel">'
            f'<div class="gallery-grid">{render_image_gallery_html(product, start_index=3)}</div>'
            "</section>"
        ),
    )


def resolve_pdf_logo_uri(project_root: Path, template_path: Path) -> str:
    logo_filename = "vent-tech-customer_site_logo_grey_bg.png"
    logo_candidates = [
        project_root / "templates" / logo_filename,
        template_path.parent / logo_filename,
    ]
    for logo_path in logo_candidates:
        if logo_path.is_file():
            return logo_path.resolve().as_uri()

    checked_paths = ", ".join(str(path) for path in logo_candidates)
    raise RuntimeError(f"PDF template logo is missing: {checked_paths}")


def build_product_pdf_html(product: Product) -> tuple[str, str]:
    template_id = product.printed_template_id or product.template_id
    template_definition = get_template_definition(template_id or product.template_id or "product-default", "product")
    if template_definition is None:
        raise RuntimeError("No product PDF template is configured.")

    project_root = Path(__file__).resolve().parents[1]
    template_path = project_root / template_definition["path"]
    if not template_path.is_file():
        raise RuntimeError(f"Product template file is missing: {template_path}")

    stylesheet_path = resolve_template_stylesheet_path(template_definition, template_path)
    logger.info(
        "[product_pdf:%s] template=%s html=%s stylesheet=%s",
        product.id,
        template_definition.get("id") or template_id or "product-default",
        template_path,
        stylesheet_path,
    )
    html_template = template_path.read_text(encoding="utf-8")
    stylesheet_text = stylesheet_path.read_text(encoding="utf-8") if stylesheet_path.is_file() else ""
    html_template = inline_template_stylesheet(html_template, stylesheet_text)

    primary_image_uri = ""
    if product.product_images:
        first_image = sorted(product.product_images, key=lambda img: (img.sort_order, img.id))[0]
        first_image_path = product_image_path(product.id, first_image.file_name)
        if first_image_path.is_file():
            primary_image_uri = first_image_path.as_uri()

    logo_uri = resolve_pdf_logo_uri(project_root, template_path)

    graph_image_uri = ""
    if product.graph_image_path:
        graph_path = Path(product.graph_image_path)
        if graph_path.is_file():
            graph_image_uri = graph_path.as_uri()

    lower_grid_layout_class, lower_visual_panels_html = render_lower_visual_panels_html(product)

    replacements = {
        "{{product.model}}": html.escape(product.model or ""),
        "{{product.product_type_label}}": html.escape(product.product_type_label or ""),
        "{{product.series_name}}": html.escape(product.series_name_value or ""),
        "{{product.description1_html}}": render_richtext_html(product.description1_html),
        "{{product.description2_html}}": render_richtext_html(product.description2_html),
        "{{product.description3_html}}": render_richtext_html(product.description3_html),
        "{{product.description_html}}": render_richtext_html(product.description1_html),
        "{{product.features_html}}": render_richtext_html(product.description2_html),
        "{{product.specifications_html}}": render_richtext_html(product.description3_html),
        "{{product.comments_html}}": render_richtext_html(product.comments_html),
        "{{product.grouped_specs_cards}}": render_grouped_specs_cards(product),
        "{{product.grouped_specs_table}}": render_grouped_specs_table(product),
        "{{product.grouped_specs_main_table}}": render_grouped_specs_table_for_group(product, "main"),
        "{{product.fan_acoustic_table}}": render_fan_acoustic_table(product),
        "{{product.image_gallery}}": render_image_gallery_html(product),
        "{{product.image_gallery_from_third}}": render_image_gallery_html(product, start_index=3),
        "{{product.secondary_product_image_html}}": render_product_image_html(
            product,
            image_index=2,
            css_class="drawing-image",
            alt_text=f"{product.model or ''} detailed view",
        ),
        "{{product.lower_grid_layout_class}}": lower_grid_layout_class,
        "{{product.lower_visual_panels_html}}": lower_visual_panels_html,
        "{{product.company_logo_url}}": logo_uri,
        "{{product.primary_product_image_url}}": primary_image_uri,
        "{{product.graph_image_url}}": graph_image_uri,
    }
    replacements.update(build_grouped_spec_group_token_map(product))
    replacements.update(build_grouped_spec_token_map(product))

    rendered = html_template
    for token, value in replacements.items():
        rendered = rendered.replace(token, value)
    return remove_empty_richtext_sections(rendered), stylesheet_text


def generate_product_pdf(
    product: Product,
    output_path: Path | None = None,
    progress_callback=None,
) -> Path:
    output_path = output_path or product_pdf_path(product)
    with tempfile.TemporaryDirectory(prefix="product-pdf-") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        base_path = temp_dir / f"product_printed_{product_slug(product)}_base.pdf"
        if progress_callback:
            progress_callback(f"Rendering product HTML for {product.model or 'product'}", 1, 3)
        html_content, stylesheet_text = build_product_pdf_html(product)
        if progress_callback:
            progress_callback(f"Writing product PDF for {product.model or 'product'}", 2, 3)
        render_pdf_from_html(html_content, base_path, stylesheet_text)
        if progress_callback:
            progress_callback(f"Copying product PDF for {product.model or 'product'}", 3, 3)
        shutil.copyfile(base_path, output_path)

    return output_path


def generate_product_pdfs(product: Product, progress_callback=None) -> Path:
    output_path = generate_product_pdf(product, progress_callback=progress_callback)
    remove_file(all_product_types_pdf_path())
    return output_path


def get_template_label(template_id: str | None, template_type: str) -> str:
    template_definition = get_template_definition(template_id, template_type)
    return str(template_definition.get("label")) if template_definition else "Default"


def series_graph_rule_label() -> str:
    return "Highest and lowest line from each product"


SERIES_PERFORMANCE_EXCLUDED_GROUP_NAMES = {
    "impeller type",
    "material",
    "motor finish",
}
SERIES_PERFORMANCE_COLUMN_LIMIT = 3


def _series_performance_candidate_columns(series: Series) -> list[tuple[str, str, str]]:
    ordered_products = sorted(series.products or [], key=lambda product: (product.model or "").lower())
    candidate_columns: list[tuple[str, str, str]] = []
    seen_columns: set[tuple[str, str]] = set()

    for product in ordered_products:
        for group in sorted(product.parameter_groups, key=lambda item: (item.sort_order, item.id)):
            group_name = (group.group_name or "").strip()
            if not group_name or group_name.casefold() in SERIES_PERFORMANCE_EXCLUDED_GROUP_NAMES:
                continue
            for parameter in sorted(group.parameters, key=lambda item: (item.sort_order, item.id)):
                parameter_name = (parameter.parameter_name or "").strip()
                key = (group_name, parameter_name)
                if not parameter_name or key in seen_columns:
                    continue
                seen_columns.add(key)
                candidate_columns.append((key[0], key[1], f"{key[0]}: {key[1]}"))
                if len(candidate_columns) >= SERIES_PERFORMANCE_COLUMN_LIMIT:
                    return candidate_columns
    return candidate_columns


def _series_type_2_performance_columns(series: Series) -> list[tuple[str, str, str]]:
    ordered_products = sorted(series.products or [], key=lambda product: (product.model or "").lower())
    selected_columns: list[tuple[str, str, str]] = []
    seen_columns: set[tuple[str, str]] = set()

    main_parameter_count = 0
    impeller_size_selected = False

    for product in ordered_products:
        for group in sorted(product.parameter_groups, key=lambda item: (item.sort_order, item.id)):
            group_name = (group.group_name or "").strip()
            if not group_name or group_name.casefold() in SERIES_PERFORMANCE_EXCLUDED_GROUP_NAMES:
                continue
            group_slug = template_token_slug(group_name)
            parameters = sorted(group.parameters, key=lambda item: (item.sort_order, item.id))

            for parameter in parameters:
                parameter_name = (parameter.parameter_name or "").strip()
                if not parameter_name:
                    continue

                key = (group_name, parameter_name)
                if key in seen_columns:
                    continue

                if group_slug == "main" and main_parameter_count < 2:
                    selected_columns.append((key[0], key[1], _series_type_2_performance_column_label(key[0], key[1])))
                    seen_columns.add(key)
                    main_parameter_count += 1
                    if len(selected_columns) >= SERIES_PERFORMANCE_COLUMN_LIMIT:
                        return selected_columns
                    continue

                if group_slug == "impeller" and parameter_name.casefold() == "size" and not impeller_size_selected:
                    selected_columns.append((key[0], key[1], _series_type_2_performance_column_label(key[0], key[1])))
                    seen_columns.add(key)
                    impeller_size_selected = True
                    if len(selected_columns) >= SERIES_PERFORMANCE_COLUMN_LIMIT:
                        return selected_columns

        if len(selected_columns) >= SERIES_PERFORMANCE_COLUMN_LIMIT:
            return selected_columns

    if len(selected_columns) < SERIES_PERFORMANCE_COLUMN_LIMIT:
        for column in _series_performance_candidate_columns(series):
            key = (column[0], column[1])
            if key in seen_columns:
                continue
            selected_columns.append((column[0], column[1], _series_type_2_performance_column_label(column[0], column[1])))
            seen_columns.add(key)
            if len(selected_columns) >= SERIES_PERFORMANCE_COLUMN_LIMIT:
                break

    return selected_columns[:SERIES_PERFORMANCE_COLUMN_LIMIT]


def _series_performance_value_map(product: Product) -> dict[tuple[str, str], str]:
    values: dict[tuple[str, str], str] = {}
    for group in product.parameter_groups:
        group_name = (group.group_name or "").strip()
        if not group_name or group_name.casefold() in SERIES_PERFORMANCE_EXCLUDED_GROUP_NAMES:
            continue
        for parameter in group.parameters:
            parameter_name = (parameter.parameter_name or "").strip()
            if not parameter_name:
                continue
            values[(group_name, parameter_name)] = format_parameter_value(parameter) or "—"
    return values


def _series_type_2_performance_column_label(group_name: str, parameter_name: str) -> str:
    if template_token_slug(group_name) == "main":
        return parameter_name
    if template_token_slug(group_name) == "impeller" and parameter_name.casefold() == "size":
        return "Impeller size"
    return f"{group_name}: {parameter_name}"


def _format_range(values: list[float], unit: str = "", precision: int = 0) -> str:
    if not values:
        return "—"
    minimum = min(values)
    maximum = max(values)
    if precision > 0:
        minimum_text = f"{minimum:.{precision}f}".rstrip("0").rstrip(".")
        maximum_text = f"{maximum:.{precision}f}".rstrip("0").rstrip(".")
    else:
        minimum_text = f"{minimum:g}"
        maximum_text = f"{maximum:g}"
    unit_text = f" {unit}" if unit else ""
    if minimum_text == maximum_text:
        return f"{minimum_text}{unit_text}"
    return f"{minimum_text} - {maximum_text}{unit_text}"


def _product_performance_ranges(product: Product) -> dict[str, str]:
    airflow_values: list[float] = []
    pressure_values: list[float] = []
    power_values: list[float] = []
    swl_values: list[float] = []

    for line in sorted(product.rpm_lines or [], key=lambda item: item.rpm):
        for point in getattr(line, "points", []) or []:
            if point.airflow is not None:
                airflow_values.append(float(point.airflow))
            if point.pressure is not None:
                pressure_values.append(float(point.pressure))

    fan_table = product.fan_acoustic_table or {}
    for row in fan_table.get("rows") or []:
        if not isinstance(row, dict):
            continue
        if row.get("peak_power_kw") not in {None, ""}:
            try:
                power_values.append(float(row["peak_power_kw"]))
            except (TypeError, ValueError):
                pass
        sound_power_levels = row.get("sound_power_levels") or {}
        if isinstance(sound_power_levels, dict):
            for value in sound_power_levels.values():
                if value in {None, ""}:
                    continue
                try:
                    swl_values.append(float(value))
                except (TypeError, ValueError):
                    continue

    return {
        "pressure_range": _format_range(pressure_values, "Pa"),
        "airflow_range": _format_range(airflow_values, "L/s"),
        "swl_range": _format_range(swl_values, "dB"),
        "power_range": _format_range(power_values, "kW", precision=2),
    }


def render_series_performance_table_rows(
    series: Series,
    selected_columns: list[tuple[str, str, str]] | None = None,
) -> str:
    ordered_products = sorted(series.products or [], key=lambda product: (product.model or "").lower())
    if not ordered_products:
        return '<tr><td colspan="8" class="placeholder">No products are linked to this series yet.</td></tr>'

    candidate_columns = (selected_columns or _series_performance_candidate_columns(series))[:SERIES_PERFORMANCE_COLUMN_LIMIT]
    while len(candidate_columns) < SERIES_PERFORMANCE_COLUMN_LIMIT:
        candidate_columns.append(("", "", ""))
    body_rows: list[str] = []
    for product in ordered_products:
        values = _series_performance_value_map(product)
        ranges = _product_performance_ranges(product)
        cells = [
            f"<td>{html.escape(product.model or '—')}</td>",
            *[
                f"<td>{html.escape(values.get((group_name, parameter_name), '—'))}</td>"
                for group_name, parameter_name, _ in candidate_columns
            ],
            f"<td>{html.escape(ranges['pressure_range'])}</td>",
            f"<td>{html.escape(ranges['airflow_range'])}</td>",
            f"<td>{html.escape(ranges['swl_range'])}</td>",
            f"<td>{html.escape(ranges['power_range'])}</td>",
        ]
        body_rows.append("<tr>" + "".join(cells) + "</tr>")

    return "".join(body_rows)


def render_series_image_html(
    series: Series,
    image_index: int,
    css_class: str,
    alt_text: str,
    placeholder_text: str,
    placeholder_class: str = "series-image-placeholder",
) -> str:
    ordered_images = sorted(series.series_images or [], key=lambda image: (image.sort_order, image.id))
    if image_index < 1 or len(ordered_images) < image_index:
        return f'<div class="{html.escape(placeholder_class)}">{html.escape(placeholder_text)}</div>'

    image = ordered_images[image_index - 1]
    image_path = series_image_path(series.id, image.file_name)
    if not image_path.is_file():
        return f'<div class="{html.escape(placeholder_class)}">{html.escape(placeholder_text)}</div>'

    return f'<img src="{html.escape(image_path.as_uri())}" alt="{html.escape(alt_text)}" class="{html.escape(css_class)}" />'


def _build_series_outline_image_path(series: Series, image_index: int, temp_dir: Path) -> Path | None:
    ordered_images = sorted(series.series_images or [], key=lambda image: (image.sort_order, image.id))
    if image_index < 1 or len(ordered_images) < image_index:
        return None

    image = ordered_images[image_index - 1]
    image_path = series_image_path(series.id, image.file_name)
    if not image_path.is_file():
        return None

    outline_path = temp_dir / f"series_{series_slug(series)}_outline_{image_index}.png"
    if outline_path.is_file():
        return outline_path

    try:
        with Image.open(image_path) as source:
            source = source.convert("RGB")
            # Cap the working size so the outline pass stays fast even for large uploads.
            scale = min(1.0, 1200 / max(source.size or (1, 1)))
            target_size = (
                max(1, int(source.width * scale)),
                max(1, int(source.height * scale)),
            )
            fitted = ImageOps.contain(source, target_size)
            grayscale = ImageOps.autocontrast(fitted.convert("L"))
            edges = grayscale.filter(ImageFilter.FIND_EDGES)
            edges = ImageOps.autocontrast(edges)
            edges = edges.point(lambda value: 255 if value > 18 else 0)
            edges = edges.filter(ImageFilter.MaxFilter(3))
            edges = edges.filter(ImageFilter.GaussianBlur(radius=0.45))

            outline = Image.new("RGBA", fitted.size, (255, 255, 255, 0))
            outline.putalpha(edges)
            outline.save(outline_path, format="PNG")
            return outline_path
    except Exception:
        return None


SERIES_COVER_PAGE_WIDTH_PX = 1800
SERIES_COVER_PAGE_HEIGHT_PX = 2546
SERIES_COVER_BACKGROUND_OPACITY = 26
SERIES_COVER_TITLE_STRIP_ALPHA = 232
SERIES_COVER_FONT_CANDIDATES = [
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    Path("/usr/share/fonts/truetype/space-grotesk/SpaceGrotesk-Bold.ttf"),
    Path("/usr/share/fonts/opentype/urw-base35/NimbusSans-Bold.otf"),
    Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc"),
]


def _load_series_cover_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for font_path in SERIES_COVER_FONT_CANDIDATES:
        if font_path.is_file():
            try:
                return ImageFont.truetype(str(font_path), size=size)
            except Exception:
                continue
    return ImageFont.load_default()


def _fit_cover_text_font(draw: ImageDraw.ImageDraw, text: str, max_width: int, start_size: int, min_size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for size in range(start_size, min_size - 1, -2):
        font = _load_series_cover_font(size)
        bbox = draw.textbbox((0, 0), text, font=font, spacing=0, align="center")
        if (bbox[2] - bbox[0]) <= max_width:
            return font
    return _load_series_cover_font(min_size)


def _draw_shadowed_centered_text(
    draw: ImageDraw.ImageDraw,
    center_x: int,
    center_y: int,
    text: str,
    font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
    fill: tuple[int, int, int, int],
    shadow_fill: tuple[int, int, int, int] = (0, 0, 0, 110),
) -> None:
    bbox = draw.textbbox((0, 0), text, font=font, spacing=0, align="center")
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    x = center_x - (text_width / 2)
    y = center_y - (text_height / 2)
    draw.text((x + 2, y + 2), text, font=font, fill=shadow_fill, spacing=0, align="center")
    draw.text((x, y), text, font=font, fill=fill, spacing=0, align="center")


def _series_cover_background(size: tuple[int, int]) -> Image.Image:
    width, height = size
    canvas = Image.new("RGBA", size, (18, 5, 8, 255))
    wash = Image.new("RGBA", size, (88, 12, 22, 115))
    return Image.alpha_composite(canvas, wash)


def _build_series_cover_background_image_path(series: Series, temp_dir: Path) -> Path | None:
    output_path = temp_dir / f"series_{series_slug(series)}_cover_background.png"
    if output_path.is_file():
        return output_path

    ordered_images = sorted(series.series_images or [], key=lambda image: (image.sort_order, image.id))
    if not ordered_images:
        return None

    image = ordered_images[0]
    image_path = series_image_path(series.id, image.file_name)
    if not image_path.is_file():
        return None

    try:
        with Image.open(image_path) as source:
            source = source.convert("RGB")
            fitted = ImageOps.fit(
                source,
                (SERIES_COVER_PAGE_WIDTH_PX, SERIES_COVER_PAGE_HEIGHT_PX),
                method=getattr(getattr(Image, "Resampling", Image), "LANCZOS"),
            )
            fitted = ImageOps.invert(ImageOps.autocontrast(fitted.convert("L"))).convert("RGBA")
            fitted = fitted.filter(ImageFilter.GaussianBlur(radius=1.0))
            fitted.putalpha(SERIES_COVER_BACKGROUND_OPACITY)

            canvas_image = Image.new("RGBA", (SERIES_COVER_PAGE_WIDTH_PX, SERIES_COVER_PAGE_HEIGHT_PX), (18, 4, 8, 255))
            canvas_image.alpha_composite(fitted)

            red_wash = Image.new("RGBA", (SERIES_COVER_PAGE_WIDTH_PX, SERIES_COVER_PAGE_HEIGHT_PX), (92, 10, 22, 120))
            canvas_image = Image.alpha_composite(canvas_image, red_wash)

            canvas_image.save(output_path, format="PNG")
            return output_path
    except Exception:
        return None


def _build_series_cover_spread_background_image_path(series: Series, temp_dir: Path) -> Path | None:
    output_path = temp_dir / f"series_{series_slug(series)}_cover_background_spread.png"
    if output_path.is_file():
        return output_path

    ordered_images = sorted(series.series_images or [], key=lambda image: (image.sort_order, image.id))
    if not ordered_images:
        return None

    image = ordered_images[0]
    image_path = series_image_path(series.id, image.file_name)
    if not image_path.is_file():
        return None

    try:
        with Image.open(image_path) as source:
            source = source.convert("RGB")
            fitted = ImageOps.fit(
                source,
                (SERIES_COVER_PAGE_WIDTH_PX * 2, SERIES_COVER_PAGE_HEIGHT_PX),
                method=getattr(getattr(Image, "Resampling", Image), "LANCZOS"),
            )
            fitted = ImageOps.invert(ImageOps.autocontrast(fitted.convert("L"))).convert("RGBA")
            fitted = fitted.filter(ImageFilter.GaussianBlur(radius=1.0))
            fitted.putalpha(SERIES_COVER_BACKGROUND_OPACITY)

            canvas_image = Image.new("RGBA", (SERIES_COVER_PAGE_WIDTH_PX * 2, SERIES_COVER_PAGE_HEIGHT_PX), (18, 4, 8, 255))
            canvas_image.alpha_composite(fitted)

            red_wash = Image.new("RGBA", (SERIES_COVER_PAGE_WIDTH_PX * 2, SERIES_COVER_PAGE_HEIGHT_PX), (92, 10, 22, 120))
            canvas_image = Image.alpha_composite(canvas_image, red_wash)

            canvas_image.save(output_path, format="PNG")
            return output_path
    except Exception:
        return None


def _hex_color_to_rgb(hex_color: str) -> tuple[int, int, int]:
    value = (hex_color or "").strip().lstrip("#")
    if len(value) == 3:
        value = "".join(ch * 2 for ch in value)
    if len(value) != 6:
        return 204, 16, 36
    try:
        return int(value[0:2], 16), int(value[2:4], 16), int(value[4:6], 16)
    except ValueError:
        return 204, 16, 36


def _build_series_cover_page_image_path(series: Series, image_index: int, temp_dir: Path, include_title: bool) -> Path | None:
    page_role = "left" if include_title else "right"
    output_path = temp_dir / f"series_{series_slug(series)}_cover_{page_role}_{image_index}_spread.png"
    if output_path.is_file():
        return output_path

    background_path = _build_series_cover_spread_background_image_path(series, temp_dir)
    title_text = series.name or ""

    try:
        if background_path is not None and background_path.is_file():
            with Image.open(background_path) as background_image:
                canvas_image = background_image.convert("RGBA").copy()
        else:
            canvas_image = _series_cover_background((SERIES_COVER_PAGE_WIDTH_PX * 2, SERIES_COVER_PAGE_HEIGHT_PX))

        crop_left = 0 if include_title else SERIES_COVER_PAGE_WIDTH_PX
        page_image = canvas_image.crop(
            (
                crop_left,
                0,
                crop_left + SERIES_COVER_PAGE_WIDTH_PX,
                SERIES_COVER_PAGE_HEIGHT_PX,
            )
        )
        draw = ImageDraw.Draw(page_image)

        if include_title:
            title_font = _fit_cover_text_font(draw, title_text, int(SERIES_COVER_PAGE_WIDTH_PX * 0.78), 132, 68) if title_text else _load_series_cover_font(112)
            title_bbox = draw.textbbox((0, 0), title_text, font=title_font, spacing=0, align="top") if title_text else (0, 0, 0, 0)
            title_width = max(0, title_bbox[2] - title_bbox[0])
            title_height = max(0, title_bbox[3] - title_bbox[1])
            box_padding_x = 22
            box_padding_y = 22
            box_width = min(
                int(SERIES_COVER_PAGE_WIDTH_PX * 0.86),
                max(title_width + (box_padding_x * 2), int(SERIES_COVER_PAGE_WIDTH_PX * 0.54)),
            )
            box_height = title_height + (box_padding_y * 2)
            box_left = (SERIES_COVER_PAGE_WIDTH_PX - box_width) // 2
            box_top = int(SERIES_COVER_PAGE_HEIGHT_PX * 0.16)
            box_fill = (*_hex_color_to_rgb(series.series_tab_color or SERIES_TAB_FALLBACK_COLOR), SERIES_COVER_TITLE_STRIP_ALPHA)
            draw.rounded_rectangle(
                (box_left, box_top, box_left + box_width, box_top + box_height),
                radius=18,
                fill=box_fill,
            )
            if title_text:
                title_x = (SERIES_COVER_PAGE_WIDTH_PX // 2) - (title_width // 2) - title_bbox[0]
                title_y = box_top + box_padding_y - title_bbox[1]
                shadow_fill = (0, 0, 0, 120)
                draw.text((title_x + 2, title_y + 2), title_text, font=title_font, fill=shadow_fill, spacing=0, align="top")
                draw.text((title_x, title_y), title_text, font=title_font, fill=(255, 255, 255, 255), spacing=0, align="top")

        page_image.save(output_path, format="PNG")
        return output_path
    except Exception:
        return None


def render_series_outline_image_html(
    series: Series,
    image_index: int,
    css_class: str,
    alt_text: str,
    placeholder_text: str,
    temp_dir: Path,
    placeholder_class: str = "series-image-placeholder",
) -> str:
    outline_path = _build_series_outline_image_path(series, image_index, temp_dir)
    if outline_path is None:
        return f'<div class="{html.escape(placeholder_class)}">{html.escape(placeholder_text)}</div>'
    return f'<img src="{html.escape(outline_path.as_uri())}" alt="{html.escape(alt_text)}" class="{html.escape(css_class)}" />'


def render_series_products_summary_table(series: Series) -> str:
    ordered_products = sorted(series.products or [], key=lambda product: (product.model or "").lower())
    if not ordered_products:
        return '<p class="placeholder">No products are linked to this series yet.</p>'

    candidate_columns: list[tuple[str, str, str]] = []
    seen_columns: set[tuple[str, str]] = set()
    for product in ordered_products:
        for group in sorted(product.parameter_groups, key=lambda item: (item.sort_order, item.id)):
            for parameter in sorted(group.parameters, key=lambda item: (item.sort_order, item.id)):
                key = (group.group_name or "", parameter.parameter_name or "")
                if not key[0] or not key[1] or key in seen_columns:
                    continue
                seen_columns.add(key)
                candidate_columns.append((key[0], key[1], f"{key[0]}: {key[1]}"))
                if len(candidate_columns) >= 6:
                    break
            if len(candidate_columns) >= 6:
                break
        if len(candidate_columns) >= 6:
            break

    def parameter_value_map(product: Product) -> dict[tuple[str, str], str]:
        values: dict[tuple[str, str], str] = {}
        for group in product.parameter_groups:
            for parameter in group.parameters:
                values[(group.group_name or "", parameter.parameter_name or "")] = format_parameter_value(parameter) or "—"
        return values

    header_cells = [
        "<th>Model</th>",
        *[f"<th>{html.escape(label)}</th>" for _, _, label in candidate_columns],
    ]

    body_rows: list[str] = []
    for product in ordered_products:
        values = parameter_value_map(product)
        data_cells = [
            f"<td>{html.escape(product.model or '—')}</td>",
            *[
                f"<td>{html.escape(values.get((group_name, parameter_name), '—'))}</td>"
                for group_name, parameter_name, _ in candidate_columns
            ],
        ]
        body_rows.append("<tr>" + "".join(data_cells) + "</tr>")

    return (
        '<table class="series-summary-table"><thead><tr>'
        + "".join(header_cells)
        + "</tr></thead><tbody>"
        + "".join(body_rows)
        + "</tbody></table>"
    )


def render_series_performance_table_html(series: Series) -> str:
    ordered_products = sorted(series.products or [], key=lambda product: (product.model or "").lower())
    if not ordered_products:
        return '<p class="performance-table__empty placeholder">No products are linked to this series yet.</p>'

    template_id = series.printed_template_id or series.template_id
    template_definition = get_template_definition(template_id or "series-default", "series")
    if template_definition and template_definition.get("id") == "series-series_type_2":
        performance_columns = _series_type_2_performance_columns(series)
    else:
        performance_columns = _series_performance_candidate_columns(series)

    performance_column_labels = [column[2] for column in performance_columns]
    while len(performance_column_labels) < SERIES_PERFORMANCE_COLUMN_LIMIT:
        performance_column_labels.append("—")

    table_rows = render_series_performance_table_rows(series, performance_columns)
    return (
        '<div class="performance-table">'
        '<table class="performance-table__table">'
        '<colgroup>'
        '<col class="performance-table__col performance-table__col--model" />'
        + "".join(
            '<col class="performance-table__col performance-table__col--spec" />'
            for _ in performance_column_labels
        )
        + '<col class="performance-table__col performance-table__col--range" />'
        + '<col class="performance-table__col performance-table__col--range" />'
        + '<col class="performance-table__col performance-table__col--range" />'
        + '<col class="performance-table__col performance-table__col--range" />'
        + '</colgroup>'
        "<thead><tr>"
        "<th>Model</th>"
        + "".join(f"<th>{html.escape(label)}</th>" for label in performance_column_labels)
        + "<th>Pressure Range</th>"
        + "<th>Airflow Range</th>"
        + "<th>SWL Range</th>"
        + "<th>Power Range</th>"
        + "</tr></thead>"
        + "<tbody>"
        + table_rows
        + "</tbody></table></div>"
    )


def build_series_graph_payload(series: Series) -> dict | None:
    product_type = series.product_type
    if not product_type or not product_type.supports_graph:
        return None

    synthetic_lines: list[dict] = []
    synthetic_points: list[dict] = []
    next_line_id = 1
    next_point_id = 1

    ordered_products = sorted(series.products or [], key=lambda product: (product.model or "").lower())
    for product in ordered_products:
        ordered_lines = sorted(product.rpm_lines or [], key=lambda line: (line.rpm, line.id))
        if not ordered_lines:
            continue
        product_color = _darken_hex_color(series_tab_color_for_identity(product.id), 0.72)

        selected_lines = [line for line in ordered_lines if getattr(line, "points", None)]
        if not selected_lines:
            continue
        if len(selected_lines) > 1:
            selected_lines = [selected_lines[0], selected_lines[-1]]

        for index, line in enumerate(selected_lines):
            display_label = (
                f"{product.model} low"
                if len(selected_lines) > 1 and index == 0
                else f"{product.model} high"
                if len(selected_lines) > 1
                else f"{product.model}"
            )
            synthetic_line_id = next_line_id
            next_line_id += 1
            synthetic_lines.append(
                {
                    "id": synthetic_line_id,
                    "product_id": product.id,
                    "rpm": synthetic_line_id,
                    "display_label": display_label,
                    "band_color": product_color,
                    "line_role": "low" if len(selected_lines) > 1 and index == 0 else "high",
                }
            )
            for point in sorted(line.points or [], key=lambda item: item.airflow):
                synthetic_points.append(
                    {
                        "id": next_point_id,
                        "product_id": product.id,
                        "rpm_line_id": synthetic_line_id,
                        "rpm": synthetic_line_id,
                        "airflow": point.airflow,
                        "pressure": point.pressure,
                    }
                )
                next_point_id += 1

    if not synthetic_points:
        return None

    return {
        "hasGraphData": True,
        "title": f"{series.name} Series Graph",
        "graphTitle": f"{series.name} Series Graph",
        "showRpmBandShading": False,
        "graphConfig": build_graph_config(product_type),
        "graphStyle": None,
        "rpmLines": synthetic_lines,
        "rpmPoints": synthetic_points,
        "efficiencyPoints": [],
    }


def series_has_graph_capable_line_data(series: Series) -> bool:
    return build_series_graph_payload(series) is not None


def generate_series_graph(series: Series) -> Path:
    payload = build_series_graph_payload(series)
    if payload is None:
        raise RuntimeError("No graph-capable products with line data are linked to this series.")

    final_path = series_graph_path(series)
    tmp_path = SERIES_GRAPHS_DIR / f"tmp_{series_graph_file_name(series)}"
    final_path.parent.mkdir(parents=True, exist_ok=True)
    if tmp_path.exists():
        tmp_path.unlink()

    result = subprocess.run(
        ["node", str(ECHARTS_RENDER_SCRIPT), str(tmp_path)],
        cwd=str(FRONTEND_DIR),
        input=json.dumps(payload),
        text=True,
        capture_output=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(
            "ECharts graph render failed: "
            + (result.stderr.strip() or result.stdout.strip() or f"exit code {result.returncode}")
        )

    shutil.move(tmp_path, final_path)
    return final_path


def build_series_pdf_html(series: Series, temp_dir: Path) -> tuple[str, str]:
    template_id = series.printed_template_id or series.template_id
    template_definition = get_template_definition(template_id or series.template_id or "series-default", "series")
    if template_definition is None:
        raise RuntimeError("No series PDF template is configured.")

    project_root = Path(__file__).resolve().parents[1]
    template_path = project_root / template_definition["path"]
    if not template_path.is_file():
        raise RuntimeError(f"Series template file is missing: {template_path}")

    stylesheet_path = resolve_template_stylesheet_path(template_definition, template_path)
    html_template = template_path.read_text(encoding="utf-8")
    stylesheet_text = stylesheet_path.read_text(encoding="utf-8") if stylesheet_path.is_file() else ""
    html_template = inline_template_stylesheet(html_template, stylesheet_text)

    graph_uri = ""
    graph_path = series_graph_path(series)
    if graph_path.is_file():
        graph_uri = graph_path.as_uri()
    series_graph_payload = build_series_graph_payload(series)

    performance_columns = _series_performance_candidate_columns(series)
    if template_definition.get("id") == "series-series_type_2":
        performance_columns = _series_type_2_performance_columns(series)
    performance_column_labels = [column[2] for column in performance_columns]
    while len(performance_column_labels) < SERIES_PERFORMANCE_COLUMN_LIMIT:
        performance_column_labels.append("—")

    logo_uri = resolve_pdf_logo_uri(project_root, template_path)

    left_cover_page_path = _build_series_cover_page_image_path(series, 1, temp_dir, include_title=True)
    right_cover_page_path = _build_series_cover_page_image_path(series, 2, temp_dir, include_title=False)

    replacements = {
        "{{series.name}}": html.escape(series.name or ""),
        "{{series.product_type_label}}": html.escape(series.product_type_label or ""),
        "{{series.series_tab_color}}": html.escape(series.series_tab_color or SERIES_TAB_FALLBACK_COLOR),
        "{{series.primary_series_image_html}}": render_series_image_html(
            series,
            image_index=1,
            css_class="series-image-card__image series-image-card__image--primary",
            alt_text=f"{series.name or ''} primary series image",
            placeholder_text="Series image unavailable",
        ),
        "{{series.secondary_series_image_html}}": render_series_image_html(
            series,
            image_index=2,
            css_class="series-image-card__image series-image-card__image--secondary",
            alt_text=f"{series.name or ''} secondary series image",
            placeholder_text="Series image unavailable",
        ),
        "{{series.cover_left_page_image_url}}": left_cover_page_path.as_uri() if left_cover_page_path else "",
        "{{series.cover_right_page_image_url}}": right_cover_page_path.as_uri() if right_cover_page_path else "",
        "{{series.description1_html}}": render_richtext_html(series.description1_html),
        "{{series.description2_html}}": render_richtext_html(series.description2_html),
        "{{series.description3_html}}": render_richtext_html(series.description3_html),
        "{{series.description4_html}}": render_richtext_html(series.description4_html),
        "{{series.comments_html}}": render_richtext_html(series.description4_html),
        "{{series.template_label}}": html.escape(get_template_label(template_id or series.template_id, "series")),
        "{{series.product_count}}": html.escape(str(series.product_count)),
        "{{series.graph_rule_label}}": html.escape(series_graph_rule_label()),
        "{{series.graph_image_url}}": graph_uri,
        "{{series.graph_payload_json}}": json_for_html_script(series_graph_payload),
        "{{series.performance_column_1_label}}": html.escape(performance_column_labels[0]),
        "{{series.performance_column_2_label}}": html.escape(performance_column_labels[1]),
        "{{series.performance_column_3_label}}": html.escape(performance_column_labels[2]),
        "{{series.performance_table_rows}}": render_series_performance_table_rows(series, performance_columns),
        "{{series.company_logo_url}}": logo_uri,
    }

    rendered = html_template
    for token, value in replacements.items():
        rendered = rendered.replace(token, value)
    return remove_empty_richtext_sections(rendered), stylesheet_text


def build_series_pdf_base(series: Series, temp_dir: Path, progress_callback=None) -> tuple[Path, int]:
    cover_base_path = temp_dir / f"series_printed_{series_slug(series)}_cover.pdf"
    cover_html, cover_stylesheet_text = build_series_pdf_html(series, temp_dir)
    if progress_callback:
        progress_callback(f"Rendering series cover for {series.name or 'series'}", 1, max(len(series.products or []) + 4, 5))
    render_pdf_from_html(cover_html, cover_base_path, cover_stylesheet_text)

    ordered_products = sorted(series.products or [], key=lambda item: (item.model or "").casefold())
    total_steps = max(len(ordered_products) + 4, 5)
    product_pdf_paths: list[Path] = []
    for index, product in enumerate(ordered_products, start=1):
        product_pdf = product_pdf_path(product)
        if not product_pdf.is_file():
            raise RuntimeError(
                f"Missing printed product PDF for {product.model or 'product'}: {product_pdf}. "
                "Generate the product PDFs first."
            )
        product_pdf_paths.append(product_pdf)
        if progress_callback:
            progress_callback(
                f"Reusing printed product {index} of {len(ordered_products)} for {series.name or 'series'}",
                index + 1,
                total_steps,
            )

    merged_base_path = temp_dir / f"series_printed_{series_slug(series)}_base.pdf"
    if progress_callback:
        progress_callback(f"Merging printed product PDFs for {series.name or 'series'}", len(ordered_products) + 2, total_steps)
    merge_pdf_files([cover_base_path, *product_pdf_paths], merged_base_path)
    page_count = pdf_page_count(merged_base_path)
    if page_count % 2 == 1:
        aligned_base_path = temp_dir / f"series_printed_{series_slug(series)}_aligned.pdf"
        if progress_callback:
            progress_callback(f"Adding buffer page for {series.name or 'series'}", len(ordered_products) + 3, total_steps)
        writer = PdfWriter()
        _append_pdf_pages(writer, merged_base_path)
        writer.add_blank_page(width=SEPARATOR_PAGE_WIDTH_PT, height=SEPARATOR_PAGE_HEIGHT_PT)
        with aligned_base_path.open("wb") as handle:
            writer.write(handle)
        merged_base_path = aligned_base_path
        page_count += 1
    if progress_callback:
        progress_callback(f"Finalising series PDF for {series.name or 'series'}", len(ordered_products) + 4, total_steps)
    return merged_base_path, page_count


def generate_series_pdf(
    series: Series,
    output_path: Path | None = None,
    progress_callback=None,
) -> Path:
    output_path = output_path or series_pdf_path(series)
    with tempfile.TemporaryDirectory(prefix="series-pdf-") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        base_path, _ = build_series_pdf_base(series, temp_dir, progress_callback=progress_callback)
        shutil.copyfile(base_path, output_path)

    return output_path


def generate_series_pdfs(series: Series, progress_callback=None) -> Path:
    output_path = generate_series_pdf(series, progress_callback=progress_callback)
    remove_file(all_product_types_pdf_path())
    return output_path


def product_primary_image_uri(product: Product) -> str:
    ordered_images = sorted(product.product_images or [], key=lambda img: (img.sort_order, img.id))
    if not ordered_images:
        return ""
    first_image_path = product_image_path(product.id, ordered_images[0].file_name)
    if not first_image_path.is_file():
        return ""
    return first_image_path.as_uri()


def series_primary_image_uri(series: Series) -> str:
    ordered_images = sorted(series.series_images or [], key=lambda img: (img.sort_order, img.id))
    if not ordered_images:
        return ""
    first_image_path = series_image_path(series.id, ordered_images[0].file_name)
    if not first_image_path.is_file():
        return ""
    return first_image_path.as_uri()


def series_secondary_image_uri(series: Series) -> str:
    ordered_images = sorted(series.series_images or [], key=lambda img: (img.sort_order, img.id))
    if len(ordered_images) < 2:
        return ""
    second_image_path = series_image_path(series.id, ordered_images[1].file_name)
    if not second_image_path.is_file():
        return ""
    return second_image_path.as_uri()


def build_product_type_contents_icon_url(product_type: ProductType) -> str:
    explicit_url = (product_type.contents_icon_url or "").strip()
    if explicit_url:
        return explicit_url

    label = (product_type.label or "PT").strip() or "PT"
    icon_text = html.escape(label[:2].upper())
    svg = (
        '<svg xmlns="http://www.w3.org/2000/svg" width="120" height="120" viewBox="0 0 120 120">'
        '<rect width="120" height="120" rx="20" fill="#21406f"/>'
        '<rect x="18" y="18" width="84" height="84" rx="12" fill="none" stroke="#ffffff" stroke-width="6" opacity="0.85"/>'
        f'<text x="60" y="72" text-anchor="middle" font-family="Arial, Helvetica, sans-serif" font-size="34" font-weight="700" fill="#ffffff">{icon_text}</text>'
        "</svg>"
    )
    return "data:image/svg+xml;charset=UTF-8," + urllib.parse.quote(svg)


def load_template_asset_config(template_path: Path) -> dict[str, str]:
    config_path = template_path.parent / "template-assets.json"
    if not config_path.is_file():
        return {}

    try:
        raw_config = json.loads(config_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Invalid template asset config: {config_path}") from exc

    if not isinstance(raw_config, dict):
        raise RuntimeError(f"Template asset config must be a JSON object: {config_path}")

    config: dict[str, str] = {}
    for key, value in raw_config.items():
        if value is None:
            continue
        text_value = str(value).strip()
        if text_value:
            config[str(key)] = text_value
    return config


def build_template_asset_uri(template_path: Path, asset_file_name: str) -> str:
    asset_path = template_path.parent / asset_file_name
    if not asset_path.is_file():
        return ""
    return asset_path.as_uri()


def build_product_type_series_names_html(series_names: list[str]) -> str:
    if not series_names:
        return '<p class="placeholder">No series are linked to this product type yet.</p>'

    items = "".join(
        f'<li class="series-names__item"><span class="series-names__name">{html.escape(str(series_name))}</span></li>'
        for series_name in series_names
        if str(series_name).strip()
    )
    if not items:
        return '<p class="placeholder">No series are linked to this product type yet.</p>'

    return '<ul class="series-names">' + items + "</ul>"


def build_product_type_series_legend_html(series_summaries: list[dict], include_page_ranges: bool = True) -> str:
    if not series_summaries:
        return '<p class="placeholder">No series are linked to this product type yet.</p>'

    items = []
    for summary in series_summaries:
        series_name = html.escape(str(summary.get("name") or "Series"))
        series_color = html.escape(str(summary.get("series_tab_color") or SERIES_TAB_FALLBACK_COLOR))
        product_count = int(summary.get("product_count") or 0)
        page_range_suffix = ""
        if include_page_ranges:
            page_start = summary.get("page_start")
            page_end = summary.get("page_end")
            if page_start and page_end:
                page_range_suffix = f" · Pages {page_start} to {page_end}"
            elif page_start:
                page_range_suffix = f" · Starts at page {page_start}"
        items.append(
            '<li class="series-legend__item">'
            f'<span class="series-legend__swatch" style="background:{series_color};"></span>'
            '<div class="series-legend__text">'
            f'<div class="series-legend__name">{series_name}</div>'
            f'<div class="series-legend__meta">{product_count} products{html.escape(page_range_suffix)}</div>'
            "</div></li>"
        )

    return '<ul class="series-legend">' + "".join(items) + "</ul>"


def build_product_type_series_groups_html(
    product_type: ProductType,
    series_summaries: list[dict],
    include_page_ranges: bool = True,
) -> str:
    if not series_summaries:
        return '<p class="placeholder">No series are linked to this product type yet.</p>'

    parts: list[str] = ['<section class="series-tile-grid">']
    for summary in series_summaries:
        series_name = html.escape(str(summary.get("name") or "Series"))
        series_color = html.escape(str(summary.get("series_tab_color") or SERIES_TAB_FALLBACK_COLOR))
        description_html = summary.get("series_description_html") or '<p class="placeholder">No description provided.</p>'
        image_uri = html.escape(str(summary.get("first_product_image_uri") or ""))
        page_start = int(summary.get("page_start") or 0)
        page_end = int(summary.get("page_end") or 0)
        if include_page_ranges:
            if page_start and page_end:
                page_range_text = f"Pages {page_start}-{page_end}"
            elif page_start:
                page_range_text = f"Page {page_start}"
            else:
                page_range_text = "Pages pending"
        else:
            page_range_text = "Pages pending"
        page_range_html = html.escape(page_range_text)
        image_html = (
            f'<img class="series-tile__image" src="{image_uri}" alt="{series_name}" />'
            if image_uri
            else '<div class="series-tile__placeholder">No image</div>'
        )

        parts.append(
            '<article class="series-tile" '
            f'style="--series-accent: {series_color};">'
            '<div class="series-tile__header">'
            f'<div class="series-tile__badge">{series_name}</div>'
            f'<div class="series-tile__range">{page_range_html}</div>'
            '</div>'
            '<div class="series-tile__image-wrap">'
            f"{image_html}"
            "</div>"
            f'<div class="series-tile__description">{description_html}</div>'
            "</article>"
        )

    parts.append("</section>")
    return "".join(parts)


def resolve_product_type_series_pdf_source(series: Series) -> Path | None:
    printed_path = series_pdf_path(series)
    if printed_path.is_file():
        return printed_path

    return None


def build_product_type_series_pdf_summaries(
    product_type: ProductType,
    strict: bool = True,
    progress_callback=None,
) -> tuple[list[dict], list[Path]]:
    ordered_series = sorted(product_type.series or [], key=lambda item: (item.name or "").casefold())
    series_summaries: list[dict] = []
    source_paths: list[Path] = []

    for series in ordered_series:
        if progress_callback:
            progress_callback(f"Inspecting series {series.name or 'Series'} for product type {product_type.label}", len(series_summaries) + 1, max(len(ordered_series), 1))
        ordered_products = sorted(series.products or [], key=lambda item: (item.model or "").casefold())
        source_path = resolve_product_type_series_pdf_source(series)
        if source_path is None:
            log_message = "[product_type_pdf:%s] missing series PDF for %s"
            if strict:
                logger.error(log_message, product_type.id, series.name)
                raise RuntimeError(f"Series PDF is missing for {series.name}. Generate the series PDF first.")
            logger.warning(log_message + "; continuing with empty context summary", product_type.id, series.name)
        else:
            source_paths.append(source_path)

        series_summaries.append(
            {
                "id": series.id,
                "name": series.name,
                "series_tab_color": series.series_tab_color or SERIES_TAB_FALLBACK_COLOR,
                "primary_series_image_uri": series_primary_image_uri(series),
                "secondary_series_image_uri": series_secondary_image_uri(series),
                "series_description_html": render_richtext_html(series.description1_html),
                "first_product_image_uri": product_primary_image_uri(ordered_products[0]) if ordered_products else "",
                "page_count": pdf_page_count(source_path) if source_path is not None else 0,
                "product_count": series.product_count,
                "products": [
                    {
                        "id": product.id,
                        "model": product.model,
                        "series_id": product.series_id,
                        "series_name": product.series_name_value,
                        "product_type_key": product.product_type_key,
                        "product_type_label": product.product_type_label,
                        "primary_product_image_uri": product_primary_image_uri(product),
                    }
                    for product in ordered_products
                ],
            }
        )

    return series_summaries, source_paths


SEPARATOR_PAGE_WIDTH_PT = float(A4[0])
SEPARATOR_PAGE_HEIGHT_PT = float(A4[1])


def _append_product_type_series_pdfs(
    writer: PdfWriter,
    series_base_paths: list[Path],
    current_page_count: int,
    progress_callback=None,
) -> int:
    total = max(len(series_base_paths), 1)
    for index, series_base_path in enumerate(series_base_paths, start=1):
        if progress_callback:
            progress_callback(f"Appending series PDF {index} of {total}", index, total)
        current_page_count += _append_pdf_pages(writer, series_base_path)
    return current_page_count


def _series_collage_target_dimensions() -> tuple[int, int]:
    return 1800, 2546


def _series_collage_load_image(image_uri: str) -> Image.Image | None:
    try:
        parsed = urllib.parse.urlparse(image_uri)
        if parsed.scheme != "file":
            return None
        image_path = Path(urllib.request.url2pathname(parsed.path))
        if not image_path.is_file():
            return None
        with Image.open(image_path) as image:
            return image.convert("RGB").copy()
    except Exception:
        return None


def build_product_type_series_collage_image_uri(series_summaries: list[dict], temp_dir: Path) -> str:
    started = time.perf_counter()
    collage_source_items: list[tuple[str, str]] = []
    for summary in series_summaries:
        image_uri = str(summary.get("primary_series_image_uri") or "").strip()
        if image_uri:
            collage_source_items.append((image_uri, str(summary.get("name") or "Series").strip() or "Series"))

    collage_width, collage_height = _series_collage_target_dimensions()
    collage_path = temp_dir / "product_type_series_collage.jpg"

    def apply_overlay_layers(base_image: Image.Image) -> Image.Image:
        canvas = base_image.convert("RGBA")
        overlay = Image.new("RGBA", canvas.size, (0, 0, 0, 0))
        overlay_draw = ImageDraw.Draw(overlay)
        overlay_draw.rectangle(
            (0, 0, canvas.width, canvas.height),
            fill=(255, 0, 0, 112),
        )

        vignette = Image.new("L", canvas.size, 0)
        vignette_draw = ImageDraw.Draw(vignette)
        vignette_draw.ellipse(
            (
                int(canvas.width * 0.05),
                int(canvas.height * 0.06),
                int(canvas.width * 0.95),
                int(canvas.height * 0.94),
            ),
            fill=255,
        )
        vignette = vignette.filter(ImageFilter.GaussianBlur(radius=max(canvas.size) * 0.10))
        vignette_rgba = Image.new("RGBA", canvas.size, (0, 0, 0, 0))
        vignette_rgba.putalpha(vignette.point(lambda value: int((255 - value) * 0.88)))

        return Image.alpha_composite(Image.alpha_composite(canvas, overlay), vignette_rgba).convert("RGB")

    if not collage_source_items:
        collage = Image.new("RGB", (collage_width, collage_height), (12, 12, 14))
        draw = ImageDraw.Draw(collage)
        message = "No series images available"
        font = ImageFont.load_default()
        text_bbox = draw.textbbox((0, 0), message, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_height = text_bbox[3] - text_bbox[1]
        draw.text(
            ((collage_width - text_width) / 2, (collage_height - text_height) / 2),
            message,
            fill=(225, 228, 234),
            font=font,
        )
        apply_overlay_layers(collage).save(collage_path, format="JPEG", quality=82)
        logger.info("[product-type-pdf] collage images=0 bytes=%d elapsed=%.1fs", collage_path.stat().st_size, time.perf_counter() - started)
        return collage_path.as_uri()

    image_count = len(collage_source_items)
    columns = max(2, min(6, math.ceil(math.sqrt(image_count * 1.15))))
    rows = math.ceil(image_count / columns)
    gap = 20
    tile_width = max(1, (collage_width - (gap * (columns + 1))) // columns)
    tile_height = max(1, (collage_height - (gap * (rows + 1))) // rows)
    resample_filter = getattr(getattr(Image, "Resampling", Image), "LANCZOS")
    background = Image.new("RGB", (collage_width, collage_height), (10, 10, 12))

    for index, (image_uri, series_name) in enumerate(collage_source_items):
        source_image = _series_collage_load_image(image_uri)
        col = index % columns
        row = index // columns
        left = gap + (col * (tile_width + gap))
        top = gap + (row * (tile_height + gap))
        tile_box = (left, top, left + tile_width, top + tile_height)

        if source_image is None:
            tile = Image.new("RGB", (tile_width, tile_height), (30, 18, 20))
            draw = ImageDraw.Draw(tile)
            font = ImageFont.load_default()
            label = series_name[:40] or "Series"
            label_bbox = draw.textbbox((0, 0), label, font=font)
            label_width = label_bbox[2] - label_bbox[0]
            label_height = label_bbox[3] - label_bbox[1]
            draw.text(
                ((tile_width - label_width) / 2, (tile_height - label_height) / 2),
                label,
                fill=(225, 225, 225),
                font=font,
            )
            background.paste(tile, tile_box)
            continue

        fitted = ImageOps.fit(source_image, (tile_width, tile_height), method=resample_filter)
        background.paste(fitted, tile_box)

    apply_overlay_layers(background).save(collage_path, format="JPEG", quality=82)
    logger.info(
        "[product-type-pdf] collage images=%d bytes=%d elapsed=%.1fs",
        image_count,
        collage_path.stat().st_size,
        time.perf_counter() - started,
    )
    return collage_path.as_uri()


def render_product_type_intro_with_page_ranges(
    product_type: ProductType,
    temp_dir: Path,
    series_summaries: list[dict],
    series_names_html: str,
    series_legend_html: str,
    series_collage_image_uri: str,
    progress_callback=None,
) -> tuple[Path, int, str, str]:
    intro_base_path = temp_dir / f"product_type_printed_{sanitize_name(product_type.key or product_type.label or 'unknown')}_intro.pdf"
    intro_page_count = 0
    rendered_series_groups_html = ""
    intro_stylesheet_text = ""

    rendered_series_groups_html = build_product_type_series_groups_html(product_type, series_summaries, include_page_ranges=False)
    intro_html, intro_stylesheet_text = build_product_type_pdf_html(
        product_type,
        series_names_html,
        rendered_series_groups_html,
        series_legend_html,
        series_collage_image_uri,
    )
    if progress_callback:
        progress_callback(f"Rendering product type intro for {product_type.label}", 1, 2)
    render_pdf_from_html(intro_html, intro_base_path, intro_stylesheet_text)
    intro_page_count = pdf_page_count(intro_base_path)
    logger.info("[product-type-pdf] first intro render pages=%d product_type=%s", intro_page_count, product_type.label)

    current_page_count = intro_page_count
    for summary in series_summaries:
        series_page_count = int(summary.get("page_count") or 0)
        summary["buffer_page"] = 0

        if series_page_count > 0 and current_page_count % 2 == 1:
            current_page_count += 1
            summary["buffer_page"] = current_page_count

        if series_page_count > 0:
            summary["page_start"] = current_page_count + 1
            current_page_count += series_page_count
            summary["page_end"] = current_page_count
        else:
            summary["page_start"] = 0
            summary["page_end"] = 0

    rendered_series_groups_html = build_product_type_series_groups_html(product_type, series_summaries, include_page_ranges=True)
    intro_html, intro_stylesheet_text = build_product_type_pdf_html(
        product_type,
        series_names_html,
        rendered_series_groups_html,
        series_legend_html,
        series_collage_image_uri,
    )
    if progress_callback:
        progress_callback(f"Rendering product type intro with page ranges for {product_type.label}", 2, 2)
    render_pdf_from_html(intro_html, intro_base_path, intro_stylesheet_text)
    intro_page_count = pdf_page_count(intro_base_path)
    logger.info("[product-type-pdf] final intro render pages=%d product_type=%s", intro_page_count, product_type.label)

    return intro_base_path, intro_page_count, rendered_series_groups_html, intro_stylesheet_text


def build_product_type_contents_html(product_type: ProductType, series_summaries: list[dict]) -> str:
    return build_product_type_series_groups_html(product_type, series_summaries)


def build_product_type_pdf_html(
    product_type: ProductType,
    series_names_html: str,
    series_groups_html: str,
    series_legend_html: str,
    series_collage_image_uri: str,
) -> tuple[str, str]:
    template_id = resolve_product_type_pdf_template_id(product_type) or "product_type-default"
    template_definition = get_template_definition(template_id, "product_type")
    if template_definition is None:
        raise RuntimeError("No product type PDF template is configured.")

    project_root = Path(__file__).resolve().parents[1]
    template_path = project_root / template_definition["path"]
    if not template_path.is_file():
        raise RuntimeError(f"Product type template file is missing: {template_path}")

    stylesheet_path = resolve_template_stylesheet_path(template_definition, template_path)
    html_template = template_path.read_text(encoding="utf-8")
    stylesheet_text = stylesheet_path.read_text(encoding="utf-8") if stylesheet_path.is_file() else ""
    html_template = inline_template_stylesheet(html_template, stylesheet_text)
    asset_config = load_template_asset_config(template_path)

    replacements = {
        "{{product_type.key}}": html.escape(product_type.key or ""),
        "{{product_type.label}}": html.escape(product_type.label or ""),
        "{{product_type.contents_icon_url}}": build_product_type_contents_icon_url(product_type),
        "{{product_type.cover_image_url}}": build_template_asset_uri(template_path, asset_config.get("cover_image", "")),
        "{{product_type.intermediate_image_url}}": build_template_asset_uri(template_path, asset_config.get("intermediate_image", "")),
        "{{product_type.contact_map_image_url}}": build_template_asset_uri(template_path, asset_config.get("contact_map_image_url", "")),
        "{{product_type.contact_shopfront_image_url}}": build_template_asset_uri(template_path, asset_config.get("contact_shopfront_image_url", "")),
        "{{product_type.series_collage_image_url}}": series_collage_image_uri,
        "{{product_type.series_names}}": html.escape(", ".join(product_type.series_names or [])),
        "{{product_type.series_names_html}}": series_names_html,
        "{{product_type.series_legend_html}}": series_legend_html,
        "{{product_type.series_groups_html}}": series_groups_html,
        "{{product_type.contents_html}}": series_groups_html,
    }

    rendered = html_template
    for token, value in replacements.items():
        rendered = rendered.replace(token, value)
    return rendered, stylesheet_text


def build_product_type_pdf_base(product_type: ProductType, temp_dir: Path, progress_callback=None) -> tuple[Path, dict]:
    series_summaries, series_base_paths = build_product_type_series_pdf_summaries(
        product_type,
        progress_callback=_make_progress_window(progress_callback, 1, 20) if progress_callback else None,
    )
    series_names_html = build_product_type_series_names_html(product_type.series_names or [])
    series_legend_html = build_product_type_series_legend_html(series_summaries)
    series_collage_image_uri = build_product_type_series_collage_image_uri(series_summaries, temp_dir)
    intro_base_path, intro_page_count, series_groups_html, _ = render_product_type_intro_with_page_ranges(
        product_type,
        temp_dir,
        series_summaries,
        series_names_html,
        series_legend_html,
        series_collage_image_uri,
        progress_callback=_make_progress_window(progress_callback, 21, 40) if progress_callback else None,
    )

    merged_base_path = temp_dir / f"product_type_printed_{sanitize_name(product_type.key or product_type.label or 'unknown')}_base.pdf"
    merged_writer = PdfWriter()
    current_page_count = _append_pdf_pages(merged_writer, intro_base_path)

    if series_summaries and current_page_count % 2 == 1:
        merged_writer.add_blank_page(width=SEPARATOR_PAGE_WIDTH_PT, height=SEPARATOR_PAGE_HEIGHT_PT)
        current_page_count += 1

    if series_summaries:
        current_page_count = _append_product_type_series_pdfs(
            merged_writer,
            series_base_paths,
            current_page_count,
            progress_callback=_make_progress_window(progress_callback, 41, 90) if progress_callback else None,
        )

    merged_base_path.parent.mkdir(parents=True, exist_ok=True)
    with merged_base_path.open("wb") as handle:
        if progress_callback:
            progress_callback(f"Writing product type PDF for {product_type.label}", 91, 100)
        merged_writer.write(handle)
    return merged_base_path, {
        "intro_page_count": intro_page_count,
        "page_count": current_page_count,
        "series_summaries": series_summaries,
        "series_names_html": series_names_html,
        "series_groups_html": series_groups_html,
        "contents_html": series_groups_html,
        "series_legend_html": series_legend_html,
        "series_collage_image_uri": series_collage_image_uri,
    }


def build_product_type_pdf_context_metadata(product_type: ProductType, temp_dir: Path) -> dict:
    series_summaries, _ = build_product_type_series_pdf_summaries(product_type, strict=False)
    series_names_html = build_product_type_series_names_html(product_type.series_names or [])
    series_legend_html = build_product_type_series_legend_html(series_summaries)
    series_collage_image_uri = build_product_type_series_collage_image_uri(series_summaries, temp_dir)
    _, intro_page_count, series_groups_html, _ = render_product_type_intro_with_page_ranges(
        product_type,
        temp_dir,
        series_summaries,
        series_names_html,
        series_legend_html,
        series_collage_image_uri,
    )

    current_page_count = intro_page_count
    if series_summaries and current_page_count % 2 == 1:
        current_page_count += 1
    for summary in series_summaries:
        current_page_count += int(summary.get("page_count") or 0)

    return {
        "intro_page_count": intro_page_count,
        "page_count": current_page_count,
        "series_summaries": series_summaries,
        "series_names_html": series_names_html,
        "series_groups_html": series_groups_html,
        "contents_html": series_groups_html,
        "series_legend_html": series_legend_html,
        "series_collage_image_uri": series_collage_image_uri,
    }


def build_product_type_page_decorations(metadata: dict, decorations: list[dict] | None = None) -> list[dict]:
    decorations = decorations or []
    series_summaries = metadata.get("series_summaries") or []
    intro_page_count = int(metadata.get("intro_page_count") or 0)
    series_tabs = [
        {
            "series_id": summary.get("id"),
            "tab_label": summary.get("name") or "Series",
            "tab_color": summary.get("series_tab_color") or SERIES_TAB_FALLBACK_COLOR,
        }
        for summary in series_summaries
    ]

    for _ in range(intro_page_count):
        decorations.append({"tabs": []})

    for active_summary in series_summaries:
        if int(active_summary.get("buffer_page") or 0):
            decorations.append({"tabs": []})
        active_id = active_summary.get("id")
        page_count = int(active_summary.get("page_count") or 0)
        for _ in range(page_count):
            decorations.append(
                {
                    "tabs": [
                        {
                            **tab,
                            "tab_opacity": 1.0 if tab.get("series_id") == active_id else 0.2,
                        }
                        for tab in series_tabs
                    ]
                }
            )

    return decorations


def generate_product_type_pdf_with_metadata(product_type: ProductType, progress_callback=None) -> tuple[Path, dict]:
    output_path = product_type_pdf_path(product_type)
    with tempfile.TemporaryDirectory(prefix="product-type-pdf-") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        base_path, metadata = build_product_type_pdf_base(
            product_type,
            temp_dir,
            progress_callback=progress_callback,
        )
        page_decorations = build_product_type_page_decorations(metadata)
        stamp_pdf_file(base_path, output_path, page_decorations)
    return output_path, metadata


def generate_product_type_pdf(product_type: ProductType, progress_callback=None) -> Path:
    output_path, _ = generate_product_type_pdf_with_metadata(product_type, progress_callback=progress_callback)
    remove_file(all_product_types_pdf_path())
    return output_path


def _product_type_pdf_intro_page_count(
    product_type: ProductType,
    source_path: Path,
    series_paths: list[Path],
) -> int:
    """Return the front-matter page count in an existing product-type PDF.

    The generated product-type PDF stores its shared introduction first, then an
    optional alignment page, then the already-generated series PDFs.  The two
    designed printed templates are four pages; the default template is normally
    one page and may therefore have one alignment page after it.
    """
    total_pages = pdf_page_count(source_path)
    series_pages = sum(pdf_page_count(path) for path in series_paths)
    remaining_pages = max(total_pages - series_pages, 0)
    template_id = resolve_product_type_pdf_template_id(product_type) or "product_type-default"
    if template_id in {"product_type-printed_type_1", "product_type-printed_type_2"}:
        return min(4, remaining_pages)
    if series_paths and remaining_pages > 1 and remaining_pages % 2 == 0:
        return remaining_pages - 1
    return remaining_pages


def _append_pdf_page_slice(writer: PdfWriter, pdf_path: Path, start_index: int) -> int:
    reader = PdfReader(str(pdf_path))
    pages_added = 0
    for page in reader.pages[max(0, start_index):]:
        writer.add_page(page)
        pages_added += 1
    return pages_added


def _render_combined_product_type_contents_page(
    product_type: ProductType,
    series_summaries: list[dict],
    page_count_before_contents: int,
    temp_dir: Path,
) -> tuple[Path, int]:
    """Render a product-type intro and return its final contents page.

    Rendering the existing product-type template lets the combined catalogue
    retain the configured visual contents layout.  Only its final page is used;
    the shared cover/contact pages are supplied once by the combined catalogue.
    """
    series_page_cursor = page_count_before_contents + 1
    if series_summaries and series_page_cursor % 2 == 1:
        series_page_cursor += 1
    for summary in series_summaries:
        page_count = int(summary.get("page_count") or 0)
        summary["page_start"] = series_page_cursor + 1 if page_count else 0
        summary["page_end"] = series_page_cursor + page_count if page_count else 0
        if page_count:
            series_page_cursor += page_count

    series_names_html = build_product_type_series_names_html(product_type.series_names or [])
    series_groups_html = build_product_type_series_groups_html(
        product_type,
        series_summaries,
        include_page_ranges=True,
    )
    series_legend_html = build_product_type_series_legend_html(
        series_summaries,
        include_page_ranges=True,
    )
    collage_uri = build_product_type_series_collage_image_uri(series_summaries, temp_dir)
    intro_html, stylesheet_text = build_product_type_pdf_html(
        product_type,
        series_names_html,
        series_groups_html,
        series_legend_html,
        collage_uri,
    )
    rendered_path = temp_dir / f"combined_{sanitize_name(product_type.key or product_type.label or 'unknown')}_intro.pdf"
    render_pdf_from_html(intro_html, rendered_path, stylesheet_text)
    return rendered_path, pdf_page_count(rendered_path)


def generate_all_product_types_pdf(
    product_types: list[ProductType],
    progress_callback=None,
) -> Path:
    """Build one catalogue from the existing product-type/series PDFs."""
    ordered_types = sorted(product_types, key=lambda item: (getattr(item, "sort_order", 0) or 0, item.id or 0))
    if not ordered_types:
        raise RuntimeError("No product types are available for the combined catalogue.")

    prepared: list[dict] = []
    for index, product_type in enumerate(ordered_types, start=1):
        source_path = product_type_pdf_path(product_type)
        if not source_path.is_file():
            if progress_callback:
                progress_callback(f"Generating missing product type PDF for {product_type.label}", 1, 1)
            # A combined-catalogue request should be able to recover when the
            # individual type PDF has not been generated yet, provided its
            # underlying printed series PDFs are available.
            generate_product_type_pdf(product_type, progress_callback=None)
            source_path = product_type_pdf_path(product_type)
            if not source_path.is_file():
                raise RuntimeError(f"Product type PDF is missing for {product_type.label} after generation.")
        series_summaries, series_paths = build_product_type_series_pdf_summaries(
            product_type,
            strict=True,
            progress_callback=None,
        )
        intro_page_count = _product_type_pdf_intro_page_count(product_type, source_path, series_paths)
        if intro_page_count < 1:
            raise RuntimeError(f"Product type PDF has no introductory pages for {product_type.label}.")
        prepared.append(
            {
                "product_type": product_type,
                "source_path": source_path,
                "series_summaries": series_summaries,
                "series_paths": series_paths,
                "intro_page_count": intro_page_count,
                "series_start_index": intro_page_count + (1 if series_paths and intro_page_count % 2 == 1 else 0),
            }
        )
        if progress_callback:
            progress_callback(f"Inspecting product type {index} of {len(ordered_types)}: {product_type.label}", index, len(ordered_types))

    # Use one existing designed product-type intro as the shared front matter.
    # Prefer the four-page printed layouts; fall back to the first available PDF.
    shared = next(
        (item for item in prepared if item["intro_page_count"] >= 4),
        prepared[0],
    )

    output_path = all_product_types_pdf_path()
    with tempfile.TemporaryDirectory(prefix="all-product-types-pdf-") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        writer = PdfWriter()
        shared_intro_count = shared["intro_page_count"]
        shared_page_count = max(shared_intro_count - 1, 0)
        _append_pdf_page_slice(writer, shared["source_path"], 0)
        # Remove the shared source's product-type contents page; it is replaced
        # below by one contents page for each product type.
        if len(writer.pages) > shared_page_count:
            writer = PdfWriter()
            reader = PdfReader(str(shared["source_path"]))
            for page in reader.pages[:shared_page_count]:
                writer.add_page(page)
        current_page_count = len(writer.pages)

        for index, item in enumerate(prepared, start=1):
            product_type = item["product_type"]
            if progress_callback:
                progress_callback(f"Rendering contents page {index} of {len(prepared)}: {product_type.label}", index, len(prepared))
            contents_path, rendered_intro_count = _render_combined_product_type_contents_page(
                product_type,
                item["series_summaries"],
                current_page_count,
                temp_dir,
            )
            # The contents page is the final page in every registered product-
            # type template, including the compact/default fallback.
            contents_reader = PdfReader(str(contents_path))
            writer.add_page(contents_reader.pages[-1])
            current_page_count += 1

            if item["series_paths"]:
                if current_page_count % 2 == 1:
                    writer.add_blank_page(width=SEPARATOR_PAGE_WIDTH_PT, height=SEPARATOR_PAGE_HEIGHT_PT)
                    current_page_count += 1
                for series_path in item["series_paths"]:
                    current_page_count += _append_pdf_pages(writer, series_path)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("wb") as handle:
            writer.write(handle)
    logger.info("[all-product-types-pdf] pages=%d output=%s", current_page_count, output_path)
    return output_path


def delete_product_image_file(image: ProductImage):
    remove_file(product_image_target_path(image.product_id, image.file_name))
    remove_file(PRODUCT_IMAGES_DIR / _normalize_media_relative_path(image.file_name))


def delete_series_image_file(image: SeriesImage):
    remove_file(series_image_target_path(image.series_id, image.file_name))
    remove_file(SERIES_IMAGES_DIR / _normalize_media_relative_path(image.file_name))


def delete_associated_document_files(owner):
    for document in list(getattr(owner, "associated_documents", []) or []):
        remove_file(associated_document_path(document.owner_type, document.owner_id, document.file_name))


def delete_product_type_assets(product_type: ProductType):
    safe_key = re.sub(r"[^a-z0-9]+", "_", (product_type.key or "").strip().lower()).strip("_") or "unknown"
    for pdf_path in PRODUCT_TYPE_PDFS_DIR.glob(f"product_type_*_{safe_key}.pdf"):
        remove_file(pdf_path)
    # The combined catalogue contains this product type and must not survive
    # as a silently stale public download after the type is removed.
    remove_file(all_product_types_pdf_path())
    delete_associated_document_files(product_type)


def build_graph_config(product_type: ProductType | None) -> dict:
    return {
        "graph_kind": product_type.graph_kind if product_type else "fan_map",
        "supports_graph_overlays": product_type.supports_graph_overlays if product_type else True,
        "supports_band_graph_style": product_type.supports_band_graph_style if product_type else True,
        "graph_line_value_label": product_type.graph_line_value_label if product_type else "RPM",
        "graph_line_value_unit": product_type.graph_line_value_unit if product_type else "RPM",
        "graph_x_axis_label": product_type.graph_x_axis_label if product_type else "Airflow",
        "graph_x_axis_unit": product_type.graph_x_axis_unit if product_type else "L/s",
        "graph_y_axis_label": product_type.graph_y_axis_label if product_type else "Pressure",
        "graph_y_axis_unit": product_type.graph_y_axis_unit if product_type else "Pa",
    }


def parse_parameter_filters(raw_filters: str | None) -> list[dict]:
    if not raw_filters:
        return []

    try:
        decoded = json.loads(raw_filters)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid parameter_filters JSON: {exc.msg}") from exc

    if not isinstance(decoded, list):
        raise HTTPException(status_code=400, detail="parameter_filters must be a JSON array")

    normalized_filters: list[dict] = []
    for item in decoded:
        if not isinstance(item, dict):
            raise HTTPException(status_code=400, detail="Each parameter filter must be an object")

        group_name = str(item.get("group_name", "")).strip()
        parameter_name = str(item.get("parameter_name", "")).strip()
        value_string = item.get("value_string")
        min_number = item.get("min_number")
        max_number = item.get("max_number")

        if not group_name or not parameter_name:
            raise HTTPException(status_code=400, detail="Each parameter filter must include group_name and parameter_name")

        try:
            normalized = {
                "group_name": group_name,
                "parameter_name": parameter_name,
                "value_string": str(value_string).strip() if value_string not in {None, ""} else None,
                "min_number": float(min_number) if min_number not in {None, ""} else None,
                "max_number": float(max_number) if max_number not in {None, ""} else None,
            }
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="parameter_filters numeric bounds must be valid numbers") from exc

        if normalized["value_string"] is None and normalized["min_number"] is None and normalized["max_number"] is None:
            raise HTTPException(
                status_code=400,
                detail="Each parameter filter must include value_string or at least one numeric bound",
            )

        if (
            normalized["min_number"] is not None and
            normalized["max_number"] is not None and
            normalized["min_number"] > normalized["max_number"]
        ):
            raise HTTPException(status_code=400, detail="parameter_filters numeric min cannot be greater than max")

        normalized_filters.append(normalized)

    return normalized_filters


def graph_filter_values(product: Product) -> dict[str, list[float]]:
    rpm_values: set[float] = set()
    airflow_values: set[float] = set()
    pressure_values: set[float] = set()

    for rpm_line in product.rpm_lines or []:
        if rpm_line.rpm is not None:
            rpm_values.add(float(rpm_line.rpm))
        for point in rpm_line.points or []:
            if point.airflow is not None:
                airflow_values.add(float(point.airflow))
            if point.pressure is not None:
                pressure_values.add(float(point.pressure))

    for point in product.efficiency_points or []:
        if point.airflow is not None:
            airflow_values.add(float(point.airflow))

    return {
        "rpm": sorted(rpm_values),
        "airflow": sorted(airflow_values),
        "pressure": sorted(pressure_values),
    }


def graph_filter_values_for_products(products: list[Product]) -> dict[str, list[float]]:
    rpm_values: set[float] = set()
    airflow_values: set[float] = set()
    pressure_values: set[float] = set()

    for product in products:
        graph_values = graph_filter_values(product)
        rpm_values.update(graph_values["rpm"])
        airflow_values.update(graph_values["airflow"])
        pressure_values.update(graph_values["pressure"])

    return {
        "rpm": sorted(rpm_values),
        "airflow": sorted(airflow_values),
        "pressure": sorted(pressure_values),
    }


def load_series_product_counts(db: Session, series_ids: list[int] | None = None) -> dict[int, int]:
    if series_ids is not None and not series_ids:
        return {}
    q = db.query(Product.series_id, func.count(Product.id))
    if series_ids is not None:
        q = q.filter(Product.series_id.in_(series_ids))
    rows = q.group_by(Product.series_id).all()
    return {int(series_id): int(count) for series_id, count in rows if series_id is not None}


def assign_series_product_counts(series_items: list[Series], counts: dict[int, int]) -> None:
    for series in series_items:
        setattr(series, "_product_count", counts.get(series.id, 0))


def notify_public_catalogue_cache_refresh():
    if not PUBLIC_CATALOGUE_SITE_URL or not CMS_API_TOKEN:
        return

    global PUBLIC_CATALOGUE_REFRESH_IN_FLIGHT
    with PUBLIC_CATALOGUE_REFRESH_LOCK:
        if PUBLIC_CATALOGUE_REFRESH_IN_FLIGHT:
            return
        PUBLIC_CATALOGUE_REFRESH_IN_FLIGHT = True

    def _post_refresh():
        global PUBLIC_CATALOGUE_REFRESH_IN_FLIGHT
        try:
            _request_public_catalogue_cache_refresh()
        except (urllib.error.HTTPError, urllib.error.URLError, ValueError) as exc:
            logger.warning("Catalogue cache refresh notification failed: %s", exc)
        finally:
            with PUBLIC_CATALOGUE_REFRESH_LOCK:
                PUBLIC_CATALOGUE_REFRESH_IN_FLIGHT = False

    threading.Thread(target=_post_refresh, daemon=True).start()


def _request_public_catalogue_cache_refresh():
    request = urllib.request.Request(
        f"{PUBLIC_CATALOGUE_SITE_URL}/api/cache/refresh",
        data=b"",
        method="POST",
        headers={"Authorization": f"Bearer {CMS_API_TOKEN}"},
    )
    with urllib.request.urlopen(request, timeout=5) as response:
        payload = response.read().decode("utf-8").strip()

    if not payload:
        return {}

    try:
        parsed = json.loads(payload)
    except json.JSONDecodeError:
        return {"raw_response": payload}

    return parsed if isinstance(parsed, dict) else {"raw_response": payload}


def refresh_public_catalogue_cache():
    if not PUBLIC_CATALOGUE_SITE_URL or not CMS_API_TOKEN:
        raise HTTPException(
            status_code=503,
            detail="Customer-facing catalogue refresh is unavailable because the site URL or token is not configured.",
        )

    try:
        return _request_public_catalogue_cache_refresh()
    except urllib.error.HTTPError as exc:
        try:
            detail = exc.read().decode("utf-8").strip()
        except Exception:
            detail = ""
        raise HTTPException(
            status_code=exc.code,
            detail=detail or f"Customer-facing catalogue refresh failed with status {exc.code}.",
        ) from exc
    except (urllib.error.URLError, ValueError) as exc:
        raise HTTPException(status_code=502, detail="Customer-facing catalogue refresh is unavailable right now.") from exc


def value_in_window(value: float, minimum: float | None, maximum: float | None) -> bool:
    if minimum is not None and value < minimum:
        return False
    if maximum is not None and value > maximum:
        return False
    return True


def product_matches_parameter_filters(product: Product, parameter_filters: list[dict]) -> bool:
    if not parameter_filters:
        return True

    product_label = f"{product.model or 'unknown model'} (id={product.id})"
    grouped_parameters: dict[tuple[str, str], list[ProductParameter]] = {}
    for group in product.parameter_groups:
            group_name = (group.group_name or "").strip().casefold()
            for parameter in group.parameters:
                parameter_name = (parameter.parameter_name or "").strip().casefold()
                grouped_parameters.setdefault((group_name, parameter_name), []).append(parameter)

    graph_values = graph_filter_values(product)

    for filter_item in parameter_filters:
        filter_key = (
            filter_item["group_name"].strip().casefold(),
            filter_item["parameter_name"].strip().casefold(),
        )
        if filter_key[0] == GRAPH_FILTER_GROUP_NAME:
            min_number = filter_item.get("min_number")
            max_number = filter_item.get("max_number")
            graph_metric_values = graph_values.get(filter_key[1], [])
            trace_product_filter(
                "product filter trace: product=%s graph_filter=%s min=%s max=%s values=%s",
                product_label,
                filter_key[1],
                min_number,
                max_number,
                graph_metric_values,
            )
            if not graph_metric_values:
                trace_product_filter(
                    "product filter trace result: product=%s graph_filter=%s matched=False reason=no_graph_values",
                    product_label,
                    filter_key[1],
                )
                return False
            graph_matches = [value_in_window(metric_value, min_number, max_number) for metric_value in graph_metric_values]
            trace_product_filter(
                "product filter trace compare: product=%s graph_filter=%s comparisons=%s mode=all",
                product_label,
                filter_key[1],
                graph_matches,
            )
            if not all(graph_matches):
                trace_product_filter(
                    "product filter trace result: product=%s graph_filter=%s matched=False reason=graph_value_out_of_window",
                    product_label,
                    filter_key[1],
                )
                return False
            continue

        matching_parameters = grouped_parameters.get(filter_key, [])
        trace_product_filter(
            "product filter trace: product=%s spec_filter=%s.%s candidates=%s",
            product_label,
            filter_key[0],
            filter_key[1],
            len(matching_parameters),
        )
        if not matching_parameters:
            trace_product_filter(
                "product filter trace result: product=%s spec_filter=%s.%s matched=False reason=no_matching_parameters",
                product_label,
                filter_key[0],
                filter_key[1],
            )
            return False

        value_string = filter_item.get("value_string")
        min_number = filter_item.get("min_number")
        max_number = filter_item.get("max_number")

        matched = False
        for parameter in matching_parameters:
            if value_string is not None:
                if (parameter.value_string or "").strip().casefold() == value_string.casefold():
                    matched = True
                    break
                continue

            if parameter.value_number is None:
                continue

            if value_in_window(parameter.value_number, min_number, max_number):
                matched = True
                break

        if not matched:
            trace_product_filter(
                "product filter trace result: product=%s spec_filter=%s.%s matched=False reason=no_parameter_match value_string=%s min=%s max=%s",
                product_label,
                filter_key[0],
                filter_key[1],
                value_string,
                min_number,
                max_number,
            )
            return False

        trace_product_filter(
            "product filter trace result: product=%s spec_filter=%s.%s matched=True value_string=%s min=%s max=%s",
            product_label,
            filter_key[0],
            filter_key[1],
            value_string,
            min_number,
            max_number,
        )

    return True


def sync_graph_image(product: Product, rpm_lines: list[RpmLine], efficiency_points: list[EfficiencyPoint]):
    rpm_point_list = sorted(
        [point for rpm_line in rpm_lines for point in rpm_line.points],
        key=lambda point: (point.rpm or 0, point.airflow),
    )
    efficiency_point_list = sorted(efficiency_points, key=lambda point: point.airflow)
    previous_path = Path(product.graph_image_path) if product.graph_image_path else None

    if not rpm_point_list and not efficiency_point_list:
        if previous_path:
            remove_file(previous_path)
        product.graph_image_path = None
        return

    final_path = PRODUCT_GRAPHS_DIR / graph_file_name(product)
    tmp_path = PRODUCT_GRAPHS_DIR / f"tmp_{graph_file_name(product)}"
    if tmp_path.exists():
        tmp_path.unlink()

    payload = {
        "title": f"{product.product_type_label} | {product.series_name} - {product.model} Performance Graph",
        "graphMode": "product",
        "showRpmBandShading": product.show_rpm_band_shading,
        "permissibleUseMode": normalize_permissible_use_mode(product.permissible_use_mode),
        "graphConfig": build_graph_config(product.product_type),
        "graphStyle": {
            "band_graph_background_color": product.band_graph_background_color,
            "band_graph_label_text_color": product.band_graph_label_text_color,
            "band_graph_faded_opacity": product.band_graph_faded_opacity,
            "band_graph_permissible_label_color": product.band_graph_permissible_label_color,
        },
        "rpmLines": [
            {
                "id": line.id,
                "product_id": line.product_id,
                "rpm": line.rpm,
                "band_color": line.band_color,
            }
            for line in sorted(rpm_lines, key=lambda line: line.rpm)
        ],
        "rpmPoints": [
            {
                "id": point.id,
                "product_id": point.product_id,
                "rpm_line_id": point.rpm_line_id,
                "rpm": point.rpm,
                "airflow": point.airflow,
                "pressure": point.pressure,
            }
            for point in rpm_point_list
        ],
        "efficiencyPoints": [
            {
                "id": point.id,
                "product_id": point.product_id,
                "airflow": point.airflow,
                "efficiency_centre": point.efficiency_centre,
                "efficiency_lower_end": point.efficiency_lower_end,
                "efficiency_higher_end": point.efficiency_higher_end,
                "permissible_use": point.permissible_use,
            }
            for point in efficiency_point_list
        ],
    }

    result = subprocess.run(
        ["node", str(ECHARTS_RENDER_SCRIPT), str(tmp_path)],
        cwd=str(FRONTEND_DIR),
        input=json.dumps(payload),
        text=True,
        capture_output=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(
            "ECharts graph render failed: "
            + (result.stderr.strip() or result.stdout.strip() or f"exit code {result.returncode}")
        )

    if previous_path and previous_path != final_path:
        remove_file(previous_path)
    shutil.move(tmp_path, final_path)
    product.graph_image_path = str(final_path)


def delete_product_assets(product: Product):
    for image in product.product_images:
        delete_product_image_file(image)
    if product.graph_image_path:
        remove_file(product.graph_image_path)


def refresh_graph_for_product(db: Session, product: Product):
    db.refresh(product)
    sync_graph_image(product, list(product.rpm_lines), list(product.efficiency_points))


def sync_fan_acoustic_table_for_product(db: Session, product: Product):
    if product.product_type_key != "fan":
        product.fan_acoustic_table = None
        return
    try:
        db.expire(product, ["rpm_lines"])
    except Exception:
        pass
    product.fan_acoustic_table = product.fan_acoustic_table


def clear_all_graph_images(db: Session) -> int:
    deleted_files = 0
    for graph_file in PRODUCT_GRAPHS_DIR.iterdir():
        if not graph_file.is_file():
            continue
        graph_file.unlink(missing_ok=True)
        deleted_files += 1

    for product in db.query(Product).all():
        product.graph_image_path = None

    return deleted_files


def require_product(db: Session, product_id: int) -> Product:
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(404, "Product not found")
    return product


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, PASSWORD_HASH_ITERATIONS)
    return f"pbkdf2_sha256${PASSWORD_HASH_ITERATIONS}${salt.hex()}${digest.hex()}"


def verify_password(password: str, password_hash: str) -> bool:
    try:
        algorithm, iteration_text, salt_hex, expected_hex = password_hash.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        iterations = int(iteration_text)
        salt = bytes.fromhex(salt_hex)
        expected = bytes.fromhex(expected_hex)
    except (ValueError, TypeError):
        return False

    actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return secrets.compare_digest(actual, expected)


def ensure_auth_config():
    missing = []
    if not SESSION_SECRET:
        missing.append("SESSION_SECRET")
    if missing:
        raise RuntimeError(
            "Authentication is enabled but required environment variables are missing: "
            + ", ".join(missing)
        )


def ensure_bootstrap_admin():
    with SessionLocal() as db:
        existing_user = db.query(User).first()
        if existing_user is not None:
            return
        if not BOOTSTRAP_ADMIN_USERNAME or not BOOTSTRAP_ADMIN_PASSWORD:
            raise RuntimeError(
                "No users exist yet. Set BOOTSTRAP_ADMIN_USERNAME and BOOTSTRAP_ADMIN_PASSWORD "
                "so the first admin account can be created."
            )
        admin_user = User(
            username=BOOTSTRAP_ADMIN_USERNAME,
            password_hash=hash_password(BOOTSTRAP_ADMIN_PASSWORD),
            is_admin=True,
            is_active=True,
        )
        db.add(admin_user)
        db.commit()


def is_authenticated(request: Request) -> bool:
    return request.session.get("authenticated") is True and request.session.get("user_id") is not None


def get_authenticated_user_id(request: Request) -> Optional[int]:
    user_id = request.session.get("user_id")
    return int(user_id) if user_id is not None else None


def get_current_user(request: Request, db: Session = Depends(get_db)) -> User:
    user_id = get_authenticated_user_id(request)
    if user_id is None:
        raise HTTPException(status_code=401, detail="Authentication required")
    user = db.get(User, user_id)
    if user is None or not user.is_active:
        request.session.clear()
        raise HTTPException(status_code=401, detail="Authentication required")
    return user


def require_admin_user(current_user: User = Depends(get_current_user)) -> User:
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user


QUOTE_REQUEST_REQUEST_TYPE_LABELS = {
    "standard": "Quote this item",
    "tailored": "Tailored product",
    "unsure": "Not sure yet",
}

QUOTE_REQUEST_ATTRIBUTE_LABELS = {
    "airflow": "Airflow",
    "pressure": "Pressure",
    "power": "Power",
    "efficiency": "Efficiency",
    "noise": "Noise",
    "size": "Size",
    "temperature": "Temperature",
    "mounting": "Mounting",
}


def _quote_request_parse_recipient_emails(raw_value: object | None) -> list[str]:
    if not raw_value:
        return []

    if isinstance(raw_value, list):
        candidates = raw_value
    else:
        candidates = str(raw_value).split(",")

    emails: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        email = _quote_request_clean(candidate, 320)
        if not email or email in seen:
            continue
        seen.add(email)
        emails.append(email)
    return emails


def _quote_request_recipient_emails_from_settings(db: Session | None = None) -> list[str]:
    configured = _quote_request_parse_recipient_emails(QUOTE_REQUEST_RECIPIENT_EMAILS)
    if db is None:
        return configured

    settings = get_or_create_app_settings(db)
    stored_emails = _quote_request_parse_recipient_emails(getattr(settings, "quote_request_recipient_emails", None))
    return stored_emails or configured


def _quote_request_throttle_key(payload: QuoteRequestCreate, request: Request) -> str:
    client_ip = _quote_request_clean(payload.client_ip or _extract_client_ip(request), 64)
    if client_ip:
        return client_ip
    return "unknown"


def _quote_request_check_throttle(key: str) -> None:
    now = time.time()
    cutoff = now - QUOTE_REQUEST_THROTTLE_WINDOW_SECONDS

    with QUOTE_REQUEST_THROTTLE_LOCK:
        bucket = QUOTE_REQUEST_THROTTLE_STATE.setdefault(key, deque())
        while bucket and bucket[0] < cutoff:
            bucket.popleft()

        if len(bucket) >= QUOTE_REQUEST_THROTTLE_MAX_ATTEMPTS:
            raise HTTPException(
                status_code=429,
                detail="Too many enquiry submissions from this address. Please wait a little before trying again.",
            )

        bucket.append(now)


def _quote_request_clean(value: object, limit: int = 200) -> str:
    cleaned = re.sub(r"\s+", " ", html.unescape(str(value or ""))).strip()
    return cleaned[:limit]


def _quote_request_clean_multiline(value: object, limit: int = 4000) -> str:
    cleaned = html.unescape(str(value or ""))
    cleaned = cleaned.replace("\r\n", "\n").replace("\r", "\n")
    cleaned = re.sub(r"[ \t]+\n", "\n", cleaned).strip()
    return cleaned[:limit]


def _quote_request_subject(record: QuoteRequest) -> str:
    page_title = _quote_request_clean(record.page_card_title or record.page_title or "Vent-Tech quote request", 120)
    type_label = QUOTE_REQUEST_REQUEST_TYPE_LABELS.get(record.request_type or "", "Quote request")
    return f"Vent-Tech quote request - {page_title} - {type_label}"


def _quote_request_body(record: QuoteRequest) -> str:
    attributes = record.attributes if isinstance(record.attributes, list) else []
    attribute_labels = [QUOTE_REQUEST_ATTRIBUTE_LABELS.get(str(attribute), str(attribute)) for attribute in attributes if str(attribute).strip()]
    lines = [
        "Vent-Tech quote request",
        "",
        f"Status: {record.status}",
        f"Request type: {QUOTE_REQUEST_REQUEST_TYPE_LABELS.get(record.request_type or '', 'Quote request')}",
        f"Name: {record.name}",
        f"Company: {record.company or 'Not provided'}",
        f"Email: {record.email}",
        f"Phone: {record.phone or 'Not provided'}",
        f"Desired attributes: {', '.join(attribute_labels) if attribute_labels else 'Not specified'}",
        f"Airflow range: {record.airflow_min or 'Not specified'} - {record.airflow_max or 'Not specified'}",
        f"Pressure range: {record.pressure_min or 'Not specified'} - {record.pressure_max or 'Not specified'}",
        f"Power limit: {record.power_limit or 'Not specified'}",
        f"Current page: {record.page_url or 'Not provided'}",
        f"Page card: {(record.page_card_title or 'Not provided')} - {(record.page_card_summary or 'Not provided')}",
        f"Page type: {record.page_type or 'Not provided'}",
        f"Verification: {record.verification_provider} / {record.verification_status}",
        f"Email status: {record.email_status}",
    ]
    if record.short_notes:
        lines.extend(["", f"Additional notes: {_quote_request_clean_multiline(record.short_notes, 300)}"])
    if record.details:
        lines.extend(["", "Details:", _quote_request_clean_multiline(record.details, 4000)])
    if record.context_json:
        lines.extend(["", "Context:", json.dumps(record.context_json, ensure_ascii=False, indent=2)])
    return "\n".join(lines)


def _build_quote_request_email(record: QuoteRequest, recipient_emails: list[str]) -> EmailMessage:
    message = EmailMessage()
    recipient_emails = recipient_emails or ["admin@venttech.co.nz"]
    from_email = SMTP_FROM_EMAIL or SMTP_USERNAME or recipient_emails[0]
    from_name = SMTP_FROM_NAME or "Vent-Tech website"

    message["To"] = ", ".join(recipient_emails)
    message["From"] = f"{from_name} <{from_email}>"
    message["Reply-To"] = record.email
    message["Subject"] = _quote_request_subject(record)
    message.set_content(_quote_request_body(record))
    return message


def _build_quote_request_test_email(recipient_email: str) -> EmailMessage:
    message = EmailMessage()
    from_email = SMTP_FROM_EMAIL or SMTP_USERNAME or recipient_email
    from_name = SMTP_FROM_NAME or "Vent-Tech website"

    message["To"] = recipient_email
    message["From"] = f"{from_name} <{from_email}>"
    message["Subject"] = "Vent-Tech SMTP test"
    message.set_content(
        "This is a test email from the Vent-Tech enquiry system.\n\n"
        "If you received this message, SMTP delivery is working."
    )
    return message


def _send_quote_request_email(message: EmailMessage) -> None:
    if not SMTP_HOST:
        raise RuntimeError("SMTP_HOST is not configured.")

    context = ssl.create_default_context()
    if SMTP_USE_SSL:
        server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=20, context=context)
    else:
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20)

    try:
        server.ehlo()
        if SMTP_USE_TLS and not SMTP_USE_SSL:
            server.starttls(context=context)
            server.ehlo()
        if SMTP_USERNAME:
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.send_message(message)
    finally:
        try:
            server.quit()
        except Exception:
            server.close()


def _normalise_quote_request_payload(payload: QuoteRequestCreate) -> dict:
    attributes = [str(item).strip().casefold() for item in (payload.attributes or []) if str(item).strip()]
    request_type = str(payload.request_type or "unsure").strip().casefold()
    if request_type not in QUOTE_REQUEST_REQUEST_TYPE_LABELS:
        request_type = "unsure"

    context_json = {
        "page_type": payload.page_type or "",
        "page_title": payload.page_title or "",
        "page_summary": payload.page_summary or "",
        "page_card_title": payload.page_card_title or "",
        "page_card_summary": payload.page_card_summary or "",
        "page_url": payload.page_url or "",
        "product_type": payload.product_type or {},
        "series": payload.series or {},
        "product": payload.product or {},
    }
    if isinstance(payload.page_context, dict):
        context_json.update({key: value for key, value in payload.page_context.items() if value not in (None, "")})

    return {
        "name": _quote_request_clean(payload.name, 120),
        "company": _quote_request_clean(payload.company, 160),
        "email": _quote_request_clean(payload.email, 160),
        "phone": _quote_request_clean(payload.phone, 80),
        "request_type": request_type,
        "attributes": attributes,
        "airflow_min": _quote_request_clean(payload.airflow_min, 40),
        "airflow_max": _quote_request_clean(payload.airflow_max, 40),
        "pressure_min": _quote_request_clean(payload.pressure_min, 40),
        "pressure_max": _quote_request_clean(payload.pressure_max, 40),
        "power_limit": _quote_request_clean(payload.power_limit, 60),
        "short_notes": _quote_request_clean_multiline(payload.short_notes, 300),
        "details": _quote_request_clean_multiline(payload.details, 4000),
        "page_type": _quote_request_clean(payload.page_type, 80),
        "page_title": _quote_request_clean(payload.page_title, 200),
        "page_summary": _quote_request_clean(payload.page_summary, 500),
        "page_card_title": _quote_request_clean(payload.page_card_title, 200),
        "page_card_summary": _quote_request_clean(payload.page_card_summary, 500),
        "page_url": _quote_request_clean(payload.page_url, 500),
        "client_ip": _quote_request_clean(payload.client_ip, 64),
        "user_agent": _quote_request_clean(payload.user_agent, 500),
        "referrer": _quote_request_clean(payload.referrer, 500),
        "origin": _quote_request_clean(payload.origin, 500),
        "context_json": context_json,
    }


bulk_import_router.add_api_route(
    "/api/bulk-import",
    bulk_import_assets,
    methods=["POST"],
    response_model=BulkImportResponse,
    dependencies=[Depends(get_current_user)],
    summary="Import workbook sheets and image assets",
)


def active_admin_count(db: Session) -> int:
    return db.query(User).filter(User.is_admin.is_(True), User.is_active.is_(True)).count()


def postgres_cli_database_url() -> str:
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL is not configured")
    if DATABASE_URL.startswith("postgresql+psycopg://"):
        return "postgresql://" + DATABASE_URL[len("postgresql+psycopg://"):]
    return DATABASE_URL


def run_command(command: list[str], *, input_bytes: bytes | None = None):
    result = subprocess.run(command, input=input_bytes, capture_output=True, check=False)
    if result.returncode != 0:
        stderr = result.stderr.decode("utf-8", errors="ignore").strip()
        stdout = result.stdout.decode("utf-8", errors="ignore").strip()
        raise RuntimeError(stderr or stdout or f"Command failed: {' '.join(command)}")
    return result


def postgres_tool_database_url() -> str:
    return postgres_cli_database_url()


def container_runtime_binary() -> str | None:
    for candidate in ("podman", "docker"):
        resolved = shutil.which(candidate)
        if resolved:
            return resolved
    return None


def run_postgres_client_tool(arguments: list[str], *, input_bytes: bytes | None = None):
    tool_name = arguments[0]
    if shutil.which(tool_name):
        direct_command = [tool_name, postgres_tool_database_url(), *arguments[1:]]
        return run_command(direct_command, input_bytes=input_bytes)

    runtime = container_runtime_binary()
    if not runtime:
        raise RuntimeError(
            f"{tool_name} is not installed and no container runtime (podman/docker) is available for fallback."
        )

    container_command = [
        runtime,
        "run",
        "--rm",
        "--network",
        "host",
        "-e",
        f"DATABASE_URL={postgres_tool_database_url()}",
        POSTGRES_CLIENT_IMAGE,
        "sh",
        "-lc",
        " ".join([tool_name, '"$DATABASE_URL"'] + [shlex.quote(arg) for arg in arguments[1:]]),
    ]
    return run_command(container_command, input_bytes=input_bytes)


def serialize_setup_log_entry(entry: dict) -> SetupLogEntryResponse:
    return SetupLogEntryResponse(**entry)


def get_recent_setup_log_entries(limit: int = 200) -> list[SetupLogEntryResponse]:
    with LOG_BUFFER_LOCK:
        entries = list(LOG_BUFFER)[-max(int(limit), 0) :]
    return [serialize_setup_log_entry(entry) for entry in entries]


def get_setup_log_entries_after_id(after_id: int) -> list[SetupLogEntryResponse]:
    with LOG_BUFFER_LOCK:
        entries = [entry for entry in LOG_BUFFER if entry["id"] > after_id]
    return [serialize_setup_log_entry(entry) for entry in entries]


def setup_log_sse_payload(entry: SetupLogEntryResponse) -> str:
    return f"event: log\ndata: {entry.model_dump_json()}\n\n"


PUBLIC_ACCESS_LOG_PREFIX = "public-access "


def parse_public_access_log_entry(entry: SetupLogEntryResponse) -> PublicAccessLogEntryResponse | None:
    message = entry.message or ""
    if not message.startswith(PUBLIC_ACCESS_LOG_PREFIX):
        return None

    payload_text = message[len(PUBLIC_ACCESS_LOG_PREFIX) :].strip()
    payload: dict[str, object] = {}
    if payload_text:
        try:
            parsed = json.loads(payload_text)
            if isinstance(parsed, dict):
                payload = parsed
        except json.JSONDecodeError:
            payload = {"raw": payload_text}

    return PublicAccessLogEntryResponse(**entry.model_dump(), payload=payload)


def get_recent_public_access_log_entries(limit: int = 200, *, site: str | None = None, route_group: str | None = None) -> list[PublicAccessLogEntryResponse]:
    with LOG_BUFFER_LOCK:
        entries = list(LOG_BUFFER)[-max(int(limit), 0) :]

    filtered: list[PublicAccessLogEntryResponse] = []
    for entry in entries:
        parsed = parse_public_access_log_entry(serialize_setup_log_entry(entry))
        if parsed is None:
            continue
        payload_site = str(parsed.payload.get("site") or "")
        payload_route_group = str(parsed.payload.get("route_group") or "")
        if site and payload_site != site:
            continue
        if route_group and payload_route_group != route_group:
            continue
        filtered.append(parsed)
    return filtered


async def fetch_customer_facing_recent_logs(limit: int = 200, public_only: bool = False) -> list[PublicAccessLogEntryResponse]:
    if not PUBLIC_CATALOGUE_SITE_URL or not CMS_API_TOKEN:
        raise HTTPException(status_code=503, detail="Customer-facing logs are unavailable because the site token is not configured.")

    query = urllib.parse.urlencode(
        {
            "limit": max(int(limit), 1),
            "public_only": "true" if public_only else "false",
        }
    )
    url = f"{PUBLIC_CATALOGUE_SITE_URL}/api/logs/recent?{query}"

    headers = {"Authorization": f"Bearer {CMS_API_TOKEN}"}
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            response = await client.get(url, headers=headers)
            response.raise_for_status()
    except httpx.HTTPStatusError as exc:
        detail = exc.response.text.strip() or f"Customer-facing log request failed with status {exc.response.status_code}."
        raise HTTPException(status_code=exc.response.status_code, detail=detail) from exc
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail="Customer-facing logs are unavailable right now.") from exc

    try:
        payload = response.json()
    except ValueError as exc:
        raise HTTPException(status_code=502, detail="Customer-facing logs returned an invalid response.") from exc
    if not isinstance(payload, list):
        raise HTTPException(status_code=502, detail="Customer-facing log payload was invalid.")

    entries: list[PublicAccessLogEntryResponse] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        try:
            entries.append(PublicAccessLogEntryResponse(**item))
        except Exception:
            continue
    return entries


def _copy_media_directories(staging_data_dir: Path, progress_callback=None, *, label_prefix: str = "Collecting", exclude_backup_dir: bool = False):
    backup_stage_total = len(DATA_BACKUP_DIRS) + 1
    for offset, media_dir in enumerate(DATA_BACKUP_DIRS, start=1):
        if exclude_backup_dir and media_dir.name == "backups":
            continue
        if media_dir.is_dir():
            if progress_callback:
                progress_callback(f"{label_prefix} {media_dir.name}", offset, backup_stage_total)
            target_dir = staging_data_dir / media_dir.name
            target_dir.parent.mkdir(parents=True, exist_ok=True)
            shutil.copytree(media_dir, target_dir, dirs_exist_ok=True)


def _write_zip_archive(source_dir: Path, archive_path: Path) -> None:
    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for root, _, files in os.walk(source_dir):
            for file_name in files:
                full_path = Path(root) / file_name
                archive.write(full_path, full_path.relative_to(source_dir))


def _write_backup_readme(staging_dir: Path, title: str, contents: list[str]) -> None:
    readme = staging_dir / "README.txt"
    readme.write_text(
        "\n".join(
            [
                title,
                "Generated by the admin maintenance API.",
                "",
                "Contents:",
                *contents,
            ]
        ),
        encoding="utf-8",
    )


def _copy_directory_to_staging(source_dir: Path, staging_root: Path, relative_target: str, progress_message: str | None = None, progress_callback=None, progress_step: int | None = None, progress_total: int | None = None):
    if not source_dir.is_dir():
        return
    if progress_callback and progress_message is not None and progress_step is not None and progress_total is not None:
        progress_callback(progress_message, progress_step, progress_total)
    destination = staging_root / relative_target
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copytree(source_dir, destination, dirs_exist_ok=True)


def create_maintenance_job(job_type: str) -> dict:
    job_id = uuid4().hex
    job = {
        "id": job_id,
        "job_type": job_type,
        "status": "queued",
        "progress_message": "Queued",
        "progress_current": None,
        "progress_total": None,
        "progress_percent": None,
        "error": None,
        "result_message": None,
        "result_download_url": None,
        "created_at": backend_now_iso(),
        "started_at": None,
        "completed_at": None,
    }
    with MAINTENANCE_JOBS_LOCK:
        MAINTENANCE_JOBS[job_id] = job
    return job


def update_maintenance_job(job_id: str, **updates):
    with MAINTENANCE_JOBS_LOCK:
        job = MAINTENANCE_JOBS.get(job_id)
        if not job:
            return
        job.update(updates)


def get_maintenance_job_or_404(job_id: str) -> dict:
    with MAINTENANCE_JOBS_LOCK:
        job = MAINTENANCE_JOBS.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Maintenance job not found")
        return dict(job)


def serialize_maintenance_job(job: dict) -> MaintenanceJobResponse:
    return MaintenanceJobResponse(**job)


def raise_job_phase_error(entity_label: str, phase_label: str, exc: Exception) -> None:
    raise RuntimeError(f"{entity_label}: {phase_label} failed: {exc}") from exc


def start_maintenance_job(job_type: str, work):
    job = create_maintenance_job(job_type)
    job_label = f"{job_type}/{job['id']}"

    def runner():
        logger.info("[maintenance:%s] queued", job_label)
        update_maintenance_job(job["id"], status="running", started_at=backend_now_iso(), progress_message="Starting")
        logger.info("[maintenance:%s] starting", job_label)

        def progress(message: str, current: int | None = None, total: int | None = None):
            updates = {"progress_message": message}
            if current is not None:
                updates["progress_current"] = current
            if total is not None:
                updates["progress_total"] = total
            if current is not None and total not in (None, 0):
                updates["progress_percent"] = round((current / total) * 100, 1)
            elif total == 0:
                updates["progress_percent"] = 100.0
            update_maintenance_job(job["id"], **updates)
            if current is not None and total is not None:
                progress_percent = updates.get("progress_percent")
                if progress_percent is not None:
                    logger.info("[maintenance:%s] %s (%s/%s, %.1f%%)", job_label, message, current, total, progress_percent)
                else:
                    logger.info("[maintenance:%s] %s (%s/%s)", job_label, message, current, total)
            else:
                logger.info("[maintenance:%s] %s", job_label, message)

        try:
            result = work(progress) or {}
            result_updates = dict(result)
            result_updates.setdefault("progress_message", result.get("progress_message") or "Completed")
            result_updates.setdefault("result_message", result.get("result_message"))
            result_updates.setdefault("result_download_url", result.get("result_download_url"))
            result_updates["progress_percent"] = 100.0
            result_updates["status"] = "completed"
            result_updates["completed_at"] = backend_now_iso()
            update_maintenance_job(job["id"], **result_updates)
            logger.info("[maintenance:%s] completed with result: %s", job_label, result_updates.get("result_message") or "Completed")
        except Exception as exc:
            logger.exception("[maintenance:%s] failed", job_label)
            update_maintenance_job(
                job["id"],
                status="failed",
                error=str(exc),
                progress_message="Failed",
                completed_at=backend_now_iso(),
            )
            logger.error("[maintenance:%s] failed with error: %s", job_label, exc)

    thread = threading.Thread(target=runner, daemon=True, name=f"maintenance-{job['id']}")
    thread.start()
    return job


def _make_progress_window(progress_callback, start_percent: int, end_percent: int):
    start_percent = max(0, min(100, int(start_percent)))
    end_percent = max(start_percent, min(100, int(end_percent)))
    span = max(0, end_percent - start_percent)

    def windowed_progress(message: str, current: int | None = None, total: int | None = None):
        if progress_callback is None:
            return
        if current is None or total in (None, 0):
            progress_callback(message, start_percent, 100)
            return
        if total <= 1 or span == 0:
            mapped = end_percent
        else:
            mapped = start_percent + round(((current - 1) / (total - 1)) * span)
        progress_callback(message, mapped, 100)

    return windowed_progress


def _make_indexed_progress_window(progress_callback, index: int, total: int):
    total = max(int(total or 0), 1)
    index = max(1, min(int(index or 1), total))
    start_percent = round(((index - 1) / total) * 100)
    end_percent = 100 if index >= total else round((index / total) * 100)
    if end_percent < start_percent:
        end_percent = start_percent
    return _make_progress_window(progress_callback, start_percent, end_percent)


def create_data_backup_bundle(progress_callback=None) -> Path:
    timestamp = backend_now().strftime("%Y%m%d_%H%M%S")
    archive_name = f"fan_graphs_media_data_backup_{timestamp}.zip"
    archive_path = BACKUP_OUTPUT_DIR / archive_name
    backup_stages_total = len(DATA_BACKUP_DIRS) + 2

    with tempfile.TemporaryDirectory() as staging_dir_raw:
        staging_dir = Path(staging_dir_raw)
        if progress_callback:
            progress_callback("Collecting media files", 1, backup_stages_total)
        data_dir = staging_dir / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        _copy_media_directories(data_dir, progress_callback, label_prefix="Collecting", exclude_backup_dir=True)
        _copy_directory_to_staging(
            TEMPLATES_DIR,
            staging_dir,
            "templates",
            "Collecting templates",
            progress_callback,
            len(DATA_BACKUP_DIRS) + 1,
            backup_stages_total,
        )

        _write_backup_readme(
            staging_dir,
            "Internal Facing media data backup archive",
            [
                *[f"- data/{name} : media assets and generated files (if present)" for name in DATA_BACKUP_DIR_NAMES],
                "- templates : template files and registry (if present)",
            ],
        )

        if progress_callback:
            progress_callback("Creating data archive", backup_stages_total, backup_stages_total)
        _write_zip_archive(staging_dir, archive_path)

    return archive_path


def create_database_backup_bundle(progress_callback=None) -> Path:
    timestamp = backend_now().strftime("%Y%m%d_%H%M%S")
    archive_name = f"fan_graphs_db_data_backup_{timestamp}.zip"
    archive_path = BACKUP_OUTPUT_DIR / archive_name
    backup_stages_total = 2

    with tempfile.TemporaryDirectory() as staging_dir_raw:
        staging_dir = Path(staging_dir_raw)
        postgres_dump_path = staging_dir / "postgres_dump.sql"
        if progress_callback:
            progress_callback("Creating PostgreSQL dump", 1, backup_stages_total)
        postgres_dump = run_postgres_client_tool(["pg_dump", "--no-owner", "--no-privileges"])
        postgres_dump_path.write_bytes(postgres_dump.stdout)

        _write_backup_readme(
            staging_dir,
            "Internal Facing DB data backup archive",
            [
                "- postgres_dump.sql : PostgreSQL database dump",
            ],
        )

        if progress_callback:
            progress_callback("Creating database archive", 2, backup_stages_total)
        _write_zip_archive(staging_dir, archive_path)

    return archive_path


def restore_backup_bundle(archive_bytes: bytes, progress_callback=None):
    restore_stages_total = len(DATA_BACKUP_DIR_NAMES) + 3
    with tempfile.TemporaryDirectory() as staging_dir_raw:
        staging_dir = Path(staging_dir_raw)
        archive_path = staging_dir / "upload.zip"
        archive_path.write_bytes(archive_bytes)
        if progress_callback:
            progress_callback("Extracting backup archive", 1, restore_stages_total)
        with zipfile.ZipFile(archive_path, "r") as archive:
            archive.extractall(staging_dir)

        postgres_dump_path = staging_dir / "postgres_dump.sql"
        if not postgres_dump_path.is_file():
            raise RuntimeError("Backup archive does not contain postgres_dump.sql")

        if progress_callback:
            progress_callback("Resetting PostgreSQL schema", 2, restore_stages_total)
        run_postgres_client_tool(
            [
                "psql",
                "-v",
                "ON_ERROR_STOP=1",
                "-c",
                f"SELECT pg_terminate_backend(pid) FROM pg_stat_activity WHERE datname = current_database() AND pid <> pg_backend_pid(); DROP SCHEMA public CASCADE; CREATE SCHEMA public; GRANT ALL ON SCHEMA public TO {POSTGRES_USER}; GRANT ALL ON SCHEMA public TO public;",
            ]
        )
        if progress_callback:
            progress_callback("Importing PostgreSQL dump", 3, restore_stages_total)
        run_postgres_client_tool(
            ["psql", "-v", "ON_ERROR_STOP=1"],
            input_bytes=postgres_dump_path.read_bytes(),
        )

        for offset, media_dir in enumerate(
            DATA_BACKUP_DIR_NAMES,
            start=4,
        ):
            source_dir = staging_dir / "data" / media_dir
            target_dir = Path(DEFAULT_DATA_DIR) / media_dir
            if source_dir.is_dir():
                if progress_callback:
                    progress_callback(f"Restoring {media_dir}", offset, restore_stages_total)
                shutil.rmtree(target_dir, ignore_errors=True)
                target_dir.mkdir(parents=True, exist_ok=True)
                shutil.copytree(source_dir, target_dir, dirs_exist_ok=True)


def restore_media_backup_bundle(archive_bytes: bytes, progress_callback=None):
    restore_stages_total = len(DATA_BACKUP_DIRS) + 2
    with tempfile.TemporaryDirectory() as staging_dir_raw:
        staging_dir = Path(staging_dir_raw)
        archive_path = staging_dir / "upload.zip"
        archive_path.write_bytes(archive_bytes)
        if progress_callback:
            progress_callback("Extracting media archive", 1, restore_stages_total)
        with zipfile.ZipFile(archive_path, "r") as archive:
            archive.extractall(staging_dir)

        for offset, media_dir in enumerate(DATA_BACKUP_DIR_NAMES, start=2):
            source_dir = staging_dir / "data" / media_dir
            target_dir = Path(DEFAULT_DATA_DIR) / media_dir
            if source_dir.is_dir():
                if progress_callback:
                    progress_callback(f"Restoring {media_dir}", offset, restore_stages_total)
                shutil.rmtree(target_dir, ignore_errors=True)
                target_dir.mkdir(parents=True, exist_ok=True)
                shutil.copytree(source_dir, target_dir, dirs_exist_ok=True)

        templates_source = staging_dir / "templates"
        if templates_source.is_dir():
            if progress_callback:
                progress_callback("Restoring templates", restore_stages_total - 1, restore_stages_total)
            shutil.rmtree(TEMPLATES_DIR, ignore_errors=True)
            TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
            shutil.copytree(templates_source, TEMPLATES_DIR, dirs_exist_ok=True)
            sync_template_registry_with_disk()


def run_post_restore_schema_prep(progress_callback=None):
    if progress_callback:
        progress_callback("Running database migrations", 9, 9)
    from backend.db_management import prepare_configured_databases

    prepare_configured_databases()
    init_db()


OPENAPI_TAGS = [
    {
        "name": "Public",
        "description": "Unauthenticated health and authentication bootstrap endpoints.",
    },
    {
        "name": "Authentication",
        "description": "Staff login and session management endpoints.",
    },
    {
        "name": "Users",
        "description": "Staff user administration endpoints.",
    },
    {
        "name": "Public Catalog",
        "description": "Read-only product, product type, and series endpoints used by the public customer-facing site.",
    },
    {
        "name": "Public Media",
        "description": "Public customer-facing product image, graph, and PDF file endpoints intended for rendered website pages.",
    },
    {
        "name": "Products",
        "description": "Product catalogue CRUD endpoints for the internal app.",
    },
    {
        "name": "Product Types",
        "description": "Product type definitions and seeded parameter preset libraries.",
    },
    {
        "name": "Series",
        "description": "Series records that group products within a product type.",
    },
    {
        "name": "Templates",
        "description": "Controlled-list template definitions used for product and series PDF/layout selection.",
    },
    {
        "name": "RPM Lines",
        "description": "RPM line management for a graph-capable product.",
    },
    {
        "name": "RPM Points",
        "description": "RPM curve point management for a graph-capable product.",
    },
    {
        "name": "Efficiency Points",
        "description": "Efficiency curve point management for a graph-capable product.",
    },
    {
        "name": "Product Images",
        "description": "Product image upload, ordering, and deletion endpoints.",
    },
    {
        "name": "Series Images",
        "description": "Series image upload, ordering, and deletion endpoints.",
    },
    {
        "name": "Media",
        "description": "Protected internal media endpoints for staff-only direct access.",
    },
    {
        "name": "Maintenance",
        "description": "Operational and data maintenance endpoints.",
    },
]

app = FastAPI(
    title="Internal Facing API",
    description=(
        "Product platform API for the Internal Facing application.\n\n"
        "Use `/api/products...` for the internal staff application.\n"
        "Use `/api/public/products...` and `/api/public/series...` for the customer-facing public site.\n"
        "Use `/api/public/media/...` for public customer-facing product images and graph files.\n"
        "Legacy `/api/fans...` aliases still work, but they are intentionally hidden from the schema."
    ),
    openapi_tags=OPENAPI_TAGS,
    docs_url=None,
    redoc_url=None,
    openapi_url=None,
)

app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET or "dev-session-secret-change-me",
    same_site="lax",
    https_only=AUTH_COOKIE_SECURE,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:8001",
        "http://127.0.0.1:8001",
        "http://xps.local:8001",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(bulk_import_router)


@app.post(
    "/api/graph-data/parse",
    dependencies=[Depends(get_current_user)],
    tags=["Products"],
    summary="Parse a graph CSV or workbook",
)
async def parse_graph_data_upload(file: UploadFile = File(...)):
    raw_name = normalize_bulk_import_source_name(file.filename or "")
    contents = await file.read()
    rows = load_graph_import_rows(raw_name, contents)
    if len(rows) < 2:
        raise HTTPException(
            status_code=400,
            detail="Choose a graph data file with a header row and at least one data row.",
        )
    return {
        "file_name": raw_name,
        "rows": rows,
    }


@app.on_event("startup")
def startup():
    ensure_auth_config()
    init_db()
    ensure_bootstrap_admin()


@app.middleware("http")
async def log_public_requests(request: Request, call_next):
    started = time.perf_counter()
    response = None
    try:
        response = await call_next(request)
        return response
    finally:
        if _extract_public_client(request).get("public_host"):
            status_code = getattr(response, "status_code", 500)
            duration_ms = (time.perf_counter() - started) * 1000.0
            _log_request_event("internal", request, status_code=status_code, duration_ms=duration_ms)


# --- Health ---
@app.get("/api/health", tags=["Public"])
def health():
    return {"ok": True}


@app.post("/api/client-telemetry", tags=["Public"])
async def client_telemetry(request: Request):
    try:
        payload = await request.json()
    except Exception:
        payload = {}

    page_url = str(payload.get("page_url") or "")
    parsed_page = urllib.parse.urlparse(page_url)
    telemetry = {
        "event": "browser-telemetry",
        "site": "internal",
        "route_group": "internal-browser-telemetry" if parsed_page.path.startswith(("/editor", "/viewer", "/template-builder", "/bulk-import", "/setup")) else "public-browser-telemetry",
        "logged_at": datetime.datetime.now(tz=APP_TIMEZONE).isoformat(timespec="seconds"),
        **_extract_public_client(request),
        "telemetry": {
            "page_url": page_url,
            "referrer": str(payload.get("referrer") or ""),
            "screen_width": payload.get("screen_width"),
            "screen_height": payload.get("screen_height"),
            "viewport_width": payload.get("viewport_width"),
            "viewport_height": payload.get("viewport_height"),
            "device_pixel_ratio": payload.get("device_pixel_ratio"),
            "color_depth": payload.get("color_depth"),
            "timezone": str(payload.get("timezone") or ""),
            "timezone_offset": payload.get("timezone_offset"),
            "language": str(payload.get("language") or ""),
            "languages": payload.get("languages") or [],
            "platform": str(payload.get("platform") or ""),
            "user_agent": str(payload.get("user_agent") or request.headers.get("user-agent", "")),
            "device_type": str(payload.get("device_type") or ""),
            "touch_points": payload.get("touch_points"),
        },
    }
    logger.info("public-access %s", json.dumps(telemetry, ensure_ascii=True, sort_keys=True, separators=(",", ":")))
    return {"ok": True}


@app.get("/openapi.json", dependencies=[Depends(get_current_user)])
def openapi_schema():
    return app.openapi()


@app.get("/docs", include_in_schema=False, dependencies=[Depends(get_current_user)])
def swagger_ui():
    return get_swagger_ui_html(openapi_url="/openapi.json", title="Internal Facing API Docs")


@app.get("/api/product-types", response_model=list[ProductTypeResponse], dependencies=[Depends(get_current_user)], tags=["Product Types"])
def list_product_types(db: Session = Depends(get_db)):
    return (
        db.query(ProductType)
        .options(
            selectinload(ProductType.series).selectinload(Series.products),
            selectinload(ProductType.products),
            selectinload(ProductType.parameter_group_presets).selectinload(
                ProductTypeParameterGroupPreset.parameter_presets
            ),
            selectinload(ProductType.rpm_line_presets).selectinload(ProductTypeRpmLinePreset.point_presets),
            selectinload(ProductType.efficiency_point_presets),
        )
        .order_by(ProductType.sort_order, ProductType.id)
        .all()
    )


@app.get(
    "/api/public/product-types",
    response_model=list[ProductTypeResponse],
    tags=["Public Catalog"],
    summary="List public product types",
    description="Returns product types for the public catalog navigation.",
)
def list_public_product_types(db: Session = Depends(get_db)):
    return (
        db.query(ProductType)
        .options(
            selectinload(ProductType.series).selectinload(Series.products),
            selectinload(ProductType.products),
            selectinload(ProductType.parameter_group_presets).selectinload(
                ProductTypeParameterGroupPreset.parameter_presets
            ),
            selectinload(ProductType.rpm_line_presets).selectinload(ProductTypeRpmLinePreset.point_presets),
            selectinload(ProductType.efficiency_point_presets),
        )
        .order_by(ProductType.sort_order, ProductType.id)
        .all()
    )


@app.post("/api/product-types", response_model=ProductTypeResponse, dependencies=[Depends(get_current_user)], tags=["Product Types"])
def create_product_type(body: ProductTypeCreate, db: Session = Depends(get_db)):
    label = (body.label or "").strip()
    if not label:
        raise HTTPException(status_code=400, detail="Product type label is required.")

    key = sanitize_name((body.key or "").strip() or label)
    if not key:
        raise HTTPException(status_code=400, detail="Product type key is required.")

    existing = db.query(ProductType).filter(ProductType.key == key).first()
    if existing:
        raise HTTPException(status_code=400, detail="A product type with that key already exists.")

    next_sort_order = db.query(func.coalesce(func.max(ProductType.sort_order), -1)).scalar()

    printed_product_template_id, online_product_template_id = resolve_template_pair(
        "product",
        body.product_template_id,
        body.printed_product_template_id,
        body.online_product_template_id,
    )

    product_type = ProductType(
        key=key,
        label=label,
        sort_order=int(next_sort_order) + 1,
        supports_graph=bool(body.supports_graph),
        graph_kind=(body.graph_kind or "").strip() or None,
        supports_graph_overlays=bool(body.supports_graph_overlays),
        supports_band_graph_style=bool(body.supports_band_graph_style),
        graph_line_value_label=(body.graph_line_value_label or "").strip() or None,
        graph_line_value_unit=(body.graph_line_value_unit or "").strip() or None,
        graph_x_axis_label=(body.graph_x_axis_label or "").strip() or None,
        graph_x_axis_unit=(body.graph_x_axis_unit or "").strip() or None,
        graph_y_axis_label=(body.graph_y_axis_label or "").strip() or None,
        graph_y_axis_unit=(body.graph_y_axis_unit or "").strip() or None,
        product_type_template_id=validate_template_id(body.product_type_template_id, "product_type"),
        series_template_id=validate_template_id(body.series_template_id, "series"),
        printed_product_template_id=printed_product_template_id,
        online_product_template_id=online_product_template_id,
        product_template_id=online_product_template_id or printed_product_template_id,
        contents_icon_url=(body.contents_icon_url or "").strip() or None,
        band_graph_background_color=normalize_color_value(body.band_graph_background_color),
        band_graph_label_text_color=normalize_color_value(body.band_graph_label_text_color),
        band_graph_faded_opacity=None if body.band_graph_faded_opacity is None else max(0, min(1, float(body.band_graph_faded_opacity))),
        band_graph_permissible_label_color=normalize_color_value(body.band_graph_permissible_label_color),
    )
    db.add(product_type)
    db.commit()
    db.refresh(product_type)
    notify_public_catalogue_cache_refresh()
    return product_type


@app.put("/api/product-types/{product_type_id}", response_model=ProductTypeResponse, dependencies=[Depends(get_current_user)], tags=["Product Types"])
def update_product_type(product_type_id: int, body: ProductTypeUpdate, db: Session = Depends(get_db)):
    product_type = db.query(ProductType).filter(ProductType.id == product_type_id).first()
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found.")

    updates = body.model_dump(exclude_unset=True)

    if "key" in updates:
        new_key = sanitize_name((updates.get("key") or "").strip())
        if not new_key:
            raise HTTPException(status_code=400, detail="Product type key cannot be blank.")
        duplicate = db.query(ProductType).filter(ProductType.key == new_key, ProductType.id != product_type_id).first()
        if duplicate:
            raise HTTPException(status_code=400, detail="A product type with that key already exists.")
        product_type.key = new_key

    if "label" in updates:
        label = (updates.get("label") or "").strip()
        if not label:
            raise HTTPException(status_code=400, detail="Product type label cannot be blank.")
        product_type.label = label

    if "series_template_id" in updates:
        product_type.series_template_id = validate_template_id(updates.pop("series_template_id"), "series")

    if any(field in updates for field in ("product_template_id", "printed_product_template_id", "online_product_template_id")):
        printed_product_template_id, online_product_template_id = resolve_template_pair(
            "product",
            updates.pop("product_template_id", None),
            updates.pop("printed_product_template_id", None),
            updates.pop("online_product_template_id", None),
        )
        if body.model_fields_set.intersection({"product_template_id", "printed_product_template_id"}):
            product_type.printed_product_template_id = printed_product_template_id
        if body.model_fields_set.intersection({"product_template_id", "online_product_template_id"}):
            product_type.online_product_template_id = online_product_template_id
        product_type.product_template_id = (
            product_type.online_product_template_id
            or product_type.printed_product_template_id
        )

    if "product_type_template_id" in updates:
        product_type.product_type_template_id = validate_template_id(updates["product_type_template_id"], "product_type")
    if "contents_icon_url" in updates:
        product_type.contents_icon_url = (updates["contents_icon_url"] or "").strip() or None

    for field in [
        "supports_graph",
        "supports_graph_overlays",
        "supports_band_graph_style",
        "graph_kind",
        "graph_line_value_label",
        "graph_line_value_unit",
        "graph_x_axis_label",
        "graph_x_axis_unit",
        "graph_y_axis_label",
        "graph_y_axis_unit",
        "band_graph_background_color",
        "band_graph_label_text_color",
        "band_graph_faded_opacity",
        "band_graph_permissible_label_color",
    ]:
        if field in updates:
            value = updates[field]
            if isinstance(value, str):
                value = value.strip() or None
            if field in {"band_graph_background_color", "band_graph_label_text_color", "band_graph_permissible_label_color"}:
                value = normalize_color_value(value)
            elif field == "band_graph_faded_opacity":
                value = None if value is None else max(0, min(1, float(value)))
            setattr(product_type, field, value)

    db.commit()
    db.refresh(product_type)
    notify_public_catalogue_cache_refresh()
    return product_type


@app.delete("/api/product-types/{product_type_id}", dependencies=[Depends(get_current_user)], tags=["Product Types"], summary="Delete a product type")
def delete_product_type(product_type_id: int, db: Session = Depends(get_db)):
    product_type = (
        db.query(ProductType)
        .options(
            selectinload(ProductType.series).selectinload(Series.products),
            selectinload(ProductType.products),
        )
        .filter(ProductType.id == product_type_id)
        .first()
    )
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found")

    delete_product_type_assets(product_type)

    for series in list(product_type.series):
        for product in list(series.products):
            sync_product_series(product, None)

    for product in list(product_type.products):
        product.product_type = None

    db.delete(product_type)
    db.commit()
    notify_public_catalogue_cache_refresh()
    return {"deleted": product_type_id}


@app.put(
    "/api/product-types/{product_type_id}/parameter-group-presets",
    response_model=ProductTypeResponse,
    dependencies=[Depends(get_current_user)],
    tags=["Product Types"],
    summary="Replace presets for a product type",
)
def update_product_type_parameter_group_presets(
    product_type_id: int,
    body: ProductTypePresetUpdate,
    db: Session = Depends(get_db),
):
    product_type = (
        db.query(ProductType)
        .options(
            selectinload(ProductType.parameter_group_presets).selectinload(
                ProductTypeParameterGroupPreset.parameter_presets
            ),
            selectinload(ProductType.rpm_line_presets).selectinload(ProductTypeRpmLinePreset.point_presets),
            selectinload(ProductType.efficiency_point_presets),
        )
        .filter(ProductType.id == product_type_id)
        .first()
    )
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found.")

    apply_product_type_presets(
        product_type,
        body.parameter_group_presets,
        body.rpm_line_presets,
        body.efficiency_point_presets,
        body.product_template_id,
        body.series_template_id,
        body.printed_product_template_id,
        body.online_product_template_id,
    )
    db.commit()
    db.refresh(product_type)
    notify_public_catalogue_cache_refresh()
    return product_type


@app.get(
    "/api/product-types/{product_type_id}/pdf-context",
    response_model=ProductTypePdfResponse,
    dependencies=[Depends(get_current_user)],
    tags=["Product Types"],
    summary="Inspect the product type PDF data context",
)
def get_product_type_pdf_context(product_type_id: int, db: Session = Depends(get_db)):
    product_type = (
        db.query(ProductType)
        .options(
            selectinload(ProductType.series).selectinload(Series.products).joinedload(Product.product_type),
        )
        .filter(ProductType.id == product_type_id)
        .first()
    )
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found.")

    with tempfile.TemporaryDirectory(prefix="product-type-context-") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        metadata = build_product_type_pdf_context_metadata(product_type, temp_dir)

    return ProductTypePdfResponse(
        id=product_type.id,
        key=product_type.key,
        label=product_type.label,
        series_names=product_type.series_names,
        series_names_html=metadata["series_names_html"],
        series_groups_html=metadata["series_groups_html"],
        contents_html=metadata["contents_html"],
        contents_icon_url=product_type.contents_icon_url,
        intro_page_count=metadata["intro_page_count"],
        page_count=metadata["page_count"],
        product_type_pdf_url=product_type.product_type_pdf_url,
        product_type_printed_pdf_url=product_type.product_type_printed_pdf_url,
        series=metadata["series_summaries"],
    )


@app.post(
    "/api/product-types/{product_type_id}/pdf/refresh",
    response_model=ProductTypePdfResponse,
    dependencies=[Depends(get_current_user)],
    tags=["Product Types"],
    summary="Generate the product type PDF",
)
def refresh_product_type_pdf(product_type_id: int, db: Session = Depends(get_db)):
    product_type = (
        db.query(ProductType)
        .options(
            selectinload(ProductType.series)
            .selectinload(Series.products)
            .selectinload(Product.product_images),
            selectinload(ProductType.series).selectinload(Series.products).selectinload(Product.product_type),
            selectinload(ProductType.series).selectinload(Series.products).selectinload(Product.series),
        )
        .filter(ProductType.id == product_type_id)
        .first()
    )
    if not product_type:
        raise HTTPException(status_code=404, detail="Product type not found.")

    try:
        _, metadata = generate_product_type_pdf_with_metadata(product_type)
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=f"Unable to generate product type PDF: {exc}") from exc

    notify_public_catalogue_cache_refresh()
    return ProductTypePdfResponse(
        id=product_type.id,
        key=product_type.key,
        label=product_type.label,
        series_names=product_type.series_names,
        series_names_html=metadata["series_names_html"],
        series_groups_html=metadata["series_groups_html"],
        contents_html=metadata["contents_html"],
        contents_icon_url=product_type.contents_icon_url,
        intro_page_count=metadata["intro_page_count"],
        page_count=metadata["page_count"],
        product_type_pdf_url=product_type.product_type_pdf_url,
        product_type_printed_pdf_url=product_type.product_type_printed_pdf_url,
        series=metadata["series_summaries"],
    )


@app.post(
    "/api/maintenance/jobs/product-types/{product_type_id}/pdf/refresh",
    response_model=MaintenanceJobResponse,
    dependencies=[Depends(get_current_user)],
    tags=["Maintenance"],
    summary="Start generating a product type PDF",
)
def start_refresh_product_type_pdf_job(product_type_id: int):
    with SessionLocal() as db:
        product_type = (
            db.query(ProductType)
            .options(
                selectinload(ProductType.series).selectinload(Series.products).joinedload(Product.product_type),
            )
            .filter(ProductType.id == product_type_id)
            .first()
        )
        if not product_type:
            raise HTTPException(status_code=404, detail="Product type not found.")
        product_type_label = product_type.label

    def work(progress):
        progress(f"Loading product type data for {product_type_label}", 1, 100)
        with SessionLocal() as db:
            product_type = (
                db.query(ProductType)
                .options(
                    selectinload(ProductType.series).selectinload(Series.products).joinedload(Product.product_type),
                )
                .filter(ProductType.id == product_type_id)
                .first()
            )
            if not product_type:
                raise HTTPException(status_code=404, detail="Product type not found.")

            try:
                progress(
                    f"Reusing existing series PDFs and building the product type contents page for {product_type.label}",
                    15,
                    100,
                )
                generate_product_type_pdf(product_type, progress_callback=_make_progress_window(progress, 15, 95))
            except Exception as exc:
                logger.exception("[maintenance:refresh_product_type_pdf_%s] render failed for %s", product_type_id, product_type.label)
                raise_job_phase_error(f"Product type {product_type.label}", "PDF rendering", exc)

            progress(f"Merging existing series PDFs into the final product type PDF for {product_type.label}", 96, 100)
            db.refresh(product_type)
            db.commit()

        progress(f"Refreshing product type context for {product_type_label}", 100, 100)
        notify_public_catalogue_cache_refresh()
        return {
            "result_message": f"Generated product type PDF for {product_type_label}.",
            "progress_message": f"Generated product type PDF for {product_type_label}.",
            "progress_current": 100,
            "progress_total": 100,
            "progress_percent": 100.0,
        }

    return serialize_maintenance_job(start_maintenance_job(f"refresh_product_type_pdf_{product_type_id}", work))


@app.get("/api/templates", response_model=TemplateRegistryResponse, dependencies=[Depends(get_current_user)], tags=["Templates"], summary="List available templates")
def list_templates():
    return sync_template_registry_with_disk()


@app.post("/api/templates/refresh", response_model=TemplateRegistryResponse, dependencies=[Depends(get_current_user)], tags=["Templates"], summary="Refresh template registry from disk")
def refresh_templates():
    return sync_template_registry_with_disk()


@app.post("/api/templates", response_model=TemplateRegistryResponse, dependencies=[Depends(get_current_user)], tags=["Templates"], summary="Create a new template")
def create_template(body: TemplateCreateRequest):
    template_type = (body.template_type or "").strip().lower()
    template_dir = template_type_directory(template_type)
    registry = sync_template_registry_with_disk()
    collection_name = template_collection_name(template_type)
    existing_ids = {str(item.get("id")).strip() for item in registry.get(collection_name, []) if isinstance(item, dict)}

    label = (body.label or "").strip()
    if not label:
        raise HTTPException(status_code=400, detail="Template label is required.")

    template_id = (body.template_id or "").strip() or f"{template_type}-{sanitize_name(label)}"
    if template_id in existing_ids:
        raise HTTPException(status_code=400, detail="A template with that id already exists.")

    directory_slug = sanitize_name(template_id.replace(f"{template_type}-", "", 1))
    destination_dir = template_dir / directory_slug
    if destination_dir.exists():
        raise HTTPException(status_code=400, detail="A template directory with that name already exists.")

    source_template_id = (body.source_template_id or "").strip() or None
    if source_template_id:
        source_definition = get_template_definition(source_template_id, template_type)
        if source_definition is None:
            raise HTTPException(status_code=404, detail="Source template not found.")
        source_path = Path(__file__).resolve().parents[1] / source_definition["path"]
        if not source_path.is_file():
            raise HTTPException(status_code=404, detail="Source template files are missing on disk.")
        shutil.copytree(source_path.parent, destination_dir)
    else:
        scaffold_blank_template(template_type, destination_dir)

    registry = sync_template_registry_with_disk()
    for item in registry.get(collection_name, []):
        if item.get("path") == str((destination_dir / "template.html").relative_to(Path(__file__).resolve().parents[1])):
            item["id"] = template_id
            item["label"] = label
            item["type"] = template_type
            break
    save_template_registry(registry)
    return sync_template_registry_with_disk()


@app.delete("/api/templates/{template_type}/{template_id}", response_model=TemplateRegistryResponse, dependencies=[Depends(get_current_user)], tags=["Templates"], summary="Delete a template")
def delete_template(template_type: str, template_id: str, db: Session = Depends(get_db)):
    normalized_type = (template_type or "").strip().lower()
    registry = sync_template_registry_with_disk()
    collection_name = template_collection_name(normalized_type)
    entry = next(
        (
            item
            for item in registry.get(collection_name, [])
            if isinstance(item, dict) and str(item.get("id")).strip() == template_id
        ),
        None,
    )
    if entry is None:
        raise HTTPException(status_code=404, detail="Template not found.")

    if normalized_type == "product":
        in_use = (
            db.query(Product)
            .filter(
                (Product.template_id == template_id)
                | (Product.printed_template_id == template_id)
                | (Product.online_template_id == template_id)
            )
            .count()
            + db.query(ProductType)
            .filter(
                (ProductType.product_template_id == template_id)
                | (ProductType.printed_product_template_id == template_id)
                | (ProductType.online_product_template_id == template_id)
            )
            .count()
        )
    elif normalized_type == "series":
        in_use = (
            db.query(Series)
            .filter(
                (Series.template_id == template_id)
                | (Series.printed_template_id == template_id)
                | (Series.online_template_id == template_id)
            )
            .count()
        )
    elif normalized_type == "product_type":
        in_use = 1 if template_id == "product_type-default" else db.query(ProductType).filter(ProductType.product_type_template_id == template_id).count()
    else:
        raise HTTPException(status_code=400, detail="Template type must be 'product', 'series', or 'product_type'.")

    if in_use:
        raise HTTPException(status_code=400, detail="Template is still assigned. Reassign records before deleting it.")

    template_path = Path(__file__).resolve().parents[1] / str(entry.get("path") or "")
    template_dir = template_path.parent
    if template_dir.exists():
        shutil.rmtree(template_dir, ignore_errors=True)

    registry[collection_name] = [
        item
        for item in registry.get(collection_name, [])
        if not (isinstance(item, dict) and str(item.get("id")).strip() == template_id)
    ]
    save_template_registry(registry)
    return sync_template_registry_with_disk()


@app.get(
    "/api/templates/{template_type}/{template_id}/files",
    response_model=TemplateFileResponse,
    dependencies=[Depends(get_current_user)],
    tags=["Templates"],
    summary="Load template source files",
)
def get_template_files(template_type: str, template_id: str):
    normalized_type = (template_type or "").strip().lower()
    if normalized_type not in {"product", "series", "product_type"}:
        raise HTTPException(status_code=400, detail="Template type must be 'product', 'series', or 'product_type'.")

    template_definition = require_template_definition(template_id, normalized_type)
    template_path = Path(__file__).resolve().parents[1] / str(template_definition.get("path") or "")
    if not template_path.is_file():
        raise HTTPException(status_code=404, detail="Template HTML file is missing.")

    stylesheet_path = resolve_template_stylesheet_path(template_definition, template_path)

    return TemplateFileResponse(
        id=str(template_definition.get("id") or template_id),
        label=str(template_definition.get("label") or template_id),
        type=normalized_type,
        html_path=str(template_path.relative_to(Path(__file__).resolve().parents[1])),
        css_path=str(stylesheet_path.relative_to(Path(__file__).resolve().parents[1])) if stylesheet_path else None,
        html_content=template_path.read_text(encoding="utf-8"),
        css_content=stylesheet_path.read_text(encoding="utf-8") if stylesheet_path and stylesheet_path.is_file() else "",
    )


@app.put(
    "/api/templates/{template_type}/{template_id}/files",
    response_model=TemplateFileResponse,
    dependencies=[Depends(get_current_user)],
    tags=["Templates"],
    summary="Save template source files",
)
def update_template_files(template_type: str, template_id: str, body: TemplateFileUpdateRequest):
    normalized_type = (template_type or "").strip().lower()
    if normalized_type not in {"product", "series", "product_type"}:
        raise HTTPException(status_code=400, detail="Template type must be 'product', 'series', or 'product_type'.")

    template_definition = require_template_definition(template_id, normalized_type)
    template_path = Path(__file__).resolve().parents[1] / str(template_definition.get("path") or "")
    if not template_path.is_file():
        raise HTTPException(status_code=404, detail="Template HTML file is missing.")

    stylesheet_path = resolve_template_stylesheet_path(template_definition, template_path) or (template_path.parent / "template.css")

    formatted_html = format_html_source(body.html_content)
    formatted_css = format_css_source(body.css_content or "")

    template_path.write_text(formatted_html, encoding="utf-8")
    stylesheet_path.parent.mkdir(parents=True, exist_ok=True)
    stylesheet_path.write_text(formatted_css, encoding="utf-8")

    registry = sync_template_registry_with_disk()
    collection_name = template_collection_name(normalized_type)
    for item in registry.get(collection_name, []):
        if isinstance(item, dict) and str(item.get("id") or "").strip() == template_id:
            item["stylesheet"] = str(stylesheet_path.relative_to(Path(__file__).resolve().parents[1]))
            break
    save_template_registry(registry)

    return TemplateFileResponse(
        id=str(template_definition.get("id") or template_id),
        label=str(template_definition.get("label") or template_id),
        type=normalized_type,
        html_path=str(template_path.relative_to(Path(__file__).resolve().parents[1])),
        css_path=str(stylesheet_path.relative_to(Path(__file__).resolve().parents[1])),
        html_content=formatted_html,
        css_content=formatted_css,
    )


@app.post(
    "/api/templates/{template_type}/{template_id}/assets",
    response_model=TemplateAssetUploadResponse,
    dependencies=[Depends(get_current_user)],
    tags=["Templates"],
    summary="Upload a template asset",
)
def upload_template_asset(template_type: str, template_id: str, body: TemplateAssetUploadRequest):
    normalized_type = (template_type or "").strip().lower()
    if normalized_type not in {"product", "series", "product_type"}:
        raise HTTPException(status_code=400, detail="Template type must be 'product', 'series', or 'product_type'.")

    template_definition = require_template_definition(template_id, normalized_type)
    template_path = Path(__file__).resolve().parents[1] / str(template_definition.get("path") or "")
    if not template_path.is_file():
        raise HTTPException(status_code=404, detail="Template HTML file is missing.")

    file_name = Path(body.filename or "").name.strip()
    if not file_name or file_name in {".", ".."}:
        raise HTTPException(status_code=400, detail="A valid filename is required.")

    data_url = (body.data_url or "").strip()
    match = re.match(r"^data:([^;]+);base64,(.+)$", data_url, re.DOTALL)
    if not match:
        raise HTTPException(status_code=400, detail="A valid base64 data URL is required.")

    try:
        payload = base64.b64decode(match.group(2), validate=False)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Unable to decode the uploaded asset.") from exc

    asset_dir = template_path.parent / "assets"
    asset_dir.mkdir(parents=True, exist_ok=True)
    asset_path = asset_dir / file_name
    asset_path.write_bytes(payload)

    return TemplateAssetUploadResponse(
        filename=file_name,
        relative_path=str(asset_path.relative_to(Path(__file__).resolve().parents[1])),
        file_url=asset_path.resolve().as_uri(),
    )


@app.get("/api/file-manager/{root_name}", response_model=FileManagerListingResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="List file manager entries")
def list_file_manager_entries(root_name: str, path: str = ""):
    return file_manager_list_directory(root_name, path)


@app.get("/api/file-manager/{root_name}/download", dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Download a file manager file")
def download_file_manager_entry(root_name: str, path: str):
    target_path = file_manager_resolve_path(root_name, path)
    if not target_path.is_file():
        raise HTTPException(status_code=404, detail="File not found.")
    return FileResponse(target_path, filename=target_path.name)


@app.get("/api/file-manager/{root_name}/content", response_model=FileManagerContentResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Read a file manager file")
def get_file_manager_entry_content(root_name: str, path: str):
    return file_manager_read_text_file(root_name, path)


@app.put("/api/file-manager/{root_name}/content", response_model=FileManagerContentResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Update a file manager file")
def update_file_manager_entry_content(root_name: str, path: str, body: FileManagerContentUpdateRequest):
    return file_manager_write_text_file(root_name, path, body.content)


@app.post("/api/file-manager/{root_name}/folders", response_model=FileManagerListingResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Create a file manager folder")
def create_file_manager_folder(root_name: str, path: str = "", body: FileManagerCreateFolderRequest | None = None):
    if body is None:
        raise HTTPException(status_code=400, detail="Folder name is required.")

    folder_name = (body.folder_name or "").strip()
    if not folder_name:
        raise HTTPException(status_code=400, detail="Folder name is required.")
    if Path(folder_name).is_absolute() or any(part in {"..", ""} for part in Path(folder_name).parts):
        raise HTTPException(status_code=400, detail="Folder name is invalid.")

    parent_path = file_manager_resolve_path(root_name, path)
    if not parent_path.exists() or not parent_path.is_dir():
        raise HTTPException(status_code=404, detail="Folder not found.")

    target = parent_path / folder_name
    if target.exists():
        raise HTTPException(status_code=409, detail="A folder with that name already exists.")
    target.mkdir(parents=True, exist_ok=False)
    if root_name.strip().lower() == "templates":
        sync_templates_after_file_change()
    return file_manager_list_directory(root_name, path)


@app.post("/api/file-manager/{root_name}/upload", response_model=FileManagerListingResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Upload files into the file manager")
async def upload_file_manager_entries(
    root_name: str,
    path: str = "",
    replace_existing: bool = False,
    files: list[UploadFile] = File(...),
):
    if not files:
        raise HTTPException(status_code=400, detail="Please choose at least one file.")

    target_dir = file_manager_resolve_path(root_name, path)
    if not target_dir.exists() or not target_dir.is_dir():
        raise HTTPException(status_code=404, detail="Folder not found.")

    uploads = []
    for upload in files:
        file_name = Path(upload.filename or "").name.strip()
        if not file_name or file_name in {".", ".."} or Path(file_name).is_absolute():
            raise HTTPException(status_code=400, detail="Each uploaded file must have a valid file name.")
        destination = target_dir / file_name
        if destination.exists() and not replace_existing:
            raise HTTPException(status_code=409, detail=f"File already exists: {file_name}")
        uploads.append((upload, destination))

    for upload, destination in uploads:
        contents = await upload.read()
        destination.parent.mkdir(parents=True, exist_ok=True)
        tmp_destination = destination.with_name(f".{destination.name}.uploading")
        tmp_destination.write_bytes(contents)
        tmp_destination.replace(destination)

    if root_name.strip().lower() == "templates":
        sync_templates_after_file_change()
    return file_manager_list_directory(root_name, path)


@app.put("/api/file-manager/{root_name}/rename", response_model=FileManagerListingResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Rename a file manager entry")
def rename_file_manager_entry(root_name: str, path: str, body: FileManagerRenameRequest):
    if not (path or "").strip():
        raise HTTPException(status_code=400, detail="Open a file or folder first.")
    target_path = file_manager_resolve_path(root_name, path)
    if not target_path.exists():
        raise HTTPException(status_code=404, detail="File or folder not found.")

    new_name = (body.new_name or "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="New name is required.")
    if Path(new_name).is_absolute() or any(part in {"..", ""} for part in Path(new_name).parts):
        raise HTTPException(status_code=400, detail="New name is invalid.")

    destination = target_path.parent / new_name
    if destination.exists():
        raise HTTPException(status_code=409, detail="A file or folder with that name already exists.")
    if file_manager_is_protected(root_name, path):
        raise HTTPException(status_code=400, detail="That item is protected and cannot be renamed.")

    target_path.rename(destination)
    if root_name.strip().lower() == "templates":
        sync_templates_after_file_change()
    return file_manager_list_directory(root_name, str(Path(path).parent) if path else "")


@app.delete("/api/file-manager/{root_name}", response_model=FileManagerListingResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Delete a file manager entry")
def delete_file_manager_entry(root_name: str, path: str, body: FileManagerDeleteRequest | None = None):
    if not (path or "").strip():
        raise HTTPException(status_code=400, detail="Open a file or folder first.")
    target_path = file_manager_resolve_path(root_name, path)
    if not target_path.exists():
        raise HTTPException(status_code=404, detail="File or folder not found.")
    if file_manager_is_protected(root_name, path):
        raise HTTPException(status_code=400, detail="That item is protected and cannot be deleted.")
    if target_path.is_dir():
        if body is None:
            body = FileManagerDeleteRequest()
        if not body.recursive and any(target_path.iterdir()):
            raise HTTPException(status_code=400, detail="Folder is not empty. Enable recursive delete to remove it.")
        shutil.rmtree(target_path)
    else:
        target_path.unlink()
    if root_name.strip().lower() == "templates":
        sync_templates_after_file_change()
    return file_manager_list_directory(root_name, str(Path(path).parent) if path else "")


@app.get("/api/series", response_model=list[SeriesResponse], dependencies=[Depends(get_current_user)], tags=["Series"], summary="List series")
def list_series(
    db: Session = Depends(get_db),
    product_type_key: Optional[str] = Query(None),
):
    q = db.query(Series).options(
        joinedload(Series.product_type),
        selectinload(Series.series_images),
    )
    if product_type_key:
        q = q.join(ProductType).filter(ProductType.key == product_type_key)
    results = q.order_by(Series.name).all()
    assign_series_product_counts(results, load_series_product_counts(db, [item.id for item in results]))
    return results


@app.get("/api/series/{series_id}", response_model=SeriesResponse, dependencies=[Depends(get_current_user)], tags=["Series"], summary="Get a series")
def get_series(series_id: int, db: Session = Depends(get_db)):
    series = load_series_graph_ready_series(db, series_id)
    return series_response_with_graph_payload(series)


@app.post("/api/series", response_model=SeriesResponse, dependencies=[Depends(get_current_user)], tags=["Series"], summary="Create a series")
def create_series(body: SeriesCreate, db: Session = Depends(get_db)):
    product_type = get_product_type_by_key(db, body.product_type_key)
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Series name is required.")
    existing = (
        db.query(Series)
        .filter(Series.product_type_id == product_type.id, Series.name.ilike(name))
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="A series with that name already exists for this product type.")
    printed_template_id, online_template_id = resolve_template_pair(
        "series",
        body.template_id,
        body.printed_template_id,
        body.online_template_id,
    )
    if printed_template_id is None and online_template_id is None:
        series_template_id = validate_template_id(product_type.series_template_id, "series")
        printed_template_id = series_template_id
        online_template_id = series_template_id

    series = Series(
        product_type_id=product_type.id,
        name=name,
        description1_html=body.description1_html,
        description2_html=body.description2_html,
        description3_html=body.description3_html,
        description4_html=body.description4_html,
        printed_template_id=printed_template_id,
        online_template_id=online_template_id,
        template_id=online_template_id or printed_template_id,
    )
    series.product_type = product_type
    db.add(series)
    db.flush()
    ensure_series_tab_color(db, series)
    db.commit()
    db.refresh(series)
    assign_series_product_counts([series], load_series_product_counts(db, [series.id]))
    notify_public_catalogue_cache_refresh()
    return series


@app.put("/api/series/{series_id}", response_model=SeriesResponse, dependencies=[Depends(get_current_user)], tags=["Series"], summary="Update a series")
def update_series(series_id: int, body: SeriesUpdate, db: Session = Depends(get_db)):
    series = db.get(Series, series_id)
    if not series:
        raise HTTPException(status_code=404, detail="Series not found")
    updates = body.model_dump(exclude_unset=True)
    if "product_type_key" in updates:
        product_type = get_product_type_by_key(db, updates.pop("product_type_key"))
        series.product_type_id = product_type.id
        series.product_type = product_type
    if "name" in updates:
        name = (updates["name"] or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Series name is required.")
        existing = (
            db.query(Series)
            .filter(
                Series.product_type_id == series.product_type_id,
                Series.name.ilike(name),
                Series.id != series_id,
            )
            .first()
        )
        if existing:
            raise HTTPException(status_code=400, detail="A series with that name already exists for this product type.")
        series.name = name
    for field in ("description1_html", "description2_html", "description3_html", "description4_html"):
        if field in updates:
            setattr(series, field, updates[field])
    if any(field in updates for field in ("template_id", "printed_template_id", "online_template_id")):
        printed_template_id, online_template_id = resolve_template_pair(
            "series",
            updates.get("template_id"),
            updates.get("printed_template_id"),
            updates.get("online_template_id"),
        )
        if "printed_template_id" in updates or "template_id" in updates:
            series.printed_template_id = printed_template_id
        if "online_template_id" in updates or "template_id" in updates:
            series.online_template_id = online_template_id
        series.template_id = series.online_template_id or series.printed_template_id
    for product in series.products:
        product.series_name = series.name
    ensure_series_tab_color(db, series)
    db.commit()
    db.refresh(series)

    # A series name/type change affects the graph payload and title for every
    # linked product. Refresh those product graph images, and the series graph
    # image when the series has graph-capable data.
    for product in list(series.products):
        sync_graph_image(product, list(product.rpm_lines), list(product.efficiency_points))
    if series.product_type and series.product_type.supports_graph and series_has_graph_capable_line_data(series):
        generate_series_graph(series)
    db.commit()
    assign_series_product_counts([series], load_series_product_counts(db, [series.id]))
    notify_public_catalogue_cache_refresh()
    return series


@app.delete("/api/series/{series_id}", dependencies=[Depends(get_current_user)], tags=["Series"], summary="Delete a series")
def delete_series(series_id: int, db: Session = Depends(get_db)):
    series = db.get(Series, series_id)
    if not series:
        raise HTTPException(status_code=404, detail="Series not found")
    delete_associated_document_files(series)
    for product in list(series.products):
        sync_product_series(product, None)
    db.delete(series)
    db.commit()
    notify_public_catalogue_cache_refresh()
    return {"deleted": series_id}


@app.post("/api/series/{series_id}/graph-image/refresh", response_model=SeriesResponse, dependencies=[Depends(get_current_user)], tags=["Series"], summary="Generate a series graph image")
def refresh_series_graph_image(series_id: int, db: Session = Depends(get_db)):
    series = load_series_graph_ready_series(db, series_id)
    try:
        generate_series_graph(series)
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=f"Unable to generate series graph: {exc}") from exc
    return series_response_with_graph_payload(series)


@app.post("/api/series/{series_id}/pdf/refresh", response_model=SeriesResponse, dependencies=[Depends(get_current_user)], tags=["Series"], summary="Generate a series PDF")
def refresh_series_pdf(series_id: int, db: Session = Depends(get_db)):
    series = load_series_graph_ready_series(db, series_id)
    try:
        if series.product_type and series.product_type.supports_graph and series_has_graph_capable_line_data(series):
            generate_series_graph(series)
        else:
            logger.info(
                "[series_pdf:%s] graph generation skipped because there is no graph-capable line data to plot",
                series_id,
            )
        generate_series_pdfs(series)
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=f"Unable to generate series PDF: {exc}") from exc
    notify_public_catalogue_cache_refresh()
    return series_response_with_graph_payload(series)


@app.post(
    "/api/maintenance/jobs/series/{series_id}/pdf/refresh",
    response_model=MaintenanceJobResponse,
    dependencies=[Depends(get_current_user)],
    tags=["Maintenance"],
    summary="Start generating a series PDF",
)
def start_refresh_series_pdf_job(series_id: int):
    with SessionLocal() as db:
        series = db.get(Series, series_id)
        if not series:
            raise HTTPException(status_code=404, detail="Series not found")
        series_name = series.name

    def work(progress):
        with SessionLocal() as db:
            series = db.get(Series, series_id)
            if not series:
                raise HTTPException(status_code=404, detail="Series not found")

            db.refresh(series)
            progress(f"Loading series data for {series.name}", 1, 100)
            if series.product_type and series.product_type.supports_graph and series_has_graph_capable_line_data(series):
                try:
                    progress(f"Refreshing series graph image for {series.name}", 10, 100)
                    generate_series_graph(series)
                except Exception as exc:
                    logger.exception("[maintenance:refresh_series_pdf_%s] graph generation failed for %s", series_id, series.name)
                    raise_job_phase_error(f"Series {series.name}", "graph image generation", exc)
            else:
                progress(f"Series graph not required for {series.name} because there is no graph-capable line data to plot", 10, 100)
                logger.info(
                    "[maintenance:refresh_series_pdf_%s] graph generation skipped for %s because there is no graph-capable line data to plot",
                    series_id,
                    series.name,
                )

            try:
                progress(f"Generating printed series PDF for {series.name}", 15, 100)
                generate_series_pdfs(series, progress_callback=_make_progress_window(progress, 15, 95))
                progress(f"Printed series PDF generated for {series.name}", 96, 100)
            except Exception as exc:
                logger.exception("[maintenance:refresh_series_pdf_%s] pdf rendering failed for %s", series_id, series.name)
                raise_job_phase_error(f"Series {series.name}", "PDF rendering", exc)
            db.commit()

        notify_public_catalogue_cache_refresh()
        return {
            "result_message": f"Generated series PDFs for {series_name}.",
            "progress_message": f"Generated series PDFs for {series_name}.",
            "progress_current": 100,
            "progress_total": 100,
            "progress_percent": 100.0,
        }

    return serialize_maintenance_job(start_maintenance_job(f"refresh_series_pdf_{series_id}", work))


def _serialize_quote_request(record: QuoteRequest) -> QuoteRequestResponse:
    return QuoteRequestResponse(
        id=record.id,
        created_at=record.created_at.isoformat() if record.created_at else None,
        updated_at=record.updated_at.isoformat() if record.updated_at else None,
        status=record.status,
        email_status=record.email_status,
        email_error=record.email_error,
        verification_provider=record.verification_provider,
        verification_status=record.verification_status,
        verification_error=record.verification_error,
        name=record.name,
        company=record.company,
        email=record.email,
        phone=record.phone,
        request_type=record.request_type,
        attributes=list(record.attributes or []),
        airflow_min=record.airflow_min,
        airflow_max=record.airflow_max,
        pressure_min=record.pressure_min,
        pressure_max=record.pressure_max,
        power_limit=record.power_limit,
        short_notes=record.short_notes,
        details=record.details,
        page_type=record.page_type,
        page_title=record.page_title,
        page_summary=record.page_summary,
        page_card_title=record.page_card_title,
        page_card_summary=record.page_card_summary,
        page_url=record.page_url,
        client_ip=record.client_ip,
        user_agent=record.user_agent,
        referrer=record.referrer,
        origin=record.origin,
        product_type=(record.context_json or {}).get("product_type") or {},
        series=(record.context_json or {}).get("series") or {},
        product=(record.context_json or {}).get("product") or {},
        context_json=record.context_json or {},
    )


@app.post("/api/quote-requests", response_model=QuoteRequestResponse, tags=["Public", "Enquiries"])
async def create_quote_request(body: QuoteRequestCreate, request: Request, db: Session = Depends(get_db)):
    if body.website:
        raise HTTPException(status_code=400, detail="Enquiry request rejected.")

    record_data = _normalise_quote_request_payload(body)
    recipient_emails = _quote_request_recipient_emails_from_settings(db)
    throttle_key = _quote_request_throttle_key(body, request)
    _quote_request_check_throttle(throttle_key)
    record = QuoteRequest(
        status="new",
        email_status="pending",
        email_error=None,
        verification_provider="honeypot",
        verification_status="passed",
        verification_error=None,
        **record_data,
    )
    db.add(record)
    db.flush()

    if not SMTP_HOST:
        record.email_status = "not_configured"
        record.email_error = "SMTP_HOST is not configured."
        logger.info("Quote request email skipped because SMTP is not configured")
    else:
        try:
            email_message = _build_quote_request_email(record, recipient_emails)
            await asyncio.to_thread(_send_quote_request_email, email_message)
        except Exception as exc:
            record.email_status = "failed"
            record.email_error = str(exc)
            logger.exception("Quote request email delivery failed")
        else:
            record.email_status = "sent"
            record.email_error = None

    db.commit()
    db.refresh(record)
    logger.info(
        "Stored quote request %s for %s (%s)",
        record.id,
        record.page_card_title or record.page_title or record.page_type,
        record.email,
    )
    return _serialize_quote_request(record)


@app.get("/api/settings/quote-request-notifications", response_model=QuoteRequestNotificationSettings, dependencies=[Depends(get_current_user)], tags=["Maintenance"])
def get_quote_request_notification_settings(db: Session = Depends(get_db)):
    return QuoteRequestNotificationSettings(
        quote_request_recipient_emails=_quote_request_recipient_emails_from_settings(db),
    )


@app.put("/api/settings/quote-request-notifications", response_model=QuoteRequestNotificationSettings, dependencies=[Depends(get_current_user)], tags=["Maintenance"])
def update_quote_request_notification_settings(body: QuoteRequestNotificationSettings, db: Session = Depends(get_db)):
    settings = get_or_create_app_settings(db)
    emails = _quote_request_parse_recipient_emails(body.quote_request_recipient_emails)
    settings.quote_request_recipient_emails = ", ".join(emails) if emails else None
    db.commit()
    db.refresh(settings)
    return QuoteRequestNotificationSettings(quote_request_recipient_emails=emails)


@app.post("/api/settings/quote-request-email-test", response_model=QuoteRequestEmailTestResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"])
async def send_quote_request_email_test(body: QuoteRequestEmailTestRequest):
    if not SMTP_HOST:
        raise HTTPException(status_code=400, detail="SMTP_HOST is not configured.")

    recipient_email = _quote_request_clean(body.recipient_email, 160)
    if not recipient_email:
        raise HTTPException(status_code=400, detail="Recipient email is required.")

    try:
        message = _build_quote_request_test_email(recipient_email)
        await asyncio.to_thread(_send_quote_request_email, message)
    except Exception as exc:
        logger.exception("SMTP test email delivery failed")
        raise HTTPException(status_code=500, detail=f"Unable to send SMTP test email: {exc}") from exc

    return QuoteRequestEmailTestResponse(
        message="SMTP test email sent.",
        recipient_email=recipient_email,
    )


@app.get("/api/quote-requests", response_model=list[QuoteRequestResponse], dependencies=[Depends(get_current_user)], tags=["Enquiries"], summary="List enquiries")
def list_quote_requests(
    db: Session = Depends(get_db),
    status: Optional[str] = Query(None),
    limit: int = Query(200, ge=1, le=1000),
):
    query = db.query(QuoteRequest)
    if status:
        query = query.filter(QuoteRequest.status == status)
    records = query.order_by(QuoteRequest.created_at.desc(), QuoteRequest.id.desc()).limit(limit).all()
    return [_serialize_quote_request(record) for record in records]


@app.patch("/api/quote-requests/{quote_request_id}", response_model=QuoteRequestResponse, dependencies=[Depends(get_current_user)], tags=["Enquiries"], summary="Update an enquiry")
def update_quote_request(
    quote_request_id: int,
    body: QuoteRequestStatusUpdate,
    db: Session = Depends(get_db),
):
    record = db.get(QuoteRequest, quote_request_id)
    if not record:
        raise HTTPException(status_code=404, detail="Enquiry not found")

    status = str(body.status or "").strip().casefold()
    if status not in {"new", "quoted", "closed"}:
        raise HTTPException(status_code=400, detail="Invalid enquiry status.")

    record.status = status
    db.commit()
    db.refresh(record)
    return _serialize_quote_request(record)


@app.delete("/api/quote-requests/{quote_request_id}", dependencies=[Depends(get_current_user)], tags=["Enquiries"], summary="Delete an enquiry")
def delete_quote_request(
    quote_request_id: int,
    db: Session = Depends(get_db),
):
    record = db.get(QuoteRequest, quote_request_id)
    if not record:
        raise HTTPException(status_code=404, detail="Enquiry not found")

    db.delete(record)
    db.commit()
    return {"deleted": True, "id": quote_request_id}


@app.get("/api/auth/session", response_model=AuthSessionResponse, tags=["Public", "Authentication"])
def get_auth_session(request: Request):
    if not is_authenticated(request):
        details = _extract_request_ip_details(request)
        device_details = _extract_device_ip_details(request)
        return AuthSessionResponse(
            authenticated=False,
            cookie_secure=AUTH_COOKIE_SECURE,
            client_ip_v4=details["public_ipv4"] or None,
            client_ip_v6=details["public_ipv6"] or None,
            client_ip=details["public_host"] or None,
            device_ip_v4=device_details["device_ip_v4"],
            device_ip_v6=device_details["device_ip_v6"],
            device_ip=device_details["device_ip"],
        )
    details = _extract_request_ip_details(request)
    device_details = _extract_device_ip_details(request)
    return AuthSessionResponse(
        authenticated=True,
        username=request.session.get("username"),
        is_admin=bool(request.session.get("is_admin")),
        cookie_secure=AUTH_COOKIE_SECURE,
        client_ip_v4=details["public_ipv4"] or None,
        client_ip_v6=details["public_ipv6"] or None,
        client_ip=details["public_host"] or None,
        device_ip_v4=device_details["device_ip_v4"],
        device_ip_v6=device_details["device_ip_v6"],
        device_ip=device_details["device_ip"],
    )


@app.post("/api/auth/login", response_model=AuthSessionResponse, tags=["Public", "Authentication"])
def login(body: LoginRequest, request: Request, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == body.username.strip()).first()
    if user is None or not user.is_active or not verify_password(body.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    request.session["authenticated"] = True
    request.session["user_id"] = user.id
    request.session["username"] = user.username
    request.session["is_admin"] = user.is_admin
    details = _extract_request_ip_details(request)
    device_details = _extract_device_ip_details(request)
    return AuthSessionResponse(
        authenticated=True,
        username=user.username,
        is_admin=user.is_admin,
        cookie_secure=AUTH_COOKIE_SECURE,
        client_ip_v4=details["public_ipv4"] or None,
        client_ip_v6=details["public_ipv6"] or None,
        client_ip=details["public_host"] or None,
        device_ip_v4=device_details["device_ip_v4"],
        device_ip_v6=device_details["device_ip_v6"],
        device_ip=device_details["device_ip"],
    )


@app.post("/api/auth/logout", response_model=AuthSessionResponse, tags=["Public", "Authentication"])
def logout(request: Request):
    request.session.clear()
    return AuthSessionResponse(authenticated=False, cookie_secure=AUTH_COOKIE_SECURE)


@app.post("/api/auth/change-password", response_model=AuthSessionResponse, tags=["Authentication"])
def change_password(
    body: AuthPasswordChangeRequest,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not verify_password(body.current_password, current_user.password_hash):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    current_user.password_hash = hash_password(body.new_password)
    db.commit()
    request.session["username"] = current_user.username
    request.session["is_admin"] = current_user.is_admin
    details = _extract_request_ip_details(request)
    device_details = _extract_device_ip_details(request)
    return AuthSessionResponse(
        authenticated=True,
        username=current_user.username,
        is_admin=current_user.is_admin,
        cookie_secure=AUTH_COOKIE_SECURE,
        client_ip_v4=details["public_ipv4"] or None,
        client_ip_v6=details["public_ipv6"] or None,
        client_ip=details["public_host"] or None,
        device_ip_v4=device_details["device_ip_v4"],
        device_ip_v6=device_details["device_ip_v6"],
        device_ip=device_details["device_ip"],
    )


@app.get("/api/users", response_model=list[UserResponse], tags=["Users"])
def list_users(_: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.query(User).order_by(User.username).all()


@app.post("/api/users", response_model=UserResponse, tags=["Users"])
def create_user(body: UserCreate, _: User = Depends(require_admin_user), db: Session = Depends(get_db)):
    username = body.username.strip()
    existing_user = db.query(User).filter(User.username == username).first()
    if existing_user is not None:
        raise HTTPException(status_code=400, detail="Username already exists")
    user = User(
        username=username,
        password_hash=hash_password(body.password),
        is_admin=body.is_admin,
        is_active=True,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@app.patch("/api/users/{user_id}", response_model=UserResponse, tags=["Users"])
def update_user(
    user_id: int,
    body: UserUpdate,
    current_user: User = Depends(require_admin_user),
    db: Session = Depends(get_db),
):
    user = db.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    updates = body.model_dump(exclude_unset=True)
    if "is_admin" in updates:
        if user.is_admin and not updates["is_admin"] and active_admin_count(db) <= 1:
            raise HTTPException(status_code=400, detail="At least one active admin account is required")
        user.is_admin = updates["is_admin"]
    if "is_active" in updates:
        if user.is_admin and not updates["is_active"] and active_admin_count(db) <= 1:
            raise HTTPException(status_code=400, detail="At least one active admin account is required")
        if user.id == current_user.id and not updates["is_active"]:
            raise HTTPException(status_code=400, detail="You cannot deactivate your own account")
        user.is_active = updates["is_active"]
    db.commit()
    db.refresh(user)
    return user


@app.put("/api/users/{user_id}/password", response_model=UserResponse, tags=["Users"])
def update_user_password(
    user_id: int,
    body: UserPasswordUpdate,
    _: User = Depends(require_admin_user),
    db: Session = Depends(get_db),
):
    user = db.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    user.password_hash = hash_password(body.password)
    db.commit()
    db.refresh(user)
    return user


@app.get("/api/settings/band-graph-style", response_model=BandGraphStyleSettings, dependencies=[Depends(get_current_user)], tags=["Maintenance"])
def get_band_graph_style_settings(db: Session = Depends(get_db)):
    return get_or_create_app_settings(db)


@app.put("/api/settings/band-graph-style", response_model=BandGraphStyleSettings, dependencies=[Depends(get_current_user)], tags=["Maintenance"])
def update_band_graph_style_settings(body: BandGraphStyleSettings, db: Session = Depends(get_db)):
    settings = get_or_create_app_settings(db)
    settings.band_graph_background_color = normalize_color_value(body.band_graph_background_color)
    settings.band_graph_label_text_color = normalize_color_value(body.band_graph_label_text_color)
    db.commit()
    db.refresh(settings)
    return settings


def _load_public_products(
    db: Session,
    search: Optional[str] = Query(None),
    product_type_key: Optional[str] = Query(None),
    series_id: Optional[int] = Query(None),
    series_name: Optional[str] = Query(None),
    parameter_filters: Optional[str] = Query(None),
):
    parsed_parameter_filters = parse_parameter_filters(parameter_filters)
    trace_product_filter(
        "product filter trace request: endpoint=/api/public/products search=%s product_type_key=%s series_id=%s series_name=%s raw_filters=%s parsed_filters=%s",
        search,
        product_type_key,
        series_id,
        series_name,
        parameter_filters,
        parsed_parameter_filters,
    )
    q = db.query(Product).options(
        joinedload(Product.product_type),
        joinedload(Product.series),
        selectinload(Product.product_images),
        selectinload(Product.parameter_groups).selectinload(ProductParameterGroup.parameters),
        selectinload(Product.rpm_lines).selectinload(RpmLine.points),
        selectinload(Product.efficiency_points),
    )
    search_joined = False
    if search and search.strip():
        s = f"%{search.strip()}%"
        q = q.join(ProductType, isouter=True).join(Series, isouter=True).filter(
            or_(
                Product.model.ilike(s),
                Product.series_name.ilike(s),
                Series.name.ilike(s),
                ProductType.key.ilike(s),
                ProductType.label.ilike(s),
            )
        )
        search_joined = True
    if product_type_key:
        if not search_joined:
            q = q.join(ProductType)
        q = q.filter(ProductType.key == product_type_key)
    if series_id is not None:
        q = q.filter(Product.series_id == series_id)
        if series_name:
            q = q.filter(Product.series_name.ilike(f"%{series_name}%"))
    results = q.order_by(Product.model).all()
    return [product for product in results if product_matches_parameter_filters(product, parsed_parameter_filters)]


@app.get("/api/public/products", response_model=list[CmsProductResponse], tags=["Public Catalog"], summary="List public products", description="Read-only product catalogue feed for the public Svelte site. Supports search, product type, series, and grouped-parameter filtering.")
def list_public_products(
    db: Session = Depends(get_db),
    search: Optional[str] = Query(None),
    product_type_key: Optional[str] = Query(None),
    series_id: Optional[int] = Query(None),
    series_name: Optional[str] = Query(None),
    parameter_filters: Optional[str] = Query(None),
):
    return _load_public_products(
        db,
        search=search,
        product_type_key=product_type_key,
        series_id=series_id,
        series_name=series_name,
        parameter_filters=parameter_filters,
    )


@app.get("/api/public/products/{product_identifier}", response_model=CmsProductResponse, tags=["Public Catalog"], summary="Get one public product", description="Returns a single product record, including grouped specifications and media URLs, for the public Svelte site.")
def get_public_product(product_identifier: str, db: Session = Depends(get_db)):
    normalized_identifier = (product_identifier or "").strip().lower()
    if not normalized_identifier:
        raise HTTPException(status_code=404, detail="Product not found")

    product_candidates = (
        db.query(Product)
        .options(joinedload(Product.product_type))
        .order_by(Product.id)
        .all()
    )
    matched_product = None
    for candidate in product_candidates:
        if normalized_identifier in {value.strip().lower() for value in product_public_identifier_candidates(candidate)}:
            matched_product = candidate
            break

    if not matched_product:
        raise HTTPException(status_code=404, detail="Product not found")

    product = (
        db.query(Product)
        .options(
            joinedload(Product.product_type),
            joinedload(Product.series),
            selectinload(Product.product_images),
            selectinload(Product.parameter_groups).selectinload(ProductParameterGroup.parameters),
            selectinload(Product.rpm_lines),
            selectinload(Product.rpm_lines).selectinload(RpmLine.points),
            selectinload(Product.efficiency_points),
        )
        .filter(Product.id == matched_product.id)
        .first()
    )
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


@app.get("/api/public/series/{series_identifier}", response_model=CmsSeriesResponse, tags=["Public Catalog"], summary="Get one public series", description="Returns a single series record for the customer-facing Svelte site, including its linked products and chart payload.")
def get_public_series(series_identifier: str, db: Session = Depends(get_db)):
    normalized_identifier = (series_identifier or "").strip().lower()
    if not normalized_identifier:
        raise HTTPException(status_code=404, detail="Series not found")

    series_candidates = (
        db.query(Series)
        .options(joinedload(Series.product_type))
        .order_by(Series.id)
        .all()
    )
    matched_series = None
    for candidate in series_candidates:
        if normalized_identifier in {value.strip().lower() for value in series_public_identifier_candidates(candidate)}:
            matched_series = candidate
            break

    if not matched_series:
        raise HTTPException(status_code=404, detail="Series not found")

    series = (
        db.query(Series)
        .options(
            joinedload(Series.product_type),
            selectinload(Series.series_images),
            selectinload(Series.products).selectinload(Product.rpm_lines).selectinload(RpmLine.points),
            selectinload(Series.products).selectinload(Product.parameter_groups).selectinload(ProductParameterGroup.parameters),
        )
        .filter(Series.id == matched_series.id)
        .first()
    )
    if not series:
        raise HTTPException(status_code=404, detail="Series not found")

    assign_series_product_counts([series], load_series_product_counts(db, [series.id]))
    response = CmsSeriesResponse.model_validate(series, from_attributes=True)
    return response.model_copy(
        update={
            "performance_table_html": render_series_performance_table_html(series),
            "series_graph_payload": build_series_graph_payload(series),
        }
    )


# --- Products CRUD ---
@app.get("/api/fans", response_model=list[ProductResponse], dependencies=[Depends(get_current_user)], tags=["Products"], include_in_schema=False)
@app.get("/api/products", response_model=list[ProductResponse], dependencies=[Depends(get_current_user)], tags=["Products"], summary="List internal products", description="Primary internal catalogue endpoint used by the Svelte staff application. Supports search, product type, series, and grouped-parameter filtering.")
def list_products(
    db: Session = Depends(get_db),
    search: Optional[str] = Query(None, description="Search model"),
    model: Optional[str] = Query(None),
    product_type_key: Optional[str] = Query(None),
    series_id: Optional[int] = Query(None),
    series_name: Optional[str] = Query(None),
    parameter_filters: Optional[str] = Query(None),
):
    parsed_parameter_filters = parse_parameter_filters(parameter_filters)
    q = db.query(Product).options(
        joinedload(Product.product_type),
        joinedload(Product.series),
        selectinload(Product.product_images),
        selectinload(Product.parameter_groups).selectinload(ProductParameterGroup.parameters),
        selectinload(Product.rpm_lines),
    )
    if search:
        s = f"%{search}%"
        q = q.filter(Product.model.ilike(s))
    if model:
        q = q.filter(Product.model.ilike(f"%{model}%"))
    if product_type_key:
        q = q.join(ProductType).filter(ProductType.key == product_type_key)
    if series_id is not None:
        q = q.filter(Product.series_id == series_id)
    if series_name:
        q = q.filter(Product.series_name.ilike(f"%{series_name}%"))
    results = q.order_by(Product.model).all()
    return [product for product in results if product_matches_parameter_filters(product, parsed_parameter_filters)]


@app.post("/api/fans", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Products"], include_in_schema=False)
@app.post("/api/products", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Products"], summary="Create a product")
def create_product(body: ProductCreate, db: Session = Depends(get_db)):
    product_data = body.model_dump()
    product_data["permissible_use_mode"] = normalize_permissible_use_mode(product_data.get("permissible_use_mode"))
    product_type = get_product_type_by_key(db, product_data.pop("product_type_key", "fan"))
    series = get_series_by_id(db, product_data.pop("series_id", None))
    parameter_groups = product_data.pop("parameter_groups", [])
    fan_acoustic_table = product_data.pop("fan_acoustic_table", None)
    rpm_line_presets = product_data.pop("rpm_lines", [])
    efficiency_point_presets = product_data.pop("efficiency_points", [])
    if not rpm_line_presets:
        rpm_line_presets = build_product_type_rpm_line_presets(product_type)
    if not efficiency_point_presets:
        efficiency_point_presets = build_product_type_efficiency_point_presets(product_type)
    band_graph_style_defaults = resolve_product_type_band_graph_style_defaults(product_type)
    product_data["band_graph_background_color"] = (
        normalize_color_value(product_data.get("band_graph_background_color"))
        or band_graph_style_defaults["band_graph_background_color"]
    )
    product_data["band_graph_label_text_color"] = (
        normalize_color_value(product_data.get("band_graph_label_text_color"))
        or band_graph_style_defaults["band_graph_label_text_color"]
    )
    product_data["band_graph_faded_opacity"] = (
        product_data.get("band_graph_faded_opacity")
        if product_data.get("band_graph_faded_opacity") is not None
        else band_graph_style_defaults["band_graph_faded_opacity"]
    )
    product_data["band_graph_permissible_label_color"] = (
        normalize_color_value(product_data.get("band_graph_permissible_label_color"))
        or band_graph_style_defaults["band_graph_permissible_label_color"]
    )
    printed_template_id, online_template_id = resolve_template_pair(
        "product",
        product_data.get("template_id"),
        product_data.get("printed_template_id"),
        product_data.get("online_template_id"),
    )
    product_data["printed_template_id"] = printed_template_id or resolve_product_type_default_template_id(product_type, "printed")
    product_data["online_template_id"] = online_template_id or resolve_product_type_default_template_id(product_type, "online")
    product_data["template_id"] = product_data["online_template_id"] or product_data["printed_template_id"]
    product_data["product_type_id"] = product_type.id
    if series is not None and series.product_type_id != product_type.id:
        raise HTTPException(status_code=400, detail="Selected series does not belong to the chosen product type.")
    product_data["series_name"] = series.name if series is not None else (product_data.get("series_name") or None)
    product = Product(**product_data)
    product.product_type = product_type
    sync_product_series(product, series)
    db.add(product)
    db.flush()
    sync_product_parameter_groups(product, parameter_groups)
    created_rpm_lines: list[RpmLine] = []
    for line_index, line in enumerate(rpm_line_presets or []):
        line_model = RpmLine(
            product_id=product.id,
            rpm=line.get("rpm"),
            band_color=normalize_color_value(line.get("band_color")) or None,
        )
        db.add(line_model)
        db.flush()
        created_rpm_lines.append(line_model)

        for point in line.get("points") or []:
            db.add(
                RpmPoint(
                    product_id=product.id,
                    rpm_line_id=line_model.id,
                    airflow=point.get("airflow"),
                    pressure=point.get("pressure"),
                )
            )

    for point_index, point in enumerate(efficiency_point_presets or []):
        db.add(
            EfficiencyPoint(
                product_id=product.id,
                airflow=point.get("airflow"),
                efficiency_centre=point.get("efficiency_centre"),
                efficiency_lower_end=point.get("efficiency_lower_end"),
                efficiency_higher_end=point.get("efficiency_higher_end"),
                permissible_use=point.get("permissible_use"),
            )
        )

    db.commit()
    db.refresh(product)
    if product.product_type_key == "fan":
        product.fan_acoustic_table = fan_acoustic_table
        db.commit()
        db.refresh(product)
    sync_graph_image(product, created_rpm_lines or list(product.rpm_lines), list(product.efficiency_points))
    db.commit()
    db.refresh(product)
    notify_public_catalogue_cache_refresh()
    return product


@app.get("/api/fans/{product_id}", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Products"], include_in_schema=False)
@app.get("/api/products/{product_id}", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Products"], summary="Get one product")
def get_product(product_id: int, db: Session = Depends(get_db)):
    product = (
        db.query(Product)
        .options(
            joinedload(Product.product_type),
            joinedload(Product.series),
            selectinload(Product.product_images),
            selectinload(Product.parameter_groups).selectinload(ProductParameterGroup.parameters),
            selectinload(Product.rpm_lines),
        )
        .filter(Product.id == product_id)
        .first()
    )
    if not product:
        raise HTTPException(404, "Product not found")
    return product


@app.put("/api/fans/{product_id}", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Products"], include_in_schema=False)
@app.put("/api/products/{product_id}", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Products"], summary="Replace a product")
def update_product(product_id: int, body: ProductUpdate, db: Session = Depends(get_db)):
    product = require_product(db, product_id)
    updates = body.model_dump(exclude_unset=True)
    fan_acoustic_table_specified = "fan_acoustic_table" in body.model_fields_set
    fan_acoustic_table = updates.pop("fan_acoustic_table", None) if fan_acoustic_table_specified else None
    next_product_type = product.product_type
    if "product_type_key" in updates:
        product_type = get_product_type_by_key(db, updates.pop("product_type_key"))
        product.product_type_id = product_type.id
        product.product_type = product_type
        next_product_type = product_type
        if product.series is not None and product.series.product_type_id != product_type.id:
            sync_product_series(product, None)
    if "series_id" in updates:
        series = get_series_by_id(db, updates.pop("series_id"))
        if series is not None and next_product_type is not None and series.product_type_id != next_product_type.id:
            raise HTTPException(status_code=400, detail="Selected series does not belong to the chosen product type.")
        sync_product_series(product, series)
    if "parameter_groups" in updates:
        sync_product_parameter_groups(product, updates.pop("parameter_groups"))
    if any(field in updates for field in ("template_id", "printed_template_id", "online_template_id")):
        printed_template_id, online_template_id = resolve_template_pair(
            "product",
            updates.pop("template_id", None),
            updates.pop("printed_template_id", None),
            updates.pop("online_template_id", None),
        )
        if body.model_fields_set.intersection({"template_id", "printed_template_id"}):
            product.printed_template_id = printed_template_id
        if body.model_fields_set.intersection({"template_id", "online_template_id"}):
            product.online_template_id = online_template_id
        product.template_id = product.online_template_id or product.printed_template_id
    for k, v in updates.items():
        if k == "permissible_use_mode":
            setattr(product, k, normalize_permissible_use_mode(v))
            continue
        if k in {"band_graph_background_color", "band_graph_label_text_color", "band_graph_permissible_label_color"}:
            setattr(product, k, normalize_color_value(v))
        elif k == "band_graph_faded_opacity":
            setattr(product, k, None if v is None else max(0, min(1, float(v))))
        else:
            setattr(product, k, v)
    if product.series is not None:
        product.series_name = product.series.name
    if product.product_type_key == "fan":
        if fan_acoustic_table_specified:
            product.fan_acoustic_table = fan_acoustic_table
        elif product._fan_acoustic_table_json is None:
            product.fan_acoustic_table = None
    else:
        product.fan_acoustic_table = None
    sync_product_image_files(product)
    db.commit()
    db.refresh(product)
    # Product metadata (including the series name, graph mode, and graph
    # styling) is part of the rendered graph payload, so every product save
    # must refresh the stored graph image rather than only graph-style saves.
    sync_graph_image(product, list(product.rpm_lines), list(product.efficiency_points))
    db.commit()
    db.refresh(product)
    notify_public_catalogue_cache_refresh()
    return product


def replace_product_graph_data(
    db: Session,
    product: Product,
    rpm_line_presets: list[dict] | None = None,
    efficiency_point_presets: list[dict] | None = None,
):
    def preset_value(item, key):
        if item is None:
            return None
        if isinstance(item, dict):
            return item.get(key)
        return getattr(item, key, None)

    product.rpm_lines.clear()
    product.efficiency_points.clear()
    db.flush()

    created_rpm_lines: list[RpmLine] = []
    for line in rpm_line_presets or []:
        line_model = RpmLine(
            product_id=product.id,
            rpm=preset_value(line, "rpm"),
            band_color=normalize_color_value(preset_value(line, "band_color")) or None,
        )
        db.add(line_model)
        db.flush()
        created_rpm_lines.append(line_model)

        for point in preset_value(line, "points") or []:
            db.add(
                RpmPoint(
                    product_id=product.id,
                    rpm_line_id=line_model.id,
                    airflow=preset_value(point, "airflow"),
                    pressure=preset_value(point, "pressure"),
                )
            )

    for point in efficiency_point_presets or []:
        db.add(
            EfficiencyPoint(
                product_id=product.id,
                airflow=preset_value(point, "airflow"),
                efficiency_centre=preset_value(point, "efficiency_centre"),
                efficiency_lower_end=preset_value(point, "efficiency_lower_end"),
                efficiency_higher_end=preset_value(point, "efficiency_higher_end"),
                permissible_use=preset_value(point, "permissible_use"),
            )
        )

    db.flush()
    sync_fan_acoustic_table_for_product(db, product)
    return created_rpm_lines


@app.patch("/api/fans/{product_id}", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Products"], include_in_schema=False)
@app.patch("/api/products/{product_id}", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Products"], summary="Partially update a product")
def patch_product(product_id: int, body: ProductUpdate, db: Session = Depends(get_db)):
    return update_product(product_id, body, db)


@app.delete("/api/fans/{product_id}", dependencies=[Depends(get_current_user)], tags=["Products"], include_in_schema=False)
@app.delete("/api/products/{product_id}", dependencies=[Depends(get_current_user)], tags=["Products"], summary="Delete a product")
def delete_product(product_id: int, db: Session = Depends(get_db)):
    product = require_product(db, product_id)
    delete_associated_document_files(product)
    delete_product_assets(product)
    db.delete(product)
    db.commit()
    notify_public_catalogue_cache_refresh()
    return {"deleted": product_id}


@app.put("/api/products/{product_id}/graph-data", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"], summary="Replace the product graph data")
def replace_product_graph_data_endpoint(product_id: int, body: ProductGraphDataReplace, db: Session = Depends(get_db)):
    product = require_product(db, product_id)
    replace_product_graph_data(db, product, body.rpm_lines, body.efficiency_points)
    db.commit()
    db.refresh(product)
    notify_public_catalogue_cache_refresh()
    return product


# --- RPM lines / points ---
@app.get("/api/fans/{product_id}/rpm-lines", response_model=list[RpmLineResponse], dependencies=[Depends(get_current_user)], tags=["RPM Lines"], include_in_schema=False)
@app.get("/api/products/{product_id}/rpm-lines", response_model=list[RpmLineResponse], dependencies=[Depends(get_current_user)], tags=["RPM Lines"])
def get_rpm_lines(product_id: int, db: Session = Depends(get_db)):
    require_product(db, product_id)
    return (
        db.query(RpmLine)
        .options(selectinload(RpmLine.points))
        .filter(RpmLine.product_id == product_id)
        .order_by(RpmLine.rpm)
        .all()
    )


@app.post("/api/fans/{product_id}/rpm-lines", response_model=RpmLineResponse, dependencies=[Depends(get_current_user)], tags=["RPM Lines"], include_in_schema=False)
@app.post("/api/products/{product_id}/rpm-lines", response_model=RpmLineResponse, dependencies=[Depends(get_current_user)], tags=["RPM Lines"])
def create_rpm_line(product_id: int, body: RpmLineCreate, db: Session = Depends(get_db)):
    product = require_product(db, product_id)
    existing = db.query(RpmLine).filter(RpmLine.product_id == product_id, RpmLine.rpm == body.rpm).first()
    if existing:
        raise HTTPException(400, "RPM line already exists")
    line = RpmLine(product_id=product_id, rpm=body.rpm, band_color=normalize_color_value(body.band_color))
    db.add(line)
    db.commit()
    db.refresh(product)
    sync_fan_acoustic_table_for_product(db, product)
    db.commit()
    refresh_graph_for_product(db, product)
    db.commit()
    db.refresh(line)
    notify_public_catalogue_cache_refresh()
    return line


@app.put("/api/fans/{product_id}/rpm-lines/{line_id}", response_model=RpmLineResponse, dependencies=[Depends(get_current_user)], tags=["RPM Lines"], include_in_schema=False)
@app.put("/api/products/{product_id}/rpm-lines/{line_id}", response_model=RpmLineResponse, dependencies=[Depends(get_current_user)], tags=["RPM Lines"])
def update_rpm_line(product_id: int, line_id: int, body: RpmLineUpdate, db: Session = Depends(get_db)):
    product = require_product(db, product_id)
    line = db.get(RpmLine, line_id)
    if not line or line.product_id != product_id:
        raise HTTPException(404, "RPM line not found")

    updates = body.model_dump(exclude_unset=True)
    if "rpm" in updates and updates["rpm"] is not None:
        existing = (
            db.query(RpmLine)
            .filter(RpmLine.product_id == product_id, RpmLine.rpm == updates["rpm"], RpmLine.id != line_id)
            .first()
        )
        if existing:
            raise HTTPException(400, "RPM line already exists")
        line.rpm = updates["rpm"]
    if "band_color" in updates:
        line.band_color = normalize_color_value(updates["band_color"])

    db.commit()
    db.refresh(product)
    sync_fan_acoustic_table_for_product(db, product)
    db.commit()
    refresh_graph_for_product(db, product)
    db.commit()
    db.refresh(line)
    notify_public_catalogue_cache_refresh()
    return line


@app.delete("/api/fans/{product_id}/rpm-lines/{line_id}", dependencies=[Depends(get_current_user)], tags=["RPM Lines"], include_in_schema=False)
@app.delete("/api/products/{product_id}/rpm-lines/{line_id}", dependencies=[Depends(get_current_user)], tags=["RPM Lines"])
def delete_rpm_line(product_id: int, line_id: int, db: Session = Depends(get_db)):
    product = require_product(db, product_id)
    line = db.get(RpmLine, line_id)
    if not line or line.product_id != product_id:
        raise HTTPException(404, "RPM line not found")
    db.delete(line)
    db.commit()
    db.refresh(product)
    sync_fan_acoustic_table_for_product(db, product)
    db.commit()
    refresh_graph_for_product(db, product)
    db.commit()
    notify_public_catalogue_cache_refresh()
    return {"deleted": line_id}


@app.get("/api/fans/{product_id}/rpm-points", response_model=list[RpmPointResponse], dependencies=[Depends(get_current_user)], tags=["RPM Points"], include_in_schema=False)
@app.get("/api/products/{product_id}/rpm-points", response_model=list[RpmPointResponse], dependencies=[Depends(get_current_user)], tags=["RPM Points"])
def get_rpm_points(product_id: int, db: Session = Depends(get_db)):
    require_product(db, product_id)
    return (
        db.query(RpmPoint)
        .options(joinedload(RpmPoint.rpm_line))
        .join(RpmLine, RpmPoint.rpm_line_id == RpmLine.id)
        .filter(RpmPoint.product_id == product_id)
        .order_by(RpmLine.rpm, RpmPoint.airflow)
        .all()
    )


@app.post("/api/fans/{product_id}/rpm-points", response_model=RpmPointResponse, dependencies=[Depends(get_current_user)], tags=["RPM Points"], include_in_schema=False)
@app.post("/api/products/{product_id}/rpm-points", response_model=RpmPointResponse, dependencies=[Depends(get_current_user)], tags=["RPM Points"])
def create_rpm_point(
    product_id: int,
    body: RpmPointCreate,
    regenerate_graph: bool = Query(True),
    db: Session = Depends(get_db),
):
    product = require_product(db, product_id)
    line = db.get(RpmLine, body.rpm_line_id)
    if not line or line.product_id != product_id:
        raise HTTPException(404, "RPM line not found")
    point = RpmPoint(product_id=product_id, **body.model_dump())
    db.add(point)
    db.commit()
    if regenerate_graph:
        refresh_graph_for_product(db, product)
        db.commit()
    db.refresh(point)
    notify_public_catalogue_cache_refresh()
    return point


@app.put("/api/fans/{product_id}/rpm-points/{point_id}", response_model=RpmPointResponse, dependencies=[Depends(get_current_user)], tags=["RPM Points"], include_in_schema=False)
@app.put("/api/products/{product_id}/rpm-points/{point_id}", response_model=RpmPointResponse, dependencies=[Depends(get_current_user)], tags=["RPM Points"])
def update_rpm_point(
    product_id: int,
    point_id: int,
    body: RpmPointCreate,
    regenerate_graph: bool = Query(True),
    db: Session = Depends(get_db),
):
    product = require_product(db, product_id)
    line = db.get(RpmLine, body.rpm_line_id)
    if not line or line.product_id != product_id:
        raise HTTPException(404, "RPM line not found")
    point = db.get(RpmPoint, point_id)
    if not point or point.product_id != product_id:
        raise HTTPException(404, "RPM point not found")
    for key, value in body.model_dump(exclude_unset=True).items():
        setattr(point, key, value)
    db.commit()
    if regenerate_graph:
        refresh_graph_for_product(db, product)
        db.commit()
    db.refresh(point)
    notify_public_catalogue_cache_refresh()
    return point


@app.delete("/api/fans/{product_id}/rpm-points/{point_id}", dependencies=[Depends(get_current_user)], tags=["RPM Points"], include_in_schema=False)
@app.delete("/api/products/{product_id}/rpm-points/{point_id}", dependencies=[Depends(get_current_user)], tags=["RPM Points"])
def delete_rpm_point(
    product_id: int,
    point_id: int,
    regenerate_graph: bool = Query(True),
    db: Session = Depends(get_db),
):
    product = require_product(db, product_id)
    point = db.get(RpmPoint, point_id)
    if not point or point.product_id != product_id:
        raise HTTPException(404, "RPM point not found")
    db.delete(point)
    db.commit()
    if regenerate_graph:
        refresh_graph_for_product(db, product)
        db.commit()
    notify_public_catalogue_cache_refresh()
    return {"deleted": point_id}


@app.get("/api/fans/{product_id}/efficiency-points", response_model=list[EfficiencyPointResponse], dependencies=[Depends(get_current_user)], tags=["Efficiency Points"], include_in_schema=False)
@app.get("/api/products/{product_id}/efficiency-points", response_model=list[EfficiencyPointResponse], dependencies=[Depends(get_current_user)], tags=["Efficiency Points"])
def get_efficiency_points(product_id: int, db: Session = Depends(get_db)):
    require_product(db, product_id)
    return (
        db.query(EfficiencyPoint)
        .filter(EfficiencyPoint.product_id == product_id)
        .order_by(EfficiencyPoint.airflow)
        .all()
    )


@app.post("/api/fans/{product_id}/efficiency-points", response_model=EfficiencyPointResponse, dependencies=[Depends(get_current_user)], tags=["Efficiency Points"], include_in_schema=False)
@app.post("/api/products/{product_id}/efficiency-points", response_model=EfficiencyPointResponse, dependencies=[Depends(get_current_user)], tags=["Efficiency Points"])
def create_efficiency_point(
    product_id: int,
    body: EfficiencyPointCreate,
    regenerate_graph: bool = Query(True),
    db: Session = Depends(get_db),
):
    product = require_product(db, product_id)
    point = EfficiencyPoint(product_id=product_id, **body.model_dump())
    db.add(point)
    db.commit()
    if regenerate_graph:
        refresh_graph_for_product(db, product)
        db.commit()
    db.refresh(point)
    notify_public_catalogue_cache_refresh()
    return point


@app.put("/api/fans/{product_id}/efficiency-points/{point_id}", response_model=EfficiencyPointResponse, dependencies=[Depends(get_current_user)], tags=["Efficiency Points"], include_in_schema=False)
@app.put("/api/products/{product_id}/efficiency-points/{point_id}", response_model=EfficiencyPointResponse, dependencies=[Depends(get_current_user)], tags=["Efficiency Points"])
def update_efficiency_point(
    product_id: int,
    point_id: int,
    body: EfficiencyPointCreate,
    regenerate_graph: bool = Query(True),
    db: Session = Depends(get_db),
):
    product = require_product(db, product_id)
    point = db.get(EfficiencyPoint, point_id)
    if not point or point.product_id != product_id:
        raise HTTPException(404, "Efficiency point not found")
    for key, value in body.model_dump(exclude_unset=True).items():
        setattr(point, key, value)
    db.commit()
    if regenerate_graph:
        refresh_graph_for_product(db, product)
        db.commit()
    db.refresh(point)
    notify_public_catalogue_cache_refresh()
    return point


@app.delete("/api/fans/{product_id}/efficiency-points/{point_id}", dependencies=[Depends(get_current_user)], tags=["Efficiency Points"], include_in_schema=False)
@app.delete("/api/products/{product_id}/efficiency-points/{point_id}", dependencies=[Depends(get_current_user)], tags=["Efficiency Points"])
def delete_efficiency_point(
    product_id: int,
    point_id: int,
    regenerate_graph: bool = Query(True),
    db: Session = Depends(get_db),
):
    product = require_product(db, product_id)
    point = db.get(EfficiencyPoint, point_id)
    if not point or point.product_id != product_id:
        raise HTTPException(404, "Efficiency point not found")
    db.delete(point)
    db.commit()
    if regenerate_graph:
        refresh_graph_for_product(db, product)
        db.commit()
    notify_public_catalogue_cache_refresh()
    return {"deleted": point_id}


@app.post("/api/fans/{product_id}/graph-image/refresh", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"], include_in_schema=False)
@app.post("/api/products/{product_id}/graph-image/refresh", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"])
def refresh_product_graph_image(product_id: int, db: Session = Depends(get_db)):
    product = require_product(db, product_id)
    refresh_graph_for_product(db, product)
    db.commit()
    return product


@app.post("/api/fans/{product_id}/pdf/refresh", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"], include_in_schema=False)
@app.post("/api/products/{product_id}/pdf/refresh", response_model=ProductResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"], summary="Generate a product PDF")
def refresh_product_pdf(product_id: int, db: Session = Depends(get_db)):
    product = require_product(db, product_id)
    graph_path = Path(product.graph_image_path) if product.graph_image_path else None
    if product.product_type and product.product_type.supports_graph and not (graph_path and graph_path.is_file()):
        refresh_graph_for_product(db, product)
    try:
        generate_product_pdfs(product)
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=f"Unable to generate product PDF: {exc}") from exc
    db.commit()
    return product


@app.post(
    "/api/maintenance/jobs/products/{product_id}/pdf/refresh",
    response_model=MaintenanceJobResponse,
    dependencies=[Depends(get_current_user)],
    tags=["Maintenance"],
    summary="Start generating a product PDF",
)
def start_refresh_product_pdf_job(product_id: int):
    with SessionLocal() as db:
        product = require_product(db, product_id)
        product_label = product.model or f"product {product_id}"

    def work(progress):
        with SessionLocal() as db:
            product = require_product(db, product_id)
            progress(f"Loading product data for {product_label}", 1, 100)
            graph_path = Path(product.graph_image_path) if product.graph_image_path else None
            if product.product_type and product.product_type.supports_graph and not (graph_path and graph_path.is_file()):
                try:
                    progress(f"Refreshing product graph image for {product_label}", 10, 100)
                    refresh_graph_for_product(db, product)
                except Exception as exc:
                    logger.exception("[maintenance:refresh_product_pdf_%s] graph refresh failed for %s", product_id, product_label)
                    raise_job_phase_error(f"Product {product_label}", "graph image generation", exc)
            else:
                progress(f"Product graph image already available for {product_label}", 10, 100)

            try:
                progress(f"Rendering printed product PDF for {product_label}", 15, 100)
                generate_product_pdfs(product, progress_callback=_make_progress_window(progress, 15, 95))
                progress(f"Printed product PDF generated for {product_label}", 96, 100)
            except Exception as exc:
                logger.exception("[maintenance:refresh_product_pdf_%s] pdf rendering failed for %s", product_id, product_label)
                raise_job_phase_error(f"Product {product_label}", "PDF rendering", exc)
            db.commit()

        notify_public_catalogue_cache_refresh()
        return {
            "result_message": f"Generated printed PDF for {product_label}.",
            "progress_message": f"Generated printed PDF for {product_label}.",
            "progress_current": 100,
            "progress_total": 100,
            "progress_percent": 100.0,
        }

    return serialize_maintenance_job(start_maintenance_job(f"refresh_product_pdf_{product_id}", work))


@app.post("/api/maintenance/graph-images/regenerate-all", response_model=GraphImageMaintenanceResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"])
def regenerate_all_graph_images(db: Session = Depends(get_db)):
    products = db.query(Product).all()
    for product in products:
        refresh_graph_for_product(db, product)
    db.commit()
    return GraphImageMaintenanceResponse(
        message="Graph images regenerated.",
        products_processed=len(products),
        files_deleted=0,
    )


@app.post("/api/maintenance/jobs/graph-images/regenerate-all", response_model=MaintenanceJobResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"], summary="Start regenerating all graph images")
def start_regenerate_all_graph_images_job():
    def work(progress):
        with SessionLocal() as db:
            products = db.query(Product).all()
            total = len(products)
            for index, product in enumerate(products, start=1):
                progress(f"Generating graph {index} of {total}: {product.model}", index, total)
                refresh_graph_for_product(db, product)
            db.commit()
        return {
            "result_message": "Graph images regenerated.",
            "products_processed": total,
        }

    return serialize_maintenance_job(start_maintenance_job("regenerate_all_graph_images", work))


@app.post("/api/maintenance/product-pdfs/regenerate-all", response_model=PdfMaintenanceResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"])
def regenerate_all_product_pdfs(db: Session = Depends(get_db)):
    products = db.query(Product).options(joinedload(Product.product_type)).all()
    processed = 0
    for product in products:
        graph_path = Path(product.graph_image_path) if product.graph_image_path else None
        if product.product_type and product.product_type.supports_graph and not (graph_path and graph_path.is_file()):
            refresh_graph_for_product(db, product)
        generate_product_pdfs(product)
        processed += 1
    db.commit()
    return PdfMaintenanceResponse(
        message="Product PDFs regenerated.",
        products_processed=processed,
    )


@app.post("/api/maintenance/jobs/product-pdfs/regenerate-all", response_model=MaintenanceJobResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"], summary="Start regenerating all product PDFs")
def start_regenerate_all_product_pdfs_job():
    def work(progress):
        with SessionLocal() as db:
            products = db.query(Product).options(joinedload(Product.product_type)).all()
            total = len(products)
            processed = 0
            for index, product in enumerate(products, start=1):
                item_progress = _make_indexed_progress_window(progress, index, total)
                item_progress(f"Loading product {index} of {total}: {product.model}", 1, 100)
                graph_path = Path(product.graph_image_path) if product.graph_image_path else None
                if product.product_type and product.product_type.supports_graph and not (graph_path and graph_path.is_file()):
                    try:
                        item_progress(f"Refreshing graph image for {product.model}", 10, 100)
                        refresh_graph_for_product(db, product)
                    except Exception as exc:
                        logger.exception("[maintenance:regenerate_all_product_pdfs] graph refresh failed for %s", product.model)
                        raise_job_phase_error(f"Product {product.model}", "graph image generation", exc)
                else:
                    item_progress(f"Graph image already available for {product.model}", 10, 100)
                try:
                    item_progress(f"Rendering printed PDF for {product.model}", 20, 100)
                    generate_product_pdfs(product, progress_callback=_make_progress_window(item_progress, 20, 90))
                    item_progress(f"Printed PDF generated for {product.model}", 95, 100)
                except Exception as exc:
                    logger.exception("[maintenance:regenerate_all_product_pdfs] pdf rendering failed for %s", product.model)
                    raise_job_phase_error(f"Product {product.model}", "PDF rendering", exc)
                processed += 1
            db.commit()
        return {
            "result_message": "Product PDFs regenerated.",
            "products_processed": processed,
        }

    return serialize_maintenance_job(start_maintenance_job("regenerate_all_product_pdfs", work))


@app.post("/api/maintenance/jobs/series-pdfs/regenerate-all", response_model=MaintenanceJobResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"], summary="Start regenerating all series PDFs")
def start_regenerate_all_series_pdfs_job():
    def work(progress):
        with SessionLocal() as db:
            series_records = db.query(Series).options(joinedload(Series.product_type)).all()
            total = len(series_records)
            for index, series in enumerate(series_records, start=1):
                item_progress = _make_indexed_progress_window(progress, index, total)
                item_progress(f"Loading series {index} of {total}: {series.name}", 1, 100)
                if series.product_type and series.product_type.supports_graph and series_has_graph_capable_line_data(series):
                    try:
                        item_progress(f"Refreshing graph image for {series.name}", 10, 100)
                        generate_series_graph(series)
                    except Exception as exc:
                        logger.exception("[maintenance:regenerate_all_series_pdfs] graph generation failed for %s", series.name)
                        raise_job_phase_error(f"Series {series.name}", "graph image generation", exc)
                else:
                    item_progress(f"Series graph image not required for {series.name} because there is no graph-capable line data to plot", 10, 100)
                    logger.info(
                        "[maintenance:regenerate_all_series_pdfs] graph generation skipped for %s because there is no graph-capable line data to plot",
                        series.name,
                    )
                try:
                    item_progress(f"Generating printed series PDF for {series.name}", 20, 100)
                    generate_series_pdfs(series, progress_callback=_make_progress_window(item_progress, 20, 90))
                    item_progress(f"Printed series PDF generated for {series.name}", 95, 100)
                except Exception as exc:
                    logger.exception("[maintenance:regenerate_all_series_pdfs] pdf rendering failed for %s", series.name)
                    raise_job_phase_error(f"Series {series.name}", "PDF rendering", exc)
            db.commit()
        return {
            "result_message": "Series PDFs regenerated.",
            "products_processed": total,
        }

    return serialize_maintenance_job(start_maintenance_job("regenerate_all_series_pdfs", work))


@app.post("/api/maintenance/jobs/product-type-pdfs/regenerate-all", response_model=MaintenanceJobResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"], summary="Start regenerating all product type PDFs")
def start_regenerate_all_product_type_pdfs_job():
    def work(progress):
        with SessionLocal() as db:
            product_types = (
                db.query(ProductType)
                .options(
                    selectinload(ProductType.series).selectinload(Series.products).joinedload(Product.product_type),
                )
                .order_by(ProductType.sort_order, ProductType.id)
                .all()
            )
            total = len(product_types)
            for index, product_type in enumerate(product_types, start=1):
                item_progress = _make_indexed_progress_window(progress, index, total)
                item_progress(f"Loading product type {index} of {total}: {product_type.label}", 1, 100)
                try:
                    item_progress(
                        f"Reusing existing series PDFs and building the product type contents page for {product_type.label}",
                        20,
                        100,
                    )
                    generate_product_type_pdf(product_type, progress_callback=_make_progress_window(item_progress, 20, 90))
                    item_progress(f"Refreshing product type context for {product_type.label}", 95, 100)
                except Exception as exc:
                    logger.exception("[maintenance:regenerate_all_product_type_pdfs] pdf generation failed for %s", product_type.label)
                    raise_job_phase_error(f"Product type {product_type.label}", "PDF rendering", exc)
            progress("Building combined all-product-types catalogue", 96, 100)
            try:
                generate_all_product_types_pdf(product_types, progress_callback=_make_progress_window(progress, 96, 100))
            except Exception as exc:
                logger.exception("[maintenance:regenerate_all_product_type_pdfs] combined PDF generation failed")
                raise_job_phase_error("All product types", "PDF rendering", exc)
            db.commit()
        return {
            "result_message": "Product type PDFs and the combined all-product-types PDF regenerated.",
            "products_processed": total,
        }

    return serialize_maintenance_job(start_maintenance_job("regenerate_all_product_type_pdfs", work))


@app.post(
    "/api/maintenance/jobs/all-product-types-pdf/refresh",
    response_model=MaintenanceJobResponse,
    dependencies=[Depends(get_current_user)],
    tags=["Maintenance"],
    summary="Build the all-product-types PDF",
)
def start_refresh_all_product_types_pdf_job():
    def work(progress):
        with SessionLocal() as db:
            product_types = (
                db.query(ProductType)
                .options(selectinload(ProductType.series).selectinload(Series.products).joinedload(Product.product_type))
                .order_by(ProductType.id)
                .all()
            )
            progress("Building combined all-product-types catalogue", 5, 100)
            try:
                output_path = generate_all_product_types_pdf(
                    product_types,
                    progress_callback=_make_progress_window(progress, 5, 100),
                )
            except Exception as exc:
                logger.exception("[maintenance:refresh_all_product_types_pdf] generation failed")
                raise_job_phase_error("All product types", "PDF rendering", exc)
            return {
                "result_message": f"Generated combined catalogue: {output_path.name}.",
                "products_processed": len(product_types),
            }

    return serialize_maintenance_job(start_maintenance_job("refresh_all_product_types_pdf", work))


@app.post(
    "/api/maintenance/jobs/customer-facing-cache/refresh",
    response_model=MaintenanceJobResponse,
    dependencies=[Depends(get_current_user)],
    tags=["Maintenance"],
    summary="Refresh the customer-facing catalogue cache",
)
def start_refresh_customer_facing_cache_job():
    def work(progress):
        progress("Sending refresh request to the customer-facing site", 1, 1)
        payload = refresh_public_catalogue_cache()
        product_types = payload.get("product_types")
        series = payload.get("series")
        products = payload.get("products")
        summary_parts = ["Customer-facing catalogue cache refreshed."]
        if isinstance(product_types, int) and isinstance(series, int) and isinstance(products, int):
            summary_parts.append(f"Loaded {product_types} product types, {series} series, and {products} products.")
        return {
            "result_message": " ".join(summary_parts),
            "progress_message": "Customer-facing catalogue cache refreshed.",
            "progress_current": 1,
            "progress_total": 1,
            "progress_percent": 100.0,
        }

    return serialize_maintenance_job(start_maintenance_job("refresh_customer_facing_cache", work))


@app.delete("/api/maintenance/graph-images", response_model=GraphImageMaintenanceResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"])
def delete_all_graph_images(db: Session = Depends(get_db)):
    deleted_files = clear_all_graph_images(db)
    db.commit()
    return GraphImageMaintenanceResponse(
        message="Graph image files deleted and product graph paths cleared.",
        products_processed=0,
        files_deleted=deleted_files,
    )


@app.post("/api/maintenance/jobs/graph-images/clear", response_model=MaintenanceJobResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"], summary="Start clearing all graph images")
def start_delete_all_graph_images_job():
    def work(progress):
        with SessionLocal() as db:
            progress("Clearing stored graph image files", 0, 1)
            deleted_files = clear_all_graph_images(db)
            db.commit()
        return {
            "result_message": "Graph image files deleted and product graph paths cleared.",
            "files_deleted": deleted_files,
            "progress_current": 1,
            "progress_total": 1,
            "progress_percent": 100.0,
        }

    return serialize_maintenance_job(start_maintenance_job("clear_all_graph_images", work))


@app.get("/api/maintenance/backups/database/download", dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Download a DB data backup")
def download_database_backup_bundle():
    try:
        archive_path = create_database_backup_bundle()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unable to create DB data backup: {exc}") from exc

    return FileResponse(
        archive_path,
        media_type="application/zip",
        filename=archive_path.name,
    )


@app.get("/api/maintenance/backups/download", dependencies=[Depends(require_admin_user)], tags=["Maintenance"], include_in_schema=False)
def download_backup_bundle():
    return download_database_backup_bundle()


@app.post("/api/maintenance/jobs/backups/database/create", response_model=MaintenanceJobResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Start creating a DB data backup")
def start_database_backup_bundle_job():
    def work(progress):
        archive_path = create_database_backup_bundle(progress)
        return {
            "result_message": f"DB data backup created: {archive_path.name}",
            "result_download_url": f"/api/maintenance/jobs/{job['id']}/download",
            "result_file_path": str(archive_path),
            "progress_current": 1,
            "progress_total": 1,
            "progress_percent": 100.0,
        }

    job = start_maintenance_job("create_database_backup_bundle", work)
    return serialize_maintenance_job(job)


@app.post("/api/maintenance/jobs/backups/create", response_model=MaintenanceJobResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], include_in_schema=False)
def start_backup_bundle_job():
    return start_database_backup_bundle_job()


@app.get("/api/maintenance/media/download", dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Download a media data backup")
def download_media_backup_bundle():
    try:
        archive_path = create_data_backup_bundle()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unable to create media data backup: {exc}") from exc

    return FileResponse(
        archive_path,
        media_type="application/zip",
        filename=archive_path.name,
    )


@app.get("/api/maintenance/data/download", dependencies=[Depends(require_admin_user)], tags=["Maintenance"], include_in_schema=False)
def download_data_backup_bundle():
    return download_media_backup_bundle()


@app.post("/api/maintenance/jobs/media/create", response_model=MaintenanceJobResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Start creating a media data backup")
def start_media_backup_bundle_job():
    def work(progress):
        archive_path = create_data_backup_bundle(progress)
        return {
            "result_message": f"Media data backup created: {archive_path.name}",
            "result_download_url": f"/api/maintenance/jobs/{job['id']}/download",
            "result_file_path": str(archive_path),
            "progress_current": 1,
            "progress_total": 1,
            "progress_percent": 100.0,
        }

    job = start_maintenance_job("create_media_data_backup_bundle", work)
    return serialize_maintenance_job(job)


@app.post("/api/maintenance/jobs/data/create", response_model=MaintenanceJobResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], include_in_schema=False)
def start_data_backup_bundle_job():
    return start_media_backup_bundle_job()


@app.post("/api/maintenance/backups/db/restore", dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Restore a DB data backup")
async def restore_db_backup_bundle_endpoint(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Please upload a .zip DB data backup.")

    try:
        archive_bytes = await file.read()
        restore_backup_bundle(archive_bytes)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unable to restore DB data backup: {exc}") from exc

    return {"message": "DB data backup restored successfully."}


@app.post("/api/maintenance/backups/restore", dependencies=[Depends(require_admin_user)], tags=["Maintenance"], include_in_schema=False)
async def restore_backup_bundle_endpoint(file: UploadFile = File(...)):
    return await restore_db_backup_bundle_endpoint(file)


@app.post("/api/maintenance/backups/media/restore", dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Restore a media data backup")
async def restore_media_backup_bundle_endpoint(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Please upload a .zip media data backup.")

    try:
        archive_bytes = await file.read()
        restore_media_backup_bundle(archive_bytes)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unable to restore media data backup: {exc}") from exc

    return {"message": "Media data backup restored successfully."}


@app.post("/api/maintenance/backups/media/restore-old", dependencies=[Depends(require_admin_user)], tags=["Maintenance"], include_in_schema=False)
async def restore_media_backup_bundle_endpoint_old(file: UploadFile = File(...)):
    return await restore_media_backup_bundle_endpoint(file)


@app.post("/api/maintenance/jobs/backups/db/restore", response_model=MaintenanceJobResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Start restoring a DB data backup")
async def start_restore_backup_bundle_job(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Please upload a .zip DB data backup.")

    archive_bytes = await file.read()

    def work(progress):
        restore_backup_bundle(archive_bytes, progress)
        run_post_restore_schema_prep(progress)
        return {
            "result_message": "DB data backup restored successfully.",
            "progress_current": 1,
            "progress_total": 1,
            "progress_percent": 100.0,
        }

    return serialize_maintenance_job(start_maintenance_job("restore_db_data_backup_bundle", work))


@app.post("/api/maintenance/jobs/backups/restore", response_model=MaintenanceJobResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], include_in_schema=False)
async def start_restore_backup_bundle_job_old(file: UploadFile = File(...)):
    return await start_restore_backup_bundle_job(file)


@app.post("/api/maintenance/jobs/backups/media/restore", response_model=MaintenanceJobResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Start restoring a media data backup")
async def start_restore_media_backup_bundle_job(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Please upload a .zip media data backup.")

    archive_bytes = await file.read()

    def work(progress):
        restore_media_backup_bundle(archive_bytes, progress)
        return {
            "result_message": "Media data backup restored successfully.",
            "progress_current": 1,
            "progress_total": 1,
            "progress_percent": 100.0,
        }

    return serialize_maintenance_job(start_maintenance_job("restore_media_data_backup_bundle", work))


@app.post("/api/maintenance/jobs/backups/media/restore-old", response_model=MaintenanceJobResponse, dependencies=[Depends(require_admin_user)], tags=["Maintenance"], include_in_schema=False)
async def start_restore_media_backup_bundle_job_old(file: UploadFile = File(...)):
    return await start_restore_media_backup_bundle_job(file)


@app.get("/api/maintenance/jobs/{job_id}", response_model=MaintenanceJobResponse, dependencies=[Depends(get_current_user)], tags=["Maintenance"], summary="Get maintenance job status")
def get_maintenance_job(job_id: str):
    return serialize_maintenance_job(get_maintenance_job_or_404(job_id))


@app.get("/api/setup/logs/recent", response_model=list[SetupLogEntryResponse], dependencies=[Depends(require_admin_user)], tags=["Setup"], summary="Get recent setup logs")
def get_setup_logs_recent(limit: int = Query(200, ge=1, le=500)):
    return get_recent_setup_log_entries(limit)


@app.get("/api/public-access/logs/recent", response_model=list[PublicAccessLogEntryResponse], dependencies=[Depends(require_admin_user)], tags=["Setup"], summary="Get recent public-access logs")
def get_public_access_logs_recent(
    limit: int = Query(200, ge=1, le=500),
    site: str | None = Query(default=None),
    route_group: str | None = Query(default=None),
):
    return get_recent_public_access_log_entries(limit, site=site, route_group=route_group)


@app.get("/api/customer-facing/logs/recent", response_model=list[PublicAccessLogEntryResponse], dependencies=[Depends(require_admin_user)], tags=["Setup"], summary="Get recent customer-facing logs")
async def get_customer_facing_logs_recent(
    limit: int = Query(200, ge=1, le=500),
    public_only: bool = Query(default=True),
):
    return await fetch_customer_facing_recent_logs(limit=limit, public_only=public_only)


@app.get("/api/setup/logs/stream", dependencies=[Depends(require_admin_user)], tags=["Setup"], summary="Stream setup logs")
async def stream_setup_logs(request: Request, after_id: int = Query(0, ge=0)):
    async def event_stream():
        last_id = after_id
        heartbeat_ticks = 0

        for entry in get_setup_log_entries_after_id(after_id):
            last_id = entry.id
            yield setup_log_sse_payload(entry)

        while True:
            if await request.is_disconnected():
                break

            new_entries = get_setup_log_entries_after_id(last_id)
            if new_entries:
                for entry in new_entries:
                    last_id = entry.id
                    yield setup_log_sse_payload(entry)
                heartbeat_ticks = 0
            else:
                heartbeat_ticks += 1
                if heartbeat_ticks >= 15:
                    heartbeat_ticks = 0
                    yield ": keep-alive\n\n"

            await asyncio.sleep(1)

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.get("/api/maintenance/jobs/{job_id}/download", dependencies=[Depends(require_admin_user)], tags=["Maintenance"], summary="Download a completed maintenance archive")
def download_maintenance_job_file(job_id: str):
    job = get_maintenance_job_or_404(job_id)
    if job.get("status") != "completed":
        raise HTTPException(status_code=409, detail="Maintenance job is not complete yet.")
    file_path = job.get("result_file_path")
    if not file_path:
        raise HTTPException(status_code=404, detail="This maintenance job does not have a downloadable file.")
    archive_path = Path(file_path)
    if not archive_path.is_file():
        raise HTTPException(status_code=404, detail="The generated file is no longer available.")
    return FileResponse(
        archive_path,
        media_type="application/zip",
        filename=archive_path.name,
    )


@app.get("/api/fans/{product_id}/product-images", response_model=list[ProductImageResponse], dependencies=[Depends(get_current_user)], tags=["Product Images"], include_in_schema=False)
@app.get("/api/products/{product_id}/product-images", response_model=list[ProductImageResponse], dependencies=[Depends(get_current_user)], tags=["Product Images"])
def get_product_images(product_id: int, db: Session = Depends(get_db)):
    product = require_product(db, product_id)
    return sorted(product.product_images, key=lambda image: (image.sort_order, image.id))


@app.post("/api/fans/{product_id}/product-images", response_model=list[ProductImageResponse], dependencies=[Depends(get_current_user)], tags=["Product Images"], include_in_schema=False)
@app.post("/api/products/{product_id}/product-images", response_model=list[ProductImageResponse], dependencies=[Depends(get_current_user)], tags=["Product Images"])
async def upload_product_images(
    product_id: int,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    product = require_product(db, product_id)
    if not files:
        raise HTTPException(400, "No files provided")

    next_order = len(product.product_images)
    for upload in files:
        suffix = os.path.splitext(upload.filename or "")[1].lower() or ".jpg"
        image = ProductImage(
            product_id=product_id,
            file_name=f"upload_{product_id}_{next_order}{suffix}",
            sort_order=next_order,
        )
        db.add(image)
        db.flush()
        contents = await upload.read()
        target_path = product_image_target_path(product_id, image.file_name)
        target_path.parent.mkdir(parents=True, exist_ok=True)
        with open(target_path, "wb") as output:
            output.write(contents)
        next_order += 1

    sync_product_image_files(product)
    db.commit()
    db.refresh(product)
    notify_public_catalogue_cache_refresh()
    return sorted(product.product_images, key=lambda image: (image.sort_order, image.id))


def _bulk_image_upload_name(upload: UploadFile, fallback_prefix: str, target_id: int, index: int) -> str:
    file_name = Path(upload.filename or "").name.strip()
    if not file_name:
        file_name = f"{fallback_prefix}_{target_id}_{index}.jpg"
    if not Path(file_name).suffix:
        file_name = f"{file_name}.jpg"
    return file_name


async def _bulk_upload_images(target_kind: str, target_id: int, files: list[UploadFile], db: Session) -> BulkImageImportResponse:
    if not files:
        raise HTTPException(400, "No files provided")

    if target_kind == "product":
        product = require_product(db, target_id)
        existing_images = list(product.product_images)
        next_order = len(existing_images)
        file_names: list[str] = []
        overwritten_file_names: list[str] = []
        for index, upload in enumerate(files):
            file_name = _bulk_image_upload_name(upload, "product_image", target_id, index)
            contents = await upload.read()
            existing_image = next((image for image in existing_images if image.file_name == file_name), None)
            if existing_image is None:
                existing_image = ProductImage(
                    product_id=target_id,
                    file_name=file_name,
                    sort_order=next_order,
                )
                db.add(existing_image)
                db.flush()
                existing_images.append(existing_image)
                next_order += 1
                file_names.append(file_name)
            else:
                overwritten_file_names.append(file_name)
            target_path = product_image_target_path(target_id, file_name)
            target_path.parent.mkdir(parents=True, exist_ok=True)
            target_path.write_bytes(contents)
        db.commit()
        db.refresh(product)
        notify_public_catalogue_cache_refresh()
        return BulkImageImportResponse(
            target_kind="product",
            target_id=target_id,
            file_names=file_names,
            overwritten_file_names=overwritten_file_names,
            image_count=len(product.product_images),
        )

    series = db.get(Series, target_id)
    if not series:
        raise HTTPException(404, "Series not found")
    existing_images = list(series.series_images)
    next_order = len(existing_images)
    file_names = []
    overwritten_file_names = []
    for index, upload in enumerate(files):
        file_name = _bulk_image_upload_name(upload, "series_image", target_id, index)
        contents = await upload.read()
        existing_image = next((image for image in existing_images if image.file_name == file_name), None)
        if existing_image is None:
            existing_image = SeriesImage(
                series_id=target_id,
                file_name=file_name,
                sort_order=next_order,
            )
            db.add(existing_image)
            db.flush()
            existing_images.append(existing_image)
            next_order += 1
            file_names.append(file_name)
        else:
            overwritten_file_names.append(file_name)
        target_path = series_image_target_path(target_id, file_name)
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_bytes(contents)
    db.commit()
    db.refresh(series)
    notify_public_catalogue_cache_refresh()
    return BulkImageImportResponse(
        target_kind="series",
        target_id=target_id,
        file_names=file_names,
        overwritten_file_names=overwritten_file_names,
        image_count=len(series.series_images),
    )


@app.post("/api/products/{product_id}/product-images/bulk", response_model=BulkImageImportResponse, dependencies=[Depends(get_current_user)], tags=["Product Images"])
async def bulk_upload_product_images(
    product_id: int,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    return await _bulk_upload_images("product", product_id, files, db)


@app.post("/api/series/{series_id}/series-images/bulk", response_model=BulkImageImportResponse, dependencies=[Depends(get_current_user)], tags=["Series Images"])
async def bulk_upload_series_images(
    series_id: int,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    return await _bulk_upload_images("series", series_id, files, db)


@app.post("/api/fans/{product_id}/product-images/reorder", response_model=list[ProductImageResponse], dependencies=[Depends(get_current_user)], tags=["Product Images"], include_in_schema=False)
@app.post("/api/products/{product_id}/product-images/reorder", response_model=list[ProductImageResponse], dependencies=[Depends(get_current_user)], tags=["Product Images"])
def reorder_product_images(product_id: int, body: ProductImageReorder, db: Session = Depends(get_db)):
    product = require_product(db, product_id)
    images_by_id = {image.id: image for image in product.product_images}
    if set(body.image_ids) != set(images_by_id.keys()):
        raise HTTPException(400, "Image order must include every existing image exactly once")

    for index, image_id in enumerate(body.image_ids):
        images_by_id[image_id].sort_order = index

    sync_product_image_files(product)
    db.commit()
    db.refresh(product)
    notify_public_catalogue_cache_refresh()
    return sorted(product.product_images, key=lambda image: (image.sort_order, image.id))


@app.delete("/api/fans/{product_id}/product-images/{image_id}", dependencies=[Depends(get_current_user)], tags=["Product Images"], include_in_schema=False)
@app.delete("/api/products/{product_id}/product-images/{image_id}", dependencies=[Depends(get_current_user)], tags=["Product Images"])
def delete_product_image(product_id: int, image_id: int, db: Session = Depends(get_db)):
    product = require_product(db, product_id)
    image = db.get(ProductImage, image_id)
    if not image or image.product_id != product_id:
        raise HTTPException(404, "Product image not found")

    delete_product_image_file(image)
    db.delete(image)
    db.flush()
    sync_product_image_files(product)
    db.commit()
    notify_public_catalogue_cache_refresh()
    return {"deleted": image_id}


@app.get("/api/series/{series_id}/series-images", response_model=list[SeriesImageResponse], dependencies=[Depends(get_current_user)], tags=["Series Images"])
def get_series_images(series_id: int, db: Session = Depends(get_db)):
    series = db.get(Series, series_id)
    if not series:
        raise HTTPException(404, "Series not found")
    return sorted(series.series_images, key=lambda image: (image.sort_order, image.id))


@app.post("/api/series/{series_id}/series-images", response_model=list[SeriesImageResponse], dependencies=[Depends(get_current_user)], tags=["Series Images"])
async def upload_series_images(
    series_id: int,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    series = db.get(Series, series_id)
    if not series:
        raise HTTPException(404, "Series not found")
    if not files:
        raise HTTPException(400, "No files provided")

    next_order = len(series.series_images)
    for upload in files:
        suffix = os.path.splitext(upload.filename or "")[1].lower() or ".jpg"
        image = SeriesImage(
            series_id=series_id,
            file_name=f"upload_{series_id}_{next_order}{suffix}",
            sort_order=next_order,
        )
        db.add(image)
        db.flush()
        contents = await upload.read()
        target_path = series_image_target_path(series_id, image.file_name)
        target_path.parent.mkdir(parents=True, exist_ok=True)
        with open(target_path, "wb") as output:
            output.write(contents)
        next_order += 1

    sync_series_image_files(series)
    db.commit()
    db.refresh(series)
    notify_public_catalogue_cache_refresh()
    return sorted(series.series_images, key=lambda image: (image.sort_order, image.id))


@app.post("/api/series/{series_id}/series-images/reorder", response_model=list[SeriesImageResponse], dependencies=[Depends(get_current_user)], tags=["Series Images"])
def reorder_series_images(series_id: int, body: SeriesImageReorder, db: Session = Depends(get_db)):
    series = db.get(Series, series_id)
    if not series:
        raise HTTPException(404, "Series not found")
    images_by_id = {image.id: image for image in series.series_images}
    if set(body.image_ids) != set(images_by_id.keys()):
        raise HTTPException(400, "Image order must include every existing image exactly once")

    for index, image_id in enumerate(body.image_ids):
        images_by_id[image_id].sort_order = index

    sync_series_image_files(series)
    db.commit()
    db.refresh(series)
    notify_public_catalogue_cache_refresh()
    return sorted(series.series_images, key=lambda image: (image.sort_order, image.id))


@app.delete("/api/series/{series_id}/series-images/{image_id}", dependencies=[Depends(get_current_user)], tags=["Series Images"])
def delete_series_image(series_id: int, image_id: int, db: Session = Depends(get_db)):
    series = db.get(Series, series_id)
    if not series:
        raise HTTPException(404, "Series not found")
    image = db.get(SeriesImage, image_id)
    if not image or image.series_id != series_id:
        raise HTTPException(404, "Series image not found")

    delete_series_image_file(image)
    db.delete(image)
    db.flush()
    sync_series_image_files(series)
    db.commit()
    notify_public_catalogue_cache_refresh()
    return {"deleted": image_id}


ASSOCIATED_DOCUMENT_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".csv", ".txt", ".rtf",
    ".dwg", ".dxf", ".png", ".jpg", ".jpeg", ".webp",
}
ASSOCIATED_DOCUMENT_MAX_BYTES = 25 * 1024 * 1024


def _associated_document_owner(db: Session, owner_type: str, owner_id: int):
    if owner_type == "product":
        return db.get(Product, owner_id)
    if owner_type == "series":
        return db.get(Series, owner_id)
    if owner_type == "product_type":
        return db.get(ProductType, owner_id)
    raise HTTPException(400, "Unsupported document owner type")


def _associated_document_query_filter(document: AssociatedDocument, owner_type: str, owner_id: int) -> bool:
    return (
        document.owner_type == owner_type
        and document.owner_id == owner_id
    )


async def _upload_associated_documents(owner_type: str, owner_id: int, files: list[UploadFile], db: Session):
    owner = _associated_document_owner(db, owner_type, owner_id)
    if not owner:
        raise HTTPException(404, f"{owner_type.replace('_', ' ').title()} not found")
    if not files:
        raise HTTPException(400, "No files provided")

    existing = list(owner.associated_documents)
    next_order = len(existing)
    for upload in files:
        original_name = Path(upload.filename or "document").name
        suffix = Path(original_name).suffix.lower()
        if suffix not in ASSOCIATED_DOCUMENT_EXTENSIONS:
            raise HTTPException(400, f"Unsupported document type: {suffix or 'unknown'}")
        contents = await upload.read(ASSOCIATED_DOCUMENT_MAX_BYTES + 1)
        if len(contents) > ASSOCIATED_DOCUMENT_MAX_BYTES:
            raise HTTPException(413, f"{original_name} exceeds the 25 MB limit")
        stored_name = f"{uuid4().hex}{suffix}"
        document = AssociatedDocument(
            owner_type=owner_type,
            original_file_name=original_name,
            file_name=stored_name,
            mime_type=upload.content_type or "application/octet-stream",
            sort_order=next_order,
        )
        setattr(document, f"{owner_type if owner_type != 'product_type' else 'product_type'}_id", owner_id)
        db.add(document)
        target_path = associated_document_directory(owner_type, owner_id) / stored_name
        target_path.write_bytes(contents)
        next_order += 1

    db.commit()
    db.refresh(owner)
    notify_public_catalogue_cache_refresh()
    return sorted(owner.associated_documents, key=lambda item: (item.sort_order, item.id))


@app.get("/api/products/{product_id}/documents", response_model=list[AssociatedDocumentResponse], dependencies=[Depends(get_current_user)], tags=["Associated Documents"])
def get_product_documents(product_id: int, db: Session = Depends(get_db)):
    owner = _associated_document_owner(db, "product", product_id)
    if not owner:
        raise HTTPException(404, "Product not found")
    return sorted(owner.associated_documents, key=lambda item: (item.sort_order, item.id))


@app.post("/api/products/{product_id}/documents", response_model=list[AssociatedDocumentResponse], dependencies=[Depends(get_current_user)], tags=["Associated Documents"])
async def upload_product_documents(product_id: int, files: list[UploadFile] = File(...), db: Session = Depends(get_db)):
    return await _upload_associated_documents("product", product_id, files, db)


@app.get("/api/series/{series_id}/documents", response_model=list[AssociatedDocumentResponse], dependencies=[Depends(get_current_user)], tags=["Associated Documents"])
def get_series_documents(series_id: int, db: Session = Depends(get_db)):
    owner = _associated_document_owner(db, "series", series_id)
    if not owner:
        raise HTTPException(404, "Series not found")
    return sorted(owner.associated_documents, key=lambda item: (item.sort_order, item.id))


@app.post("/api/series/{series_id}/documents", response_model=list[AssociatedDocumentResponse], dependencies=[Depends(get_current_user)], tags=["Associated Documents"])
async def upload_series_documents(series_id: int, files: list[UploadFile] = File(...), db: Session = Depends(get_db)):
    return await _upload_associated_documents("series", series_id, files, db)


@app.get("/api/product-types/{product_type_id}/documents", response_model=list[AssociatedDocumentResponse], dependencies=[Depends(get_current_user)], tags=["Associated Documents"])
def get_product_type_documents(product_type_id: int, db: Session = Depends(get_db)):
    owner = _associated_document_owner(db, "product_type", product_type_id)
    if not owner:
        raise HTTPException(404, "Product type not found")
    return sorted(owner.associated_documents, key=lambda item: (item.sort_order, item.id))


@app.post("/api/product-types/{product_type_id}/documents", response_model=list[AssociatedDocumentResponse], dependencies=[Depends(get_current_user)], tags=["Associated Documents"])
async def upload_product_type_documents(product_type_id: int, files: list[UploadFile] = File(...), db: Session = Depends(get_db)):
    return await _upload_associated_documents("product_type", product_type_id, files, db)


@app.delete("/api/{owner_type}/{owner_id}/documents/{document_id}", response_model=dict, dependencies=[Depends(get_current_user)], tags=["Associated Documents"])
def delete_associated_document(owner_type: str, owner_id: int, document_id: int, db: Session = Depends(get_db)):
    owner = _associated_document_owner(db, owner_type, owner_id)
    if not owner:
        raise HTTPException(404, f"{owner_type.replace('_', ' ').title()} not found")
    document = db.get(AssociatedDocument, document_id)
    if not document or not _associated_document_query_filter(document, owner_type, owner_id):
        raise HTTPException(404, "Document not found")
    remove_file(associated_document_path(owner_type, owner_id, document.file_name))
    db.delete(document)
    db.commit()
    notify_public_catalogue_cache_refresh()
    return {"deleted": document_id}


@app.get("/api/media/product_images/{product_id}/{file_name:path}", dependencies=[Depends(get_current_user)], tags=["Media"])
def serve_product_image(product_id: int, file_name: str):
    file_path = product_image_path(product_id, file_name)
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Product image not found")
    return FileResponse(file_path)


@app.get("/api/media/series_images/{series_id}/{file_name:path}", dependencies=[Depends(get_current_user)], tags=["Media"])
def serve_series_image(series_id: int, file_name: str):
    file_path = series_image_path(series_id, file_name)
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Series image not found")
    return FileResponse(file_path)


@app.get("/api/media/product_graphs/{file_name}", dependencies=[Depends(get_current_user)], tags=["Media"])
def serve_product_graph(file_name: str):
    file_path = PRODUCT_GRAPHS_DIR / file_name
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Product graph not found")
    return FileResponse(file_path)


@app.get("/api/media/product_pdfs/{file_name}", dependencies=[Depends(get_current_user)], tags=["Media"])
def serve_product_pdf(file_name: str):
    file_path = PRODUCT_PDFS_DIR / file_name
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Product PDF not found")
    return pdf_file_response(file_path)


@app.get("/api/media/product_type_pdfs/{file_name}", dependencies=[Depends(get_current_user)], tags=["Media"])
def serve_product_type_pdf(file_name: str):
    file_path = PRODUCT_TYPE_PDFS_DIR / file_name
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Product type PDF not found")
    return pdf_file_response(file_path)


@app.get(
    "/api/media/all-product-types-pdf",
    dependencies=[Depends(get_current_user)],
    tags=["Media"],
    summary="Get the combined all-product-types PDF",
)
def serve_all_product_types_pdf():
    file_path = all_product_types_pdf_path()
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Combined all-product-types PDF not found")
    return pdf_file_response(file_path)


@app.get("/api/media/series_graphs/{file_name}", dependencies=[Depends(get_current_user)], tags=["Media"])
def serve_series_graph(file_name: str):
    file_path = SERIES_GRAPHS_DIR / file_name
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Series graph not found")
    return FileResponse(file_path)


@app.get("/api/media/series_pdfs/{file_name}", dependencies=[Depends(get_current_user)], tags=["Media"])
def serve_series_pdf(file_name: str):
    file_path = SERIES_PDFS_DIR / file_name
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Series PDF not found")
    return pdf_file_response(file_path)


@app.get(
    "/api/public/media/product_images/{product_id}/{file_name:path}",
    tags=["Public Media"],
    summary="Get a public customer product image",
    description="Public product image endpoint intended for rendered customer-facing pages.",
)
def serve_cms_product_image(product_id: int, file_name: str):
    file_path = product_image_path(product_id, file_name)
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Product image not found")
    return FileResponse(file_path)


@app.get(
    "/api/public/media/series_images/{series_id}/{file_name:path}",
    tags=["Public Media"],
    summary="Get a public customer series image",
    description="Public series image endpoint intended for rendered customer-facing pages.",
)
def serve_cms_series_image(series_id: int, file_name: str):
    file_path = series_image_path(series_id, file_name)
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Series image not found")
    return FileResponse(file_path)


@app.get(
    "/api/public/media/product_graphs/{file_name}",
    tags=["Public Media"],
    summary="Get a public customer product graph image",
    description="Public graph image endpoint intended for rendered customer-facing pages and downloads.",
)
def serve_cms_product_graph(file_name: str):
    file_path = PRODUCT_GRAPHS_DIR / file_name
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Product graph not found")
    return FileResponse(file_path)


@app.get(
    "/api/public/media/product_pdfs/{file_name}",
    tags=["Public Media"],
    summary="Get a public customer product PDF",
    description="Public product PDF endpoint intended for customer-facing downloads.",
)
def serve_cms_product_pdf(file_name: str):
    file_path = PRODUCT_PDFS_DIR / file_name
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Product PDF not found")
    return pdf_file_response(file_path)


@app.get(
    "/api/public/media/product_type_pdfs/{file_name}",
    tags=["Public Media"],
    summary="Get a public customer product type PDF",
    description="Public product type PDF endpoint intended for customer-facing downloads.",
)
def serve_cms_product_type_pdf(file_name: str):
    file_path = PRODUCT_TYPE_PDFS_DIR / file_name
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Product type PDF not found")
    return pdf_file_response(file_path)


@app.get(
    "/api/public/media/all-product-types-pdf",
    tags=["Public Media"],
    summary="Get the public combined all-product-types PDF",
    description="Public combined product catalogue PDF endpoint intended for customer-facing downloads.",
)
def serve_cms_all_product_types_pdf():
    file_path = all_product_types_pdf_path()
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Combined all-product-types PDF not found")
    return pdf_file_response(file_path)


@app.get(
    "/api/public/catalogue/all-product-types-pdf",
    tags=["Public Catalog"],
    summary="Get combined catalogue PDF metadata",
)
def get_public_all_product_types_pdf_metadata(response: Response):
    file_path = all_product_types_pdf_path()
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return {
        "available": file_path.is_file(),
        "url": all_product_types_pdf_public_url(file_path),
        "size_bytes": file_path.stat().st_size if file_path.is_file() else None,
    }


@app.get(
    "/api/public/media/series_graphs/{file_name}",
    tags=["Public Media"],
    summary="Get a public customer series graph image",
    description="Public series graph endpoint intended for rendered customer-facing pages and downloads.",
)
def serve_cms_series_graph(file_name: str):
    file_path = SERIES_GRAPHS_DIR / file_name
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Series graph not found")
    return FileResponse(file_path)


@app.get(
    "/api/public/media/series_pdfs/{file_name}",
    tags=["Public Media"],
    summary="Get a public customer series PDF",
    description="Public series PDF endpoint intended for customer-facing downloads.",
)
def serve_cms_series_pdf(file_name: str):
    file_path = SERIES_PDFS_DIR / file_name
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Series PDF not found")
    return pdf_file_response(file_path)


@app.get("/api/media/associated_documents/{owner_type}/{owner_id}/{file_name:path}", dependencies=[Depends(get_current_user)], tags=["Media"])
def serve_associated_document(owner_type: str, owner_id: int, file_name: str, db: Session = Depends(get_db)):
    file_path = associated_document_path(owner_type, owner_id, file_name)
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Associated document not found")
    document = db.query(AssociatedDocument).filter(
        AssociatedDocument.file_name == file_name,
        AssociatedDocument.owner_type == owner_type,
    ).first()
    return FileResponse(file_path, media_type=None, filename=document.original_file_name if document else file_path.name)


@app.get("/api/public/media/associated_documents/{owner_type}/{owner_id}/{file_name:path}", tags=["Public Media"])
def serve_public_associated_document(owner_type: str, owner_id: int, file_name: str, db: Session = Depends(get_db)):
    file_path = associated_document_path(owner_type, owner_id, file_name)
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Associated document not found")
    document = db.query(AssociatedDocument).filter(
        AssociatedDocument.file_name == file_name,
        AssociatedDocument.owner_type == owner_type,
    ).first()
    return FileResponse(file_path, media_type=None, filename=document.original_file_name if document else file_path.name)


# --- Frontend static serving ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(BASE_DIR)
FRONTEND_BUILD_DIR = os.path.join(PROJECT_ROOT, "frontend", "build")

if os.path.isdir(FRONTEND_BUILD_DIR):
    app.mount("/_app", StaticFiles(directory=os.path.join(FRONTEND_BUILD_DIR, "_app")), name="frontend_app")

    @app.get("/")
    def serve_index():
        return FileResponse(os.path.join(FRONTEND_BUILD_DIR, "index.html"))

    @app.get("/{full_path:path}")
    def serve_spa(full_path: str):
        if full_path.startswith("api/"):
            raise HTTPException(status_code=404, detail="Not found")
        requested_path = os.path.join(FRONTEND_BUILD_DIR, full_path)
        if os.path.isfile(requested_path):
            return FileResponse(requested_path)
        return FileResponse(os.path.join(FRONTEND_BUILD_DIR, "index.html"))
